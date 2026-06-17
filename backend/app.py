"""
ScreenPlant Pro — Flask Backend
ID Card Generation System for Tariq Printing Solutions

[ENHANCED] Added analytics stats, daily trends, school health, paginated submissions,
partial submission editing, photo review state, template duplication, analytics CSV
export, and runtime admin password change support.
"""

import os, json, uuid, zipfile, io, base64, binascii, csv, re, tempfile, difflib, shutil, sqlite3, threading, time, hmac, hashlib, logging, secrets, urllib.request, queue
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, date
from functools import wraps
from contextlib import closing
from flask import Flask, request, jsonify, send_file, send_from_directory, session, g, Response, stream_with_context
from flask_cors import CORS
from werkzeug.utils import secure_filename
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_OK = True
except ImportError:
    openpyxl = None
    Font = Alignment = PatternFill = Border = Side = None
    OPENPYXL_OK = False
try:
    import fitz  # PyMuPDF — for reading CorelDraw PDFs
    PYMUPDF_OK = True
except ImportError:
    PYMUPDF_OK = False
try:
    import qrcode
    QRCODE_OK = True
except ImportError:
    qrcode = None
    QRCODE_OK = False

try:
    import numpy as np
    NUMPY_OK = True
except ImportError:
    np = None
    NUMPY_OK = False

try:
    import boto3
    from botocore.exceptions import BotoCoreError, ClientError
    BOTO3_OK = True
except ImportError:
    boto3 = None
    BotoCoreError = ClientError = Exception
    BOTO3_OK = False

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR     = os.path.abspath(os.path.join(BASE_DIR, '..'))
# Support both layouts:
#   backend/app.py with index.html in the project root
#   backend/app.py with frontend/index.html
#   app.py next to index.html
_root_frontend = PROJECT_DIR if os.path.exists(os.path.join(PROJECT_DIR, 'index.html')) else ''
_frontend_dir = os.path.join(PROJECT_DIR, 'frontend')
if os.environ.get('SCREENPLANT_FRONTEND_DIR'):
    FRONTEND_DIR = os.path.abspath(os.environ['SCREENPLANT_FRONTEND_DIR'])
elif _root_frontend:
    FRONTEND_DIR = _root_frontend
elif os.path.isdir(_frontend_dir):
    FRONTEND_DIR = os.path.abspath(_frontend_dir)
else:
    FRONTEND_DIR = BASE_DIR

app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path='')

_secret_key = os.environ.get('SCREENPLANT_SECRET_KEY', '')
_is_https = os.environ.get('SCREENPLANT_HTTPS', '').lower() in ('1', 'true', 'yes')

if not _secret_key or _secret_key == 'change-this-secret-before-production':
    _secret_key = hashlib.sha256(os.urandom(32)).hexdigest()
    logger_tmp = logging.getLogger('screenplant.startup')
    logger_tmp.warning(
        "⚠️  SCREENPLANT_SECRET_KEY not set — using a random key. "
        "All sessions will be invalidated on every restart. "
        "Set a stable SCREENPLANT_SECRET_KEY before production deployment."
    )

app.config.update(
    SECRET_KEY=_secret_key,
    MAX_CONTENT_LENGTH=int(os.environ.get('SCREENPLANT_MAX_UPLOAD_MB', '25')) * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_is_https,   # Set SCREENPLANT_HTTPS=1 in production behind HTTPS
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)
CORS(app, origins=os.environ.get('SCREENPLANT_CORS_ORIGINS', '*').split(','), supports_credentials=True)

logging.basicConfig(
    level=os.environ.get('SCREENPLANT_LOG_LEVEL', 'INFO'),
    format='%(asctime)s %(levelname)s %(name)s: %(message)s'
)
logger = logging.getLogger('screenplant')

# ── Startup safety warnings ────────────────────────────────────────────────────
_default_admin_pass = 'admin123'
_configured_pass = os.environ.get('SCREENPLANT_ADMIN_PASSWORD', _default_admin_pass)
if _configured_pass == _default_admin_pass:
    logger.warning(
        "⚠️  SCREENPLANT_ADMIN_PASSWORD is using the default 'admin123'. "
        "Set a strong password via environment variable before any public deployment."
    )

# Warn when running on a cloud platform with an ephemeral filesystem.
# On Railway/Render/Heroku, local files (photos, templates, backups) are lost
# on every redeploy. Set up Cloudflare R2 / Backblaze B2 for persistent storage.
_on_railway = bool(os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('RAILWAY_PROJECT_ID'))
_on_render  = bool(os.environ.get('RENDER'))
if _on_railway or _on_render:
    logger.warning(
        "⚠️  Running on an ephemeral cloud filesystem. "
        "Uploaded photos, templates, and local backups WILL BE LOST on redeploy. "
        "Set SCREENPLANT_STORAGE_BACKEND=r2 and configure R2 credentials for persistent file storage."
    )

# ── Cloudflare R2 Storage ──────────────────────────────────────────────────────
# Set these environment variables to enable R2:
#   SCREENPLANT_STORAGE_BACKEND=r2
#   R2_ACCOUNT_ID=your_cloudflare_account_id
#   R2_ACCESS_KEY_ID=your_r2_access_key
#   R2_SECRET_ACCESS_KEY=your_r2_secret_key
#   R2_BUCKET_NAME=your_bucket_name
#   R2_PUBLIC_URL=https://pub-xxxx.r2.dev  (optional: public bucket URL for direct photo access)

_USE_R2 = os.environ.get('SCREENPLANT_STORAGE_BACKEND', '').lower() == 'r2'
_R2_ACCOUNT_ID       = os.environ.get('R2_ACCOUNT_ID', '')
_R2_ACCESS_KEY_ID    = os.environ.get('R2_ACCESS_KEY_ID', '')
_R2_SECRET_ACCESS_KEY = os.environ.get('R2_SECRET_ACCESS_KEY', '')
_R2_BUCKET_NAME      = os.environ.get('R2_BUCKET_NAME', 'screenplant-photos')
_R2_PUBLIC_URL       = os.environ.get('R2_PUBLIC_URL', '').rstrip('/')

_r2_client = None

def _get_r2_client():
    global _r2_client
    if _r2_client is not None:
        return _r2_client
    if not BOTO3_OK:
        raise RuntimeError("R2 requires boto3. Run: pip install boto3")
    if not _R2_ACCOUNT_ID or not _R2_ACCESS_KEY_ID or not _R2_SECRET_ACCESS_KEY:
        raise RuntimeError(
            "R2 credentials not set. Set R2_ACCOUNT_ID, R2_ACCESS_KEY_ID, R2_SECRET_ACCESS_KEY."
        )
    _r2_client = boto3.client(
        's3',
        endpoint_url=f'https://{_R2_ACCOUNT_ID}.r2.cloudflarestorage.com',
        aws_access_key_id=_R2_ACCESS_KEY_ID,
        aws_secret_access_key=_R2_SECRET_ACCESS_KEY,
        region_name='auto',
    )
    return _r2_client

def r2_upload(local_path_or_bytes, r2_key, content_type='image/jpeg'):
    """Upload a file to R2. Accepts a local file path (str) or raw bytes."""
    client = _get_r2_client()
    if isinstance(local_path_or_bytes, str):
        client.upload_file(local_path_or_bytes, _R2_BUCKET_NAME, r2_key,
                           ExtraArgs={'ContentType': content_type})
    else:
        client.put_object(Bucket=_R2_BUCKET_NAME, Key=r2_key,
                          Body=local_path_or_bytes, ContentType=content_type)
    logger.info("R2 upload: %s", r2_key)

def r2_delete(r2_key):
    """Delete a single object from R2. Silent if not found."""
    if not r2_key:
        return
    try:
        client = _get_r2_client()
        client.delete_object(Bucket=_R2_BUCKET_NAME, Key=r2_key)
        logger.info("R2 delete: %s", r2_key)
    except Exception as e:
        logger.warning("R2 delete failed for %s: %s", r2_key, e)

def r2_delete_many(r2_keys):
    """Batch-delete a list of R2 keys (max 1000 per call)."""
    keys = [k for k in r2_keys if k]
    if not keys:
        return
    try:
        client = _get_r2_client()
        for i in range(0, len(keys), 1000):
            batch = [{'Key': k} for k in keys[i:i+1000]]
            client.delete_objects(Bucket=_R2_BUCKET_NAME, Delete={'Objects': batch})
        logger.info("R2 batch delete: %d keys", len(keys))
    except Exception as e:
        logger.warning("R2 batch delete failed: %s", e)

def r2_download_bytes(r2_key):
    """Download an R2 object and return raw bytes."""
    client = _get_r2_client()
    resp = client.get_object(Bucket=_R2_BUCKET_NAME, Key=r2_key)
    return resp['Body'].read()

def r2_key_from_path(photo_path):
    """
    Convert a local photo_path like /app/uploads/A001_Rayan.jpg
    to an R2 key like  photos/A001_Rayan.jpg
    """
    if not photo_path:
        return ''
    return 'photos/' + os.path.basename(photo_path)

def photo_url_for_sub(sub, host_url):
    """
    Return the URL to serve a submission's photo.
    - R2 mode + public bucket → direct R2 URL (no proxy needed)
    - R2 mode, no public URL  → proxied through /uploads/<filename>
    - Local mode              → /uploads/<filename>
    """
    path = sub.get('photo_path', '')
    if not path:
        return ''
    filename = os.path.basename(path)
    if _USE_R2 and _R2_PUBLIC_URL:
        return f"{_R2_PUBLIC_URL}/photos/{filename}"
    return host_url.rstrip('/') + '/uploads/' + filename


UPLOADS_DIR     = os.path.join(BASE_DIR, 'uploads')
GENERATED_DIR   = os.path.join(BASE_DIR, 'generated')
TEMPLATES_DIR   = os.path.join(BASE_DIR, 'templates_store')
DB_FILE         = os.path.join(BASE_DIR, 'database.json')
SQLITE_FILE     = os.path.join(BASE_DIR, 'screenplant.sqlite3')
DATABASE_URL    = (
    os.environ.get('SCREENPLANT_DATABASE_URL')
    or os.environ.get('DATABASE_URL')
)
BACKUPS_DIR     = os.path.join(BASE_DIR, 'backups')
FONTS_DIR       = os.path.join(BASE_DIR, 'fonts')

for d in [UPLOADS_DIR, GENERATED_DIR, TEMPLATES_DIR, FONTS_DIR, BACKUPS_DIR]:
    os.makedirs(d, exist_ok=True)

# ── Database (JSON file — simple, portable) ────────────────────────────────────
DEFAULT_DB = {"schools": [], "submissions": [], "templates": [], "orders": [], "settings": {
    "biz_name": "Tariq Printing Solutions",
    "owner": "Tariq Bhaiya",
    "phone": "+91 98765 43210",
    "email": "",
    "price_per_card": 25,
    "recovery_phone": "8553031925"
}, "deleted_items": []}

db_lock = threading.RLock()
db_init_lock = threading.RLock()
db_initialized = False
font_cache = {}
template_cache = {}
generation_jobs = {}
generation_jobs_lock = threading.RLock()
generation_executor = ThreadPoolExecutor(
    max_workers=max(2, int(os.environ.get('SCREENPLANT_GENERATION_WORKERS', '4'))),
    thread_name_prefix='screenplant-gen'
)
rate_bucket = {}
login_failures = {}   # ip -> {'count': int, 'locked_until': float}
reset_otps = {}       # phone -> {'hash': str, 'expires': float, 'attempts': int}
admin_password_hash_override = ''

# Stats cache — avoids running 8+ DB queries on every dashboard refresh
# TTL set to 30s by default; admin sees data at most 30s old, DB load drops 10x+
_stats_cache = {'data': None, 'expires': 0.0}
_stats_cache_lock = threading.Lock()
_STATS_CACHE_TTL = int(os.environ.get('SCREENPLANT_STATS_CACHE_TTL', '30'))

# Analytics daily cache — these counts change slowly; 60s TTL is fine
_analytics_daily_cache = {}           # key: days -> {'data': ..., 'expires': float}
_analytics_daily_cache_lock = threading.Lock()
_ANALYTICS_CACHE_TTL = int(os.environ.get('SCREENPLANT_ANALYTICS_CACHE_TTL', '60'))

# [ENHANCED] Runtime password hashing for the change-password endpoint. Existing
# env-based plaintext auth remains supported for backward compatibility.
def password_hash(raw):
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac('sha256', str(raw).encode('utf-8'), salt, 120000)
    return base64.b64encode(salt + digest).decode('ascii')

def password_matches(raw, stored_hash):
    try:
        data = base64.b64decode(stored_hash.encode('ascii'), validate=True)
        salt, expected = data[:16], data[16:]
        actual = hashlib.pbkdf2_hmac('sha256', str(raw).encode('utf-8'), salt, 120000)
        return hmac.compare_digest(actual, expected)
    except (binascii.Error, ValueError):
        return False

def current_admin_password_ok(raw):
    configured_pass = os.environ.get('SCREENPLANT_ADMIN_PASSWORD', _default_admin_pass)
    try:
        stored_hash = db_get_settings().get('admin_password_hash', '')
        if stored_hash:
            return password_matches(raw, stored_hash)
    except Exception:
        logger.exception("Could not read stored admin password hash")
    if admin_password_hash_override:
        return password_matches(raw, admin_password_hash_override)
    configured_hash = os.environ.get('SCREENPLANT_ADMIN_PASSWORD_HASH', '')
    if configured_hash:
        return password_matches(raw, configured_hash)
    return hmac.compare_digest(str(raw), configured_pass)

def normalize_phone(value):
    digits = re.sub(r'\D+', '', str(value or ''))
    if digits.startswith('91') and len(digits) == 12:
        digits = digits[2:]
    return digits[-10:] if len(digits) >= 10 else digits

def otp_hash(phone, otp):
    secret = app.config.get('SECRET_KEY', '')
    return hmac.new(str(secret).encode('utf-8'), f"{phone}:{otp}".encode('utf-8'), hashlib.sha256).hexdigest()

def send_reset_otp(phone, otp):
    """
    Production hook: set SCREENPLANT_OTP_WEBHOOK_URL to an SMS provider endpoint
    that accepts JSON {"phone": "...", "message": "..."}.
    Local fallback logs the OTP and returns it only for localhost requests.
    """
    message = f"ScreenPlant Pro password reset OTP: {otp}. It expires in 10 minutes."
    webhook = os.environ.get('SCREENPLANT_OTP_WEBHOOK_URL', '').strip()
    if webhook:
        payload = json.dumps({"phone": phone, "message": message}).encode('utf-8')
        req = urllib.request.Request(webhook, data=payload, headers={'Content-Type': 'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status >= 400:
                raise RuntimeError(f"OTP provider returned HTTP {resp.status}")
        return "sms"
    logger.warning("PASSWORD RESET OTP for %s: %s", phone, otp)
    return "local_log"

def pg_driver():
    try:
        import psycopg
        from psycopg.rows import tuple_row
        from psycopg.types.json import Jsonb
        return psycopg, tuple_row, Jsonb
    except ImportError as e:
        raise RuntimeError(
            "PostgreSQL support requires psycopg. Run `pip install -r requirements.txt` "
            "and set SCREENPLANT_DATABASE_URL or DATABASE_URL if your PostgreSQL server is not "
            "postgresql://postgres:postgres@127.0.0.1:5432/screenplant."
        ) from e

def use_postgres():
    return bool(DATABASE_URL)

def sqlite_conn():
    conn = sqlite3.connect(SQLITE_FILE, timeout=30)
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA busy_timeout=30000')
    conn.execute('PRAGMA synchronous=NORMAL')
    conn.execute('PRAGMA cache_size=10000')
    return conn

# ── PostgreSQL connection pool ────────────────────────────────────────────────
_pg_pool = None
_pg_pool_lock = threading.Lock()

def _pg_pool_check(conn):
    """
    Health-check callback for psycopg_pool.
    Called before each connection is handed to a caller from the pool.
    Rolls back any dirty transaction state and issues a lightweight ping.
    On failure the pool discards the connection and opens a fresh one.
    This is the primary fix for AdminShutdown / rolling-back warnings.
    """
    try:
        # transaction_status: 0=IDLE, 1=INTRANS, 2=INERROR, 3=UNKNOWN
        if conn.info.transaction_status != 0:
            conn.rollback()
        conn.execute("SELECT 1")
    except Exception:
        raise   # pool discards + reconnects automatically


def _get_pg_pool():
    global _pg_pool
    if _pg_pool is not None:
        return _pg_pool
    with _pg_pool_lock:
        if _pg_pool is not None:
            return _pg_pool
        try:
            from psycopg_pool import ConnectionPool
            psycopg, tuple_row, _ = pg_driver()
            min_size = int(os.environ.get('SCREENPLANT_PG_POOL_MIN', '3'))
            max_size = int(os.environ.get('SCREENPLANT_PG_POOL_MAX', '25'))
            pool_timeout = float(os.environ.get('SCREENPLANT_PG_POOL_TIMEOUT', '10'))
            _pg_pool = ConnectionPool(
                DATABASE_URL,
                min_size=min_size,
                max_size=max_size,
                timeout=pool_timeout,
                # check= validates every connection before use.
                # Dead connections are caught here and replaced transparently.
                check=_pg_pool_check,
                # reconnect_timeout: max time to keep retrying a dead connection.
                reconnect_timeout=30,
                # wait=False: pool opens without blocking startup.
                # Connections are established in background threads immediately.
                # This prevents the first request from waiting for pool warmup.
                open=True,
                wait=False,
                kwargs={
                    # connect_timeout=5: faster failure if DB is unreachable.
                    # Railway PostgreSQL is always on so 5s is plenty.
                    'connect_timeout': 5,
                    'row_factory': tuple_row,
                    # TCP keepalives prevent silent connection drops on idle pools.
                    'keepalives': 1,
                    'keepalives_idle': 30,
                    'keepalives_interval': 5,
                    'keepalives_count': 5,
                },
            )
            # Warm up the pool immediately — ensures min_size connections are
            # ready before the first real request hits, preventing 10s warmup delay.
            _pg_pool.check()
            logger.info(
                "PostgreSQL pool started (min=%d max=%d timeout=%.0fs check=on keepalives=on)",
                min_size, max_size, pool_timeout
            )
            # Keep-alive ping: fires every 45s to prevent Neon from cold-starting.
            # Uses a direct connection instead of the pool so it never occupies
            # a pool slot and cannot contribute to pool exhaustion.
            _keepalive_db_url = DATABASE_URL
            def _pg_keepalive():
                time.sleep(20)  # brief delay so app finishes starting first
                while True:
                    try:
                        import psycopg as _psycopg_ka
                        with _psycopg_ka.connect(_keepalive_db_url, connect_timeout=5) as _kconn:
                            _kconn.execute("SELECT 1")
                    except Exception:
                        pass
                    time.sleep(45)
            threading.Thread(target=_pg_keepalive, daemon=True, name='pg-keepalive').start()
            logger.info("PostgreSQL keep-alive thread started (interval=45s, direct connection)")
        except ImportError:
            logger.warning(
                "psycopg_pool not installed — falling back to per-request connections. "
                "Run `pip install psycopg-pool` for much better performance."
            )
            _pg_pool = None
    return _pg_pool


# Transient PostgreSQL error codes safe to retry automatically.
# Neon terminates idle connections with AdminShutdown (57P01).
_PG_RETRYABLE_ERRCODES = frozenset({
    "57P01",  # AdminShutdown — primary Neon idle-disconnect error
    "57P02",  # CrashShutdown
    "08000",  # connection_exception
    "08003",  # connection_does_not_exist
    "08006",  # connection_failure
    "08001",  # sqlclient_unable_to_establish_sqlconnection
    "08004",  # sqlserver_rejected_establishment_of_sqlconnection
    "40001",  # serialization_failure
    "40P01",  # deadlock_detected
})

def pg_conn():
    """Legacy — use pg_connection() context manager for pooled connections."""
    pool = _get_pg_pool()
    if pool is not None:
        return pool.getconn()
    psycopg, tuple_row, _ = pg_driver()
    return psycopg.connect(DATABASE_URL, row_factory=tuple_row, connect_timeout=5)

def _release_pg_conn(conn):
    """Return a connection to the pool, or close it if pooling is unavailable."""
    pool = _get_pg_pool()
    if pool is not None:
        try:
            pool.putconn(conn)
        except Exception:
            try:
                conn.close()
            except Exception:
                pass
    else:
        try:
            conn.close()
        except Exception:
            pass

from contextlib import contextmanager as _contextmanager

@_contextmanager
def pg_connection(max_retries: int = 2, _retry_delay: float = 0.25):
    """
    Pool-safe context manager with automatic retry on transient Neon errors.

    The pool's check= callback handles most dead connections before they reach
    the caller, but a connection can still go bad between the health-check and
    the actual query (race window). This retry wrapper catches those cases.

    Retries on: AdminShutdown (57P01), connection_exception (08xxx),
    serialization_failure (40001), deadlock (40P01).

    Usage:
        with pg_connection() as conn:
            conn.execute(...)
            conn.commit()
    """
    pool = _get_pg_pool()
    last_exc = None

    for attempt in range(max_retries + 1):
        if attempt:
            time.sleep(_retry_delay * (2 ** (attempt - 1)))
            logger.warning(
                "pg_connection: retry %d/%d after transient error: %s",
                attempt, max_retries, last_exc
            )

        if pool is not None:
            conn = pool.getconn()
            try:
                yield conn
                return
            except Exception as exc:
                last_exc = exc
                pgcode = getattr(exc, 'pgcode', None) or getattr(getattr(exc, 'diag', None), 'sqlstate', None) or ''
                try:
                    conn.rollback()
                except Exception:
                    pass
                try:
                    pool.putconn(conn)
                except Exception:
                    try:
                        conn.close()
                    except Exception:
                        pass
                # Only retry on known transient connection errors
                if pgcode in _PG_RETRYABLE_ERRCODES and attempt < max_retries:
                    continue
                raise
        else:
            psycopg, tuple_row, _ = pg_driver()
            conn = psycopg.connect(DATABASE_URL, row_factory=tuple_row, connect_timeout=10)
            try:
                yield conn
                return
            except Exception as exc:
                last_exc = exc
                pgcode = getattr(exc, 'pgcode', None) or ''
                try:
                    conn.rollback()
                except Exception:
                    pass
                try:
                    conn.close()
                except Exception:
                    pass
                if pgcode in _PG_RETRYABLE_ERRCODES and attempt < max_retries:
                    continue
                raise

    raise last_exc  # exhausted all retries

@_contextmanager
def pg_ro_connection():
    """
    Read-only connection context manager with autocommit=True.
    Use this for SELECT-only helpers (stats, lists, counts) to avoid
    the 'rolling back returned connection' pool warnings caused by
    psycopg3 implicitly starting a transaction on the first query.
    Autocommit means no BEGIN is issued, so there is nothing to roll back.

    Also safe to use for DDL statements (CREATE INDEX IF NOT EXISTS etc.)
    since those require autocommit — they cannot run inside a transaction block.
    """
    pool = _get_pg_pool()
    if pool is not None:
        # Borrow a pooled connection and temporarily enable autocommit.
        # psycopg3 may have implicitly issued BEGIN on a previously-used
        # pooled connection (status INTRANS), which prevents flipping
        # autocommit. We roll back first to clear any implicit transaction.
        conn = pool.getconn()
        try:
            try:
                conn.rollback()  # clear implicit BEGIN if connection is INTRANS
            except Exception:
                pass
            conn.autocommit = True
            yield conn
        except Exception as exc:
            pgcode = getattr(exc, 'pgcode', None) or ''
            if pgcode in _PG_RETRYABLE_ERRCODES:
                logger.warning("pg_ro_connection transient error (not retried): %s", exc)
            raise
        finally:
            try:
                conn.autocommit = False  # restore before returning to pool
                pool.putconn(conn)
            except Exception:
                try:
                    conn.close()
                except Exception:
                    pass
    else:
        psycopg, tuple_row, _ = pg_driver()
        conn = psycopg.connect(
            DATABASE_URL, row_factory=tuple_row,
            connect_timeout=10, autocommit=True
        )
        try:
            yield conn
        finally:
            try:
                conn.close()
            except Exception:
                pass

def load_sqlite_state():
    if not os.path.exists(SQLITE_FILE):
        return None
    try:
        with closing(sqlite3.connect(SQLITE_FILE, timeout=30)) as conn:
            row = conn.execute('SELECT payload FROM data_state WHERE id = 1').fetchone()
            if row:
                logger.info("Migrating existing SQLite state into PostgreSQL")
                return ensure_db_shape(json.loads(row[0]))
    except (sqlite3.DatabaseError, OSError, json.JSONDecodeError) as e:
        logger.warning("Could not migrate SQLite state; trying JSON backup next: %s", e)
    return None

def load_json_state():
    if not os.path.exists(DB_FILE):
        return None
    try:
        with open(DB_FILE, 'r', encoding='utf-8') as f:
            logger.info("Migrating database.json into PostgreSQL")
            return ensure_db_shape(json.load(f))
    except (OSError, json.JSONDecodeError) as e:
        logger.exception("Could not read database.json; starting fresh: %s", e)
    return None

def db_payload(db):
    return json.loads(json.dumps(ensure_db_shape(db), default=str))

def pg_create_tables(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_schools (
            id TEXT PRIMARY KEY,
            name TEXT,
            school_type TEXT,
            created_at TEXT,
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_submissions (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            unique_id TEXT,
            name TEXT,
            roll_no TEXT,
            class_dept TEXT,
            status TEXT,
            submitted_at TEXT,
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_templates (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            name TEXT,
            file_path TEXT,
            created_at TEXT,
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_orders (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            status TEXT,
            order_date TEXT,
            total NUMERIC,
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_settings (
            id SMALLINT PRIMARY KEY CHECK (id = 1),
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_deleted_items (
            id TEXT PRIMARY KEY,
            item_type TEXT,
            item_id TEXT,
            deleted_at TEXT,
            updated_at TEXT NOT NULL DEFAULT NOW()::TEXT,
            data JSONB NOT NULL
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_school ON screenplant_submissions (school_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_status ON screenplant_submissions (status)')
    conn.execute("CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_identity ON screenplant_submissions (school_id, (data->>'submission_identity_hash'))")
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_templates_school ON screenplant_templates (school_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_orders_school ON screenplant_orders (school_id)')
    # --- DB1: Additional indexes for 15k-row query patterns ---
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_school_status ON screenplant_submissions (school_id, status)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_submitted_at ON screenplant_submissions (submitted_at)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_orders_status_date ON screenplant_orders (status, order_date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_deleted_items_deleted_at ON screenplant_deleted_items (deleted_at)')

def pg_has_relational_data(conn):
    tables = (
        'screenplant_schools',
        'screenplant_submissions',
        'screenplant_templates',
        'screenplant_orders',
        'screenplant_deleted_items',
        'screenplant_settings',
    )
    total = 0
    for table in tables:
        total += conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
    return total > 0

def pg_load_legacy_state(conn):
    legacy_exists = conn.execute("SELECT to_regclass('public.data_state')").fetchone()[0]
    if not legacy_exists:
        return None
    row = conn.execute('SELECT payload FROM data_state WHERE id = 1').fetchone()
    if not row:
        return None
    payload = row[0]
    if isinstance(payload, str):
        payload = json.loads(payload)
    logger.info("Migrating legacy data_state JSONB payload into relational PostgreSQL tables")
    return ensure_db_shape(payload)

def pg_insert_rows(conn, db, Jsonb):
    payload = db_payload(db)
    conn.execute('''
        TRUNCATE screenplant_deleted_items, screenplant_orders, screenplant_templates,
                 screenplant_submissions, screenplant_schools, screenplant_settings
    ''')
    for school in payload.get('schools', []):
        conn.execute(
            '''
            INSERT INTO screenplant_schools (id, name, school_type, created_at, updated_at, data)
            VALUES (%s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                school.get('id') or gen_id('sch_'),
                school.get('name', ''),
                school.get('type', ''),
                school.get('created_at', ''),
                Jsonb(school),
            )
        )
    for sub in payload.get('submissions', []):
        conn.execute(
            '''
            INSERT INTO screenplant_submissions
                (id, school_id, unique_id, name, roll_no, class_dept, status, submitted_at, updated_at, data)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                sub.get('id') or gen_id('sub_'),
                sub.get('school_id', ''),
                sub.get('unique_id', ''),
                sub.get('name', ''),
                sub.get('roll_no', ''),
                sub.get('class_dept', ''),
                sub.get('status', ''),
                sub.get('submitted_at', ''),
                Jsonb(sub),
            )
        )
    for tpl in payload.get('templates', []):
        conn.execute(
            '''
            INSERT INTO screenplant_templates (id, school_id, name, file_path, created_at, updated_at, data)
            VALUES (%s, %s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                tpl.get('id') or gen_id('tpl_'),
                tpl.get('school_id', ''),
                tpl.get('name', ''),
                tpl.get('file', ''),
                tpl.get('created_at', ''),
                Jsonb(tpl),
            )
        )
    for order in payload.get('orders', []):
        total = order.get('total') or order.get('amount') or order.get('price')
        try:
            total = float(total) if total not in (None, '') else None
        except (TypeError, ValueError):
            total = None
        conn.execute(
            '''
            INSERT INTO screenplant_orders (id, school_id, status, order_date, total, updated_at, data)
            VALUES (%s, %s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                order.get('id') or gen_id('ord_'),
                order.get('school_id', ''),
                order.get('status', ''),
                order.get('date') or order.get('created_at', ''),
                total,
                Jsonb(order),
            )
        )
    conn.execute(
        '''
        INSERT INTO screenplant_settings (id, updated_at, data)
        VALUES (1, NOW()::TEXT, %s)
        ''',
        (Jsonb(payload.get('settings', {})),)
    )
    for item in payload.get('deleted_items', []):
        conn.execute(
            '''
            INSERT INTO screenplant_deleted_items (id, item_type, item_id, deleted_at, updated_at, data)
            VALUES (%s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                item.get('id') or gen_id('del_'),
                item.get('type', ''),
                item.get('item_id', ''),
                item.get('deleted_at', ''),
                Jsonb(item),
            )
        )

def pg_load_relational_db(conn):
    db = {
        "schools": [row[0] for row in conn.execute('SELECT data FROM screenplant_schools ORDER BY name, id').fetchall()],
        "submissions": [row[0] for row in conn.execute('SELECT data FROM screenplant_submissions ORDER BY submitted_at, id').fetchall()],
        "templates": [row[0] for row in conn.execute('SELECT data FROM screenplant_templates ORDER BY created_at, id').fetchall()],
        "orders": [row[0] for row in conn.execute('SELECT data FROM screenplant_orders ORDER BY order_date, id').fetchall()],
        "settings": {},
        "deleted_items": [row[0] for row in conn.execute('SELECT data FROM screenplant_deleted_items ORDER BY deleted_at, id').fetchall()],
    }
    row = conn.execute('SELECT data FROM screenplant_settings WHERE id = 1').fetchone()
    if row:
        db["settings"] = row[0]
    return ensure_db_shape(db)

def sqlite_create_tables(conn):
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_schools (
            id TEXT PRIMARY KEY,
            name TEXT,
            school_type TEXT,
            created_at TEXT,
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_submissions (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            unique_id TEXT,
            name TEXT,
            roll_no TEXT,
            class_dept TEXT,
            status TEXT,
            submitted_at TEXT,
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_templates (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            name TEXT,
            file_path TEXT,
            created_at TEXT,
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_orders (
            id TEXT PRIMARY KEY,
            school_id TEXT,
            status TEXT,
            order_date TEXT,
            total REAL,
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS screenplant_deleted_items (
            id TEXT PRIMARY KEY,
            item_type TEXT,
            item_id TEXT,
            deleted_at TEXT,
            updated_at TEXT NOT NULL,
            data TEXT NOT NULL
        )
    ''')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_school ON screenplant_submissions (school_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_status ON screenplant_submissions (status)')
    conn.execute("CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_identity ON screenplant_submissions (school_id, json_extract(data, '$.submission_identity_hash'))")
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_templates_school ON screenplant_templates (school_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_orders_school ON screenplant_orders (school_id)')
    # --- DB1: Additional indexes for 15k-row query patterns ---
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_school_status ON screenplant_submissions (school_id, status)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_submissions_submitted_at ON screenplant_submissions (submitted_at)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_orders_status_date ON screenplant_orders (status, order_date)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_screenplant_deleted_items_deleted_at ON screenplant_deleted_items (deleted_at)')

def sqlite_has_relational_data(conn):
    tables = (
        'screenplant_schools',
        'screenplant_submissions',
        'screenplant_templates',
        'screenplant_orders',
        'screenplant_deleted_items',
        'screenplant_settings',
    )
    return sum(conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0] for table in tables) > 0

def sqlite_load_legacy_state(conn):
    row = conn.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'data_state'").fetchone()
    if not row:
        return None
    row = conn.execute('SELECT payload FROM data_state WHERE id = 1').fetchone()
    if not row:
        return None
    logger.info("Migrating legacy SQLite data_state payload into relational SQLite tables")
    return ensure_db_shape(json.loads(row[0]))

def sqlite_insert_rows(conn, db):
    payload = db_payload(db)
    now = datetime.now().isoformat()
    for table in (
        'screenplant_deleted_items',
        'screenplant_orders',
        'screenplant_templates',
        'screenplant_submissions',
        'screenplant_schools',
        'screenplant_settings',
    ):
        conn.execute(f'DELETE FROM {table}')
    for school in payload.get('schools', []):
        conn.execute(
            'INSERT INTO screenplant_schools (id, name, school_type, created_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?)',
            (school.get('id') or gen_id('sch_'), school.get('name', ''), school.get('type', ''), school.get('created_at', ''), now, json.dumps(school, default=str))
        )
    for sub in payload.get('submissions', []):
        conn.execute(
            '''
            INSERT INTO screenplant_submissions
                (id, school_id, unique_id, name, roll_no, class_dept, status, submitted_at, updated_at, data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                sub.get('id') or gen_id('sub_'), sub.get('school_id', ''), sub.get('unique_id', ''),
                sub.get('name', ''), sub.get('roll_no', ''), sub.get('class_dept', ''),
                sub.get('status', ''), sub.get('submitted_at', ''), now, json.dumps(sub, default=str)
            )
        )
    for tpl in payload.get('templates', []):
        conn.execute(
            'INSERT INTO screenplant_templates (id, school_id, name, file_path, created_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (tpl.get('id') or gen_id('tpl_'), tpl.get('school_id', ''), tpl.get('name', ''), tpl.get('file', ''), tpl.get('created_at', ''), now, json.dumps(tpl, default=str))
        )
    for order in payload.get('orders', []):
        total = order.get('total') or order.get('amount') or order.get('price')
        try:
            total = float(total) if total not in (None, '') else None
        except (TypeError, ValueError):
            total = None
        conn.execute(
            'INSERT INTO screenplant_orders (id, school_id, status, order_date, total, updated_at, data) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (order.get('id') or gen_id('ord_'), order.get('school_id', ''), order.get('status', ''), order.get('date') or order.get('created_at', ''), total, now, json.dumps(order, default=str))
        )
    conn.execute(
        'INSERT INTO screenplant_settings (id, updated_at, data) VALUES (1, ?, ?)',
        (now, json.dumps(payload.get('settings', {}), default=str))
    )
    for item in payload.get('deleted_items', []):
        conn.execute(
            'INSERT INTO screenplant_deleted_items (id, item_type, item_id, deleted_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?)',
            (item.get('id') or gen_id('del_'), item.get('type', ''), item.get('item_id', ''), item.get('deleted_at', ''), now, json.dumps(item, default=str))
        )

def sqlite_load_relational_db(conn):
    db = {
        "schools": [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_schools ORDER BY name, id').fetchall()],
        "submissions": [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_submissions ORDER BY submitted_at, id').fetchall()],
        "templates": [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_templates ORDER BY created_at, id').fetchall()],
        "orders": [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_orders ORDER BY order_date, id').fetchall()],
        "settings": {},
        "deleted_items": [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_deleted_items ORDER BY deleted_at, id').fetchall()],
    }
    row = conn.execute('SELECT data FROM screenplant_settings WHERE id = 1').fetchone()
    if row:
        db["settings"] = json.loads(row[0])
    return ensure_db_shape(db)

def init_sqlite_db():
    with closing(sqlite_conn()) as conn:
        sqlite_create_tables(conn)
        if not sqlite_has_relational_data(conn):
            payload = sqlite_load_legacy_state(conn) or load_json_state() or ensure_db_shape(DEFAULT_DB.copy())
            sqlite_insert_rows(conn, payload)
        conn.commit()

def init_db():
    global db_initialized
    if db_initialized:
        return
    with db_init_lock:
        if db_initialized:
            return
        if not use_postgres():
            init_sqlite_db()
            db_initialized = True
            return
        _, _, Jsonb = pg_driver()

        # ── Step 1: DDL in autocommit mode ────────────────────────────────────
        # CREATE TABLE / CREATE INDEX IF NOT EXISTS must run outside a transaction
        # block. Using pg_ro_connection() (autocommit=True) avoids the
        # UniqueViolation on pg_class_relname_nsp_index that occurs when multiple
        # Gunicorn workers race to create the same index concurrently inside a
        # regular transaction — a common pattern on Railway/Heroku where all
        # workers boot simultaneously against a shared Postgres instance.
        with pg_ro_connection() as ddl_conn:
            pg_create_tables(ddl_conn)

        # ── Step 2: Seed data in a proper transaction ──────────────────────────
        # Only insert legacy/default data if the DB is genuinely empty. This runs
        # in a normal transactional connection so the seed is fully atomic.
        with pg_connection() as conn:
            if not pg_has_relational_data(conn):
                payload = (
                    pg_load_legacy_state(conn)
                    or load_sqlite_state()
                    or load_json_state()
                    or ensure_db_shape(DEFAULT_DB.copy())
                )
                pg_insert_rows(conn, payload, Jsonb)
            conn.commit()

        db_initialized = True

def recover_from_json_backup():
    backups = sorted(
        [os.path.join(BACKUPS_DIR, f) for f in os.listdir(BACKUPS_DIR) if f.endswith('.json')],
        key=os.path.getmtime,
        reverse=True
    )
    for path in backups:
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return ensure_db_shape(json.load(f))
        except (OSError, json.JSONDecodeError):
            logger.warning("Skipping corrupted backup %s", path)
    return ensure_db_shape(DEFAULT_DB.copy())

def load_db():
    init_db()
    if not use_postgres():
        try:
            with closing(sqlite_conn()) as conn:
                return sqlite_load_relational_db(conn)
        except (sqlite3.DatabaseError, json.JSONDecodeError) as e:
            logger.exception("SQLite relational state is corrupted; recovering from latest JSON backup: %s", e)
            db = recover_from_json_backup()
            save_db(db)
            return db
    try:
        with pg_connection() as conn:
            return pg_load_relational_db(conn)
    except (Exception, json.JSONDecodeError) as e:
        logger.exception("PostgreSQL relational state is unavailable or corrupted; recovering from latest JSON backup: %s", e)
        db = recover_from_json_backup()
        save_db(db)
        return db

def save_db(db):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            sqlite_insert_rows(conn, db)
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        # Serialize whole-state saves while the existing route layer still edits
        # a reconstructed in-memory state object.
        conn.execute('SELECT pg_advisory_xact_lock(1)')
        pg_insert_rows(conn, db, Jsonb)
        conn.commit()

def db_get_school(school_id):
    init_db()
    if not school_id:
        return None
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute(
                'SELECT data FROM screenplant_schools WHERE id = ?',
                (school_id,)
            ).fetchone()
            return json.loads(row[0]) if row else None
    with pg_ro_connection() as conn:
        row = conn.execute(
            'SELECT data FROM screenplant_schools WHERE id = %s',
            (school_id,)
        ).fetchone()
        return row[0] if row else None

def db_get_settings():
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_settings WHERE id = 1').fetchone()
            return json.loads(row[0]) if row else {}
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_settings WHERE id = 1').fetchone()
        return row[0] if row else {}

def public_school_payload(school):
    return {
        "id": school.get("id"),
        "name": school.get("name", ""),
        "type": school.get("type", "School"),
        "color": school.get("color", "#1a73e8"),
        "form_schema": school.get("form_schema", []),
        "classes": school.get("classes", []),
    }

def db_list_schools(public_only=True):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute('SELECT data FROM screenplant_schools ORDER BY name, id').fetchall()
            schools = [json.loads(row[0]) for row in rows]
            if public_only:
                return [public_school_payload(s) for s in schools]
            counts = dict(conn.execute('SELECT school_id, COUNT(*) FROM screenplant_submissions GROUP BY school_id').fetchall())
            for school in schools:
                school['submission_count'] = int(counts.get(school.get('id'), 0) or 0)
            return schools
    with pg_ro_connection() as conn:
        rows = conn.execute('SELECT data FROM screenplant_schools ORDER BY name, id').fetchall()
        schools = [row[0] for row in rows]
        if public_only:
            return [public_school_payload(s) for s in schools]
        count_rows = conn.execute('SELECT school_id, COUNT(*) FROM screenplant_submissions GROUP BY school_id').fetchall()
        counts = {row[0]: int(row[1] or 0) for row in count_rows}
        for school in schools:
            school['submission_count'] = counts.get(school.get('id'), 0)
        return schools

def db_get_submission(sub_id):
    init_db()
    if not sub_id:
        return None
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_submissions WHERE id = ?', (sub_id,)).fetchone()
            return json.loads(row[0]) if row else None
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_submissions WHERE id = %s', (sub_id,)).fetchone()
        return row[0] if row else None

def db_update_submission(sub):
    init_db()
    sub_id = sub.get('id')
    if not sub_id:
        raise ValueError("Submission id is required")
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                '''
                UPDATE screenplant_submissions
                SET school_id = ?, unique_id = ?, name = ?, roll_no = ?, class_dept = ?,
                    status = ?, submitted_at = ?, updated_at = ?, data = ?
                WHERE id = ?
                ''',
                (
                    sub.get('school_id', ''), sub.get('unique_id', ''), sub.get('name', ''),
                    sub.get('roll_no', ''), sub.get('class_dept', ''), sub.get('status', ''),
                    sub.get('submitted_at', ''), now, json.dumps(sub, default=str), sub_id
                )
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            '''
            UPDATE screenplant_submissions
            SET school_id = %s, unique_id = %s, name = %s, roll_no = %s, class_dept = %s,
                status = %s, submitted_at = %s, updated_at = NOW()::TEXT, data = %s
            WHERE id = %s
            ''',
            (
                sub.get('school_id', ''), sub.get('unique_id', ''), sub.get('name', ''),
                sub.get('roll_no', ''), sub.get('class_dept', ''), sub.get('status', ''),
                sub.get('submitted_at', ''), Jsonb(sub), sub_id
            )
        )
        conn.commit()

def db_list_templates():
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_templates ORDER BY created_at, id').fetchall()]
    with pg_ro_connection() as conn:
        return [row[0] for row in conn.execute('SELECT data FROM screenplant_templates ORDER BY created_at, id').fetchall()]

def db_get_template_for_school(school_id):
    init_db()
    if not school_id:
        return {}
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_templates WHERE school_id = ? ORDER BY created_at DESC, id DESC LIMIT 1', (school_id,)).fetchone()
            return json.loads(row[0]) if row else {}
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_templates WHERE school_id = %s ORDER BY created_at DESC, id DESC LIMIT 1', (school_id,)).fetchone()
        return row[0] if row else {}

def db_list_orders():
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return [json.loads(row[0]) for row in conn.execute('SELECT data FROM screenplant_orders ORDER BY order_date, id').fetchall()]
    with pg_ro_connection() as conn:
        return [row[0] for row in conn.execute('SELECT data FROM screenplant_orders ORDER BY order_date, id').fetchall()]

# ── Surgical school helpers ────────────────────────────────────────────────────

def db_insert_school(school):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'INSERT INTO screenplant_schools (id, name, school_type, created_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?)',
                (school.get('id'), school.get('name', ''), school.get('type', ''), school.get('created_at', ''), now, json.dumps(school, default=str))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'INSERT INTO screenplant_schools (id, name, school_type, created_at, updated_at, data) VALUES (%s, %s, %s, %s, NOW()::TEXT, %s)',
            (school.get('id'), school.get('name', ''), school.get('type', ''), school.get('created_at', ''), Jsonb(school))
        )
        conn.commit()

def db_update_school(school):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'UPDATE screenplant_schools SET name=?, school_type=?, updated_at=?, data=? WHERE id=?',
                (school.get('name', ''), school.get('type', ''), now, json.dumps(school, default=str), school.get('id'))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'UPDATE screenplant_schools SET name=%s, school_type=%s, updated_at=NOW()::TEXT, data=%s WHERE id=%s',
            (school.get('name', ''), school.get('type', ''), Jsonb(school), school.get('id'))
        )
        conn.commit()

def db_delete_school(school_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute('DELETE FROM screenplant_schools WHERE id = ?', (school_id,))
            conn.commit()
        return
    with pg_connection() as conn:
        conn.execute('DELETE FROM screenplant_schools WHERE id = %s', (school_id,))
        conn.commit()

def db_delete_submissions_by_school(school_id):
    """Return deleted submission records and remove them from DB atomically."""
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute('SELECT data FROM screenplant_submissions WHERE school_id = ?', (school_id,)).fetchall()
            deleted = [json.loads(r[0]) for r in rows]
            conn.execute('DELETE FROM screenplant_submissions WHERE school_id = ?', (school_id,))
            conn.commit()
        return deleted
    with pg_connection() as conn:
        rows = conn.execute('SELECT data FROM screenplant_submissions WHERE school_id = %s', (school_id,)).fetchall()
        deleted = [r[0] for r in rows]
        conn.execute('DELETE FROM screenplant_submissions WHERE school_id = %s', (school_id,))
        conn.commit()
    return deleted

def db_delete_submission_by_id(sub_id):
    """Return the deleted submission record and remove it from DB."""
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_submissions WHERE id = ?', (sub_id,)).fetchone()
            sub = json.loads(row[0]) if row else None
            conn.execute('DELETE FROM screenplant_submissions WHERE id = ?', (sub_id,))
            conn.commit()
        return sub
    with pg_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_submissions WHERE id = %s', (sub_id,)).fetchone()
        sub = row[0] if row else None
        conn.execute('DELETE FROM screenplant_submissions WHERE id = %s', (sub_id,))
        conn.commit()
    return sub

def db_delete_all_submissions():
    """Return all submissions as deleted records and clear the table."""
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute('SELECT data FROM screenplant_submissions').fetchall()
            deleted = [json.loads(r[0]) for r in rows]
            conn.execute('DELETE FROM screenplant_submissions')
            conn.commit()
        return deleted
    with pg_connection() as conn:
        rows = conn.execute('SELECT data FROM screenplant_submissions').fetchall()
        deleted = [r[0] for r in rows]
        conn.execute('DELETE FROM screenplant_submissions')
        conn.commit()
    return deleted

# ── Surgical template helpers ──────────────────────────────────────────────────

def db_insert_template(tpl):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'INSERT INTO screenplant_templates (id, school_id, name, file_path, created_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?, ?)',
                (tpl.get('id'), tpl.get('school_id', ''), tpl.get('name', ''), tpl.get('file', ''), tpl.get('created_at', ''), now, json.dumps(tpl, default=str))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'INSERT INTO screenplant_templates (id, school_id, name, file_path, created_at, updated_at, data) VALUES (%s, %s, %s, %s, %s, NOW()::TEXT, %s)',
            (tpl.get('id'), tpl.get('school_id', ''), tpl.get('name', ''), tpl.get('file', ''), tpl.get('created_at', ''), Jsonb(tpl))
        )
        conn.commit()

def db_update_template(tpl):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'UPDATE screenplant_templates SET name=?, school_id=?, file_path=?, updated_at=?, data=? WHERE id=?',
                (tpl.get('name', ''), tpl.get('school_id', ''), tpl.get('file', ''), now, json.dumps(tpl, default=str), tpl.get('id'))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'UPDATE screenplant_templates SET name=%s, school_id=%s, file_path=%s, updated_at=NOW()::TEXT, data=%s WHERE id=%s',
            (tpl.get('name', ''), tpl.get('school_id', ''), tpl.get('file', ''), Jsonb(tpl), tpl.get('id'))
        )
        conn.commit()

def db_delete_template(tpl_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute('DELETE FROM screenplant_templates WHERE id = ?', (tpl_id,))
            conn.commit()
        return
    with pg_connection() as conn:
        conn.execute('DELETE FROM screenplant_templates WHERE id = %s', (tpl_id,))
        conn.commit()

def db_get_template(tpl_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_templates WHERE id = ?', (tpl_id,)).fetchone()
            return json.loads(row[0]) if row else None
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_templates WHERE id = %s', (tpl_id,)).fetchone()
        return row[0] if row else None

# ── Surgical order helpers ─────────────────────────────────────────────────────

def db_insert_order(order):
    init_db()
    now = datetime.now().isoformat()
    total = order.get('total')
    try:
        total = float(total) if total not in (None, '') else None
    except (TypeError, ValueError):
        total = None
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'INSERT INTO screenplant_orders (id, school_id, status, order_date, total, updated_at, data) VALUES (?, ?, ?, ?, ?, ?, ?)',
                (order.get('id'), order.get('school_id', ''), order.get('status', ''), order.get('date', now), total, now, json.dumps(order, default=str))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'INSERT INTO screenplant_orders (id, school_id, status, order_date, total, updated_at, data) VALUES (%s, %s, %s, %s, %s, NOW()::TEXT, %s)',
            (order.get('id'), order.get('school_id', ''), order.get('status', ''), order.get('date', now), total, Jsonb(order))
        )
        conn.commit()

def db_update_order(order):
    init_db()
    now = datetime.now().isoformat()
    total = order.get('total')
    try:
        total = float(total) if total not in (None, '') else None
    except (TypeError, ValueError):
        total = None
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'UPDATE screenplant_orders SET status=?, total=?, updated_at=?, data=? WHERE id=?',
                (order.get('status', ''), total, now, json.dumps(order, default=str), order.get('id'))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'UPDATE screenplant_orders SET status=%s, total=%s, updated_at=NOW()::TEXT, data=%s WHERE id=%s',
            (order.get('status', ''), total, Jsonb(order), order.get('id'))
        )
        conn.commit()

def db_delete_order(order_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute('DELETE FROM screenplant_orders WHERE id = ?', (order_id,))
            conn.commit()
        return
    with pg_connection() as conn:
        conn.execute('DELETE FROM screenplant_orders WHERE id = %s', (order_id,))
        conn.commit()

def db_get_order(order_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_orders WHERE id = ?', (order_id,)).fetchone()
            return json.loads(row[0]) if row else None
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_orders WHERE id = %s', (order_id,)).fetchone()
        return row[0] if row else None

# ── Surgical settings helpers ──────────────────────────────────────────────────

def db_update_settings(settings_dict):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'INSERT INTO screenplant_settings (id, updated_at, data) VALUES (1, ?, ?) ON CONFLICT(id) DO UPDATE SET updated_at=excluded.updated_at, data=excluded.data',
                (now, json.dumps(settings_dict, default=str))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'INSERT INTO screenplant_settings (id, updated_at, data) VALUES (1, NOW()::TEXT, %s) ON CONFLICT(id) DO UPDATE SET updated_at=NOW()::TEXT, data=EXCLUDED.data',
            (Jsonb(settings_dict),)
        )
        conn.commit()

# ── Surgical deleted_items helpers ────────────────────────────────────────────

def db_insert_deleted_item(item):
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'INSERT INTO screenplant_deleted_items (id, item_type, item_id, deleted_at, updated_at, data) VALUES (?, ?, ?, ?, ?, ?)',
                (item.get('id'), item.get('type', ''), item.get('item_id', ''), item.get('deleted_at', ''), now, json.dumps(item, default=str))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'INSERT INTO screenplant_deleted_items (id, item_type, item_id, deleted_at, updated_at, data) VALUES (%s, %s, %s, %s, NOW()::TEXT, %s)',
            (item.get('id'), item.get('type', ''), item.get('item_id', ''), item.get('deleted_at', ''), Jsonb(item))
        )
        conn.commit()

def db_update_deleted_item(item):
    """Update a deleted_item record in-place (e.g. after purging photos)."""
    init_db()
    now = datetime.now().isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.execute(
                'UPDATE screenplant_deleted_items SET updated_at=?, data=? WHERE id=?',
                (now, json.dumps(item, default=str), item.get('id'))
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            'UPDATE screenplant_deleted_items SET updated_at=NOW()::TEXT, data=%s WHERE id=%s',
            (Jsonb(item), item.get('id'))
        )
        conn.commit()

def db_get_deleted_item(deleted_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute('SELECT data FROM screenplant_deleted_items WHERE id = ?', (deleted_id,)).fetchone()
            return json.loads(row[0]) if row else None
    with pg_ro_connection() as conn:
        row = conn.execute('SELECT data FROM screenplant_deleted_items WHERE id = %s', (deleted_id,)).fetchone()
        return row[0] if row else None

def db_insert_deleted_item_and_submission(sub):
    """Atomically record a deletion + remove the submission row (SQLite only; PG uses same approach)."""
    rec = deleted_record('submission', sub)
    db_insert_deleted_item(rec)
    db_delete_submission_by_id(sub.get('id'))
    return rec

def db_list_deleted_items():
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return [json.loads(r[0]) for r in conn.execute('SELECT data FROM screenplant_deleted_items ORDER BY deleted_at DESC, id DESC').fetchall()]
    with pg_ro_connection() as conn:
        return [r[0] for r in conn.execute('SELECT data FROM screenplant_deleted_items ORDER BY deleted_at DESC, id DESC').fetchall()]

def db_restore_submission_from_deleted(item):
    """Move a submission from deleted_items back into submissions."""
    data = item.get('data') or {}
    db_insert_submission(data)
    item['restored'] = True
    item['restored_at'] = datetime.now().isoformat()
    db_update_deleted_item(item)
    return data

def db_restore_school_from_deleted(item):
    data = item.get('data') or {}
    db_insert_school(data)
    item['restored'] = True
    item['restored_at'] = datetime.now().isoformat()
    db_update_deleted_item(item)
    return data

def db_restore_order_from_deleted(item):
    data = item.get('data') or {}
    db_insert_order(data)
    item['restored'] = True
    item['restored_at'] = datetime.now().isoformat()
    db_update_deleted_item(item)
    return data

def db_insert_submission(sub):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            now = datetime.now().isoformat()
            conn.execute(
                '''
                INSERT INTO screenplant_submissions
                    (id, school_id, unique_id, name, roll_no, class_dept, status, submitted_at, updated_at, data)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    sub.get('id'), sub.get('school_id', ''), sub.get('unique_id', ''),
                    sub.get('name', ''), sub.get('roll_no', ''), sub.get('class_dept', ''),
                    sub.get('status', ''), sub.get('submitted_at', ''), now, json.dumps(sub, default=str)
                )
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        conn.execute(
            '''
            INSERT INTO screenplant_submissions
                (id, school_id, unique_id, name, roll_no, class_dept, status, submitted_at, updated_at, data)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW()::TEXT, %s)
            ''',
            (
                sub.get('id'), sub.get('school_id', ''), sub.get('unique_id', ''),
                sub.get('name', ''), sub.get('roll_no', ''), sub.get('class_dept', ''),
                sub.get('status', ''), sub.get('submitted_at', ''), Jsonb(sub)
            )
        )
        conn.commit()

def db_count_submissions_by_identity(school_id, identity_hash):
    init_db()
    if not school_id or not identity_hash:
        return 0
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return conn.execute(
                '''
                SELECT COUNT(*) FROM screenplant_submissions
                WHERE school_id = ?
                AND json_extract(data, '$.submission_identity_hash') = ?
                ''',
                (school_id, identity_hash)
            ).fetchone()[0]
    with pg_ro_connection() as conn:
        return conn.execute(
            '''
            SELECT COUNT(*) FROM screenplant_submissions
            WHERE school_id = %s
            AND data->>'submission_identity_hash' = %s
            ''',
            (school_id, identity_hash)
        ).fetchone()[0]

def db_list_submissions(school_id='', status='', page=None, per_page=50):
    init_db()
    where = []
    params = []
    if school_id:
        where.append('school_id = ?')
        params.append(school_id)
    if status:
        where.append('status = ?')
        params.append(status)
    where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
    limit_sql = ''
    if page is not None:
        start = (page - 1) * per_page
        limit_sql = ' LIMIT ? OFFSET ?'
        params.extend([per_page, start])
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute(
                f'SELECT data FROM screenplant_submissions{where_sql} ORDER BY submitted_at DESC, id DESC{limit_sql}',
                tuple(params)
            ).fetchall()
            return [json.loads(row[0]) for row in rows]
    pg_where_sql = where_sql.replace('?', '%s')
    pg_limit_sql = limit_sql.replace('?', '%s')
    with pg_ro_connection() as conn:
        rows = conn.execute(
            f'SELECT data FROM screenplant_submissions{pg_where_sql} ORDER BY submitted_at DESC, id DESC{pg_limit_sql}',
            tuple(params)
        ).fetchall()
        return [row[0] for row in rows]

def db_list_submissions_for_generation(school_id='', sub_ids=None):
    init_db()
    if sub_ids:
        ids = [str(x) for x in sub_ids if str(x)]
        if not ids:
            return []
        if not use_postgres():
            placeholders = ','.join(['?'] * len(ids))
            with closing(sqlite_conn()) as conn:
                rows = conn.execute(
                    f'SELECT data FROM screenplant_submissions WHERE id IN ({placeholders})',
                    tuple(ids)
                ).fetchall()
                loaded = [json.loads(row[0]) for row in rows]
                by_id = {sub.get('id'): sub for sub in loaded}
        else:
            with pg_ro_connection() as conn:
                rows = conn.execute(
                    'SELECT data FROM screenplant_submissions WHERE id = ANY(%s)',
                    (ids,)
                ).fetchall()
                by_id = {row[0].get('id'): row[0] for row in rows}
        return [by_id[sub_id] for sub_id in ids if sub_id in by_id]
    if not school_id:
        return []
    return db_list_submissions(school_id=school_id)

def db_mark_submissions_generated(subs):
    init_db()
    if not subs:
        return
    now = datetime.now().isoformat()
    for sub in subs:
        sub['status'] = 'generated'
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            conn.executemany(
                '''
                UPDATE screenplant_submissions
                SET status = ?, updated_at = ?, data = ?
                WHERE id = ?
                ''',
                [('generated', now, json.dumps(sub, default=str), sub.get('id')) for sub in subs]
            )
            conn.commit()
        return
    _, _, Jsonb = pg_driver()
    with pg_connection() as conn:
        with conn.cursor() as cur:
            cur.executemany(
                '''
                UPDATE screenplant_submissions
                SET status = %s, updated_at = NOW()::TEXT, data = %s
                WHERE id = %s
                ''',
                [('generated', Jsonb(sub), sub.get('id')) for sub in subs]
            )
        conn.commit()

def db_count_submissions(school_id='', status=''):
    init_db()
    where = []
    params = []
    if school_id:
        where.append('school_id = ?')
        params.append(school_id)
    if status:
        where.append('status = ?')
        params.append(status)
    where_sql = (' WHERE ' + ' AND '.join(where)) if where else ''
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return conn.execute(
                f'SELECT COUNT(*) FROM screenplant_submissions{where_sql}',
                tuple(params)
            ).fetchone()[0]
    with pg_ro_connection() as conn:
        return conn.execute(
            f"SELECT COUNT(*) FROM screenplant_submissions{where_sql.replace('?', '%s')}",
            tuple(params)
        ).fetchone()[0]

def db_school_submission_stats(school_id):
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            row = conn.execute(
                '''
                SELECT
                    COUNT(*),
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status = 'generated' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status = 'printed' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN COALESCE(json_extract(data, '$.photo_path'), '') != '' THEN 1 ELSE 0 END)
                FROM screenplant_submissions
                WHERE school_id = ?
                ''',
                (school_id,)
            ).fetchone()
    else:
        with pg_ro_connection() as conn:
            row = conn.execute(
                '''
                SELECT
                    COUNT(*),
                    COUNT(*) FILTER (WHERE status = 'pending'),
                    COUNT(*) FILTER (WHERE status = 'generated'),
                    COUNT(*) FILTER (WHERE status = 'printed'),
                    COUNT(*) FILTER (WHERE COALESCE(data->>'photo_path', '') != '')
                FROM screenplant_submissions
                WHERE school_id = %s
                ''',
                (school_id,)
            ).fetchone()
    total, pending, generated, printed, with_photo = [int(x or 0) for x in row]
    return {
        "students": total,
        "pending": pending,
        "generated": generated,
        "printed": printed,
        "with_photo": with_photo,
        "missing_photo": max(total - with_photo, 0),
        "photo_pct": round((with_photo / total) * 100, 1) if total else 0,
        "generated_pct": round(((generated + printed) / total) * 100, 1) if total else 0,
        "printed_pct": round((printed / total) * 100, 1) if total else 0,
    }

def db_stats_payload():
    init_db()
    today = date.today().isoformat()
    today_end = (date.today() + timedelta(days=1)).isoformat()
    week_start = (datetime.now() - timedelta(days=7)).isoformat()
    month_start = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            schools_count = conn.execute('SELECT COUNT(*) FROM screenplant_schools').fetchone()[0]
            templates_count = conn.execute('SELECT COUNT(*) FROM screenplant_templates').fetchone()[0]
            row = conn.execute(
                '''
                SELECT
                    COUNT(*),
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status = 'generated' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN status = 'printed' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN COALESCE(json_extract(data, '$.photo_path'), '') != '' THEN 1 ELSE 0 END),
                    SUM(CASE WHEN submitted_at >= ? AND submitted_at < ? THEN 1 ELSE 0 END),
                    SUM(CASE WHEN submitted_at >= ? THEN 1 ELSE 0 END)
                FROM screenplant_submissions
                ''',
                (today, today_end, week_start)
            ).fetchone()
            total, pending, generated, printed, with_photo, today_count, week_count = [int(x or 0) for x in row]
            revenue = conn.execute(
                '''
                SELECT COALESCE(SUM(total), 0) FROM screenplant_orders
                WHERE status = 'paid' AND COALESCE(order_date, '') >= ?
                ''',
                (month_start,)
            ).fetchone()[0] or 0
            top_rows = conn.execute(
                '''
                SELECT
                    sc.id,
                    sc.name,
                    COALESCE(json_extract(sc.data, '$.color'), '#1a73e8') AS color,
                    COUNT(sub.id) AS submissions,
                    SUM(CASE WHEN sub.status = 'pending' THEN 1 ELSE 0 END) AS pending,
                    SUM(CASE WHEN COALESCE(json_extract(sub.data, '$.photo_path'), '') != '' THEN 1 ELSE 0 END) AS with_photo
                FROM screenplant_schools sc
                LEFT JOIN screenplant_submissions sub ON sub.school_id = sc.id
                GROUP BY sc.id, sc.name, color
                ORDER BY submissions DESC
                LIMIT 8
                '''
            ).fetchall()
    else:
        with pg_ro_connection() as conn:
            schools_count = conn.execute('SELECT COUNT(*) FROM screenplant_schools').fetchone()[0]
            templates_count = conn.execute('SELECT COUNT(*) FROM screenplant_templates').fetchone()[0]
            row = conn.execute(
                '''
                SELECT
                    COUNT(*),
                    COUNT(*) FILTER (WHERE status = 'pending'),
                    COUNT(*) FILTER (WHERE status = 'generated'),
                    COUNT(*) FILTER (WHERE status = 'printed'),
                    COUNT(*) FILTER (WHERE COALESCE(data->>'photo_path', '') != ''),
                    COUNT(*) FILTER (WHERE submitted_at >= %s AND submitted_at < %s),
                    COUNT(*) FILTER (WHERE submitted_at >= %s)
                FROM screenplant_submissions
                ''',
                (today, today_end, week_start)
            ).fetchone()
            total, pending, generated, printed, with_photo, today_count, week_count = [int(x or 0) for x in row]
            revenue = conn.execute(
                '''
                SELECT COALESCE(SUM(total), 0) FROM screenplant_orders
                WHERE status = 'paid' AND COALESCE(order_date, '') >= %s
                ''',
                (month_start,)
            ).fetchone()[0] or 0
            top_rows = conn.execute(
                '''
                SELECT
                    sc.id,
                    sc.name,
                    COALESCE(sc.data->>'color', '#1a73e8') AS color,
                    COUNT(sub.id) AS submissions,
                    COUNT(sub.id) FILTER (WHERE sub.status = 'pending') AS pending,
                    COUNT(sub.id) FILTER (WHERE COALESCE(sub.data->>'photo_path', '') != '') AS with_photo
                FROM screenplant_schools sc
                LEFT JOIN screenplant_submissions sub ON sub.school_id = sc.id
                GROUP BY sc.id, sc.name, color
                ORDER BY submissions DESC
                LIMIT 8
                '''
            ).fetchall()

    top_schools = []
    for sid, name, color, sub_count, pending_count, photo_count in top_rows:
        sub_count = int(sub_count or 0)
        photo_count = int(photo_count or 0)
        top_schools.append({
            "id": sid,
            "name": name,
            "color": color or "#1a73e8",
            "submissions": sub_count,
            "pending": int(pending_count or 0),
            "with_photo": photo_count,
            "completion_pct": round((photo_count / sub_count) * 100, 1) if sub_count else 0
        })
    active_top = [s for s in top_schools if s["submissions"]]
    top_school = max(active_top, key=lambda s: (s.get('completion_pct', 0), s.get('submissions', 0)), default=None)
    missing_photo = max(total - with_photo, 0)
    return {
        "schools": schools_count,
        "submissions": total,
        "pending": pending,
        "generated": generated,
        "printed": printed,
        "with_photo": with_photo,
        "missing_photo": missing_photo,
        "today_count": today_count,
        "this_week_count": week_count,
        "photo_completion_pct": round((with_photo / total) * 100, 1) if total else 0,
        "top_school": top_school,
        "top_schools": top_schools,
        "revenue": float(revenue or 0),
        "templates": templates_count,
    }

def db_analytics_daily_counts(days):
    init_db()
    start = date.today() - timedelta(days=days - 1)
    buckets = {(start + timedelta(days=i)).isoformat(): 0 for i in range(days)}
    start_key = start.isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute(
                '''
                SELECT substr(submitted_at, 1, 10) AS day, COUNT(*)
                FROM screenplant_submissions
                WHERE substr(submitted_at, 1, 10) >= ?
                GROUP BY day
                ''',
                (start_key,)
            ).fetchall()
    else:
        with pg_ro_connection() as conn:
            rows = conn.execute(
                '''
                SELECT substring(submitted_at from 1 for 10) AS day, COUNT(*)
                FROM screenplant_submissions
                WHERE substring(submitted_at from 1 for 10) >= %s
                GROUP BY day
                ''',
                (start_key,)
            ).fetchall()
    for day, count in rows:
        if day in buckets:
            buckets[day] = int(count or 0)
    return [{"date": key, "count": buckets[key]} for key in sorted(buckets)]

def db_analytics_by_school_rows():
    init_db()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            rows = conn.execute(
                '''
                SELECT
                    sc.id,
                    sc.name,
                    COUNT(sub.id) AS total,
                    SUM(CASE WHEN COALESCE(json_extract(sub.data, '$.photo_path'), '') != '' THEN 1 ELSE 0 END) AS with_photo,
                    SUM(CASE WHEN sub.status = 'pending' THEN 1 ELSE 0 END) AS pending,
                    SUM(CASE WHEN sub.status = 'generated' THEN 1 ELSE 0 END) AS generated,
                    SUM(CASE WHEN sub.status = 'printed' THEN 1 ELSE 0 END) AS printed,
                    MAX(sub.submitted_at) AS last_submission
                FROM screenplant_schools sc
                LEFT JOIN screenplant_submissions sub ON sub.school_id = sc.id
                GROUP BY sc.id, sc.name
                ORDER BY total DESC
                '''
            ).fetchall()
    else:
        with pg_ro_connection() as conn:
            rows = conn.execute(
                '''
                SELECT
                    sc.id,
                    sc.name,
                    COUNT(sub.id) AS total,
                    COUNT(sub.id) FILTER (WHERE COALESCE(sub.data->>'photo_path', '') != '') AS with_photo,
                    COUNT(sub.id) FILTER (WHERE sub.status = 'pending') AS pending,
                    COUNT(sub.id) FILTER (WHERE sub.status = 'generated') AS generated,
                    COUNT(sub.id) FILTER (WHERE sub.status = 'printed') AS printed,
                    MAX(sub.submitted_at) AS last_submission
                FROM screenplant_schools sc
                LEFT JOIN screenplant_submissions sub ON sub.school_id = sc.id
                GROUP BY sc.id, sc.name
                ORDER BY total DESC
                '''
            ).fetchall()
    out = []
    for sid, name, total, with_photo, pending, generated, printed, last_submission in rows:
        total = int(total or 0)
        with_photo = int(with_photo or 0)
        pct = round((with_photo / total) * 100, 1) if total else 0
        out.append({
            "school_id": sid,
            "name": name or "",
            "total": total,
            "with_photo": with_photo,
            "pct": pct,
            "pending": int(pending or 0),
            "generated": int(generated or 0),
            "printed": int(printed or 0),
            "health": health_label(pct),
            "last_submission": last_submission or ""
        })
    return out

def db_analytics_csv_rows(days):
    init_db()
    start = date.today() - timedelta(days=days - 1)
    start_key = start.isoformat()
    if not use_postgres():
        with closing(sqlite_conn()) as conn:
            return conn.execute(
                '''
                SELECT
                    substr(sub.submitted_at, 1, 10) AS day,
                    sc.name,
                    COUNT(sub.id) AS submissions,
                    SUM(CASE WHEN COALESCE(json_extract(sub.data, '$.photo_path'), '') != '' THEN 1 ELSE 0 END) AS photos
                FROM screenplant_schools sc
                LEFT JOIN screenplant_submissions sub
                    ON sub.school_id = sc.id
                    AND substr(sub.submitted_at, 1, 10) >= ?
                GROUP BY day, sc.id, sc.name
                HAVING day IS NOT NULL
                ORDER BY day, sc.name
                ''',
                (start_key,)
            ).fetchall()
    with pg_ro_connection() as conn:
        return conn.execute(
            '''
            SELECT
                substring(sub.submitted_at from 1 for 10) AS day,
                sc.name,
                COUNT(sub.id) AS submissions,
                COUNT(sub.id) FILTER (WHERE COALESCE(sub.data->>'photo_path', '') != '') AS photos
            FROM screenplant_schools sc
            LEFT JOIN screenplant_submissions sub
                ON sub.school_id = sc.id
                AND substring(sub.submitted_at from 1 for 10) >= %s
            GROUP BY day, sc.id, sc.name
            HAVING substring(sub.submitted_at from 1 for 10) IS NOT NULL
            ORDER BY day, sc.name
            ''',
            (start_key,)
        ).fetchall()

def create_backup(label='manual'):
    db = load_db()
    return write_backup_payload(label, db)

def write_backup_payload(label, payload):
    safe_label = re.sub(r'[^a-zA-Z0-9_-]+', '_', label)[:30] or 'manual'
    fname = f"screenplant_{safe_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path = os.path.join(BACKUPS_DIR, fname)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f, indent=2, default=str)
    except OSError as e:
        logger.warning("Could not write backup to disk (ephemeral filesystem?): %s", e)
    return fname, path

def backup_metadata(payload):
    if isinstance(payload, dict) and payload.get('backup_type') == 'screenplant_scope_backup':
        data = payload.get('data') or {}
        return {
            "backup_type": "scope",
            "scope": payload.get('scope', ''),
            "label": payload.get('label', ''),
            "schools": len(data.get('schools', [])),
            "submissions": len(data.get('submissions', [])),
            "templates": len(data.get('templates', [])),
            "orders": len(data.get('orders', [])),
        }
    if isinstance(payload, dict):
        return {
            "backup_type": "full",
            "scope": "all",
            "schools": len(payload.get('schools', [])),
            "submissions": len(payload.get('submissions', [])),
            "templates": len(payload.get('templates', [])),
            "orders": len(payload.get('orders', [])),
        }
    return {"backup_type": "unknown", "scope": ""}

def scoped_backup_payload(db, scope, school_id='', submission_ids=None):
    submission_ids = set(submission_ids or [])
    scope = clean_text(scope, 40)
    schools = db.get('schools', [])
    submissions = db.get('submissions', [])
    templates = db.get('templates', [])
    orders = db.get('orders', [])

    if scope == 'all':
        return db_payload(db), 'all'
    if scope in {'school', 'org', 'school_students'}:
        school = next((s for s in schools if s.get('id') == school_id), None)
        if not school:
            raise ValueError("School not found")
        selected_schools = [school] if scope in {'school', 'org'} else []
        selected_subs = [s for s in submissions if s.get('school_id') == school_id] if scope in {'org', 'school_students'} else []
        selected_templates = [t for t in templates if t.get('school_id') == school_id] if scope == 'org' else []
        selected_orders = [o for o in orders if o.get('school_id') == school_id] if scope == 'org' else []
        label = f"{scope}_{school.get('name') or school_id}"
    elif scope == 'selected_students':
        selected_subs = [s for s in submissions if s.get('id') in submission_ids]
        if not selected_subs:
            raise ValueError("No matching students selected")
        school_ids = {s.get('school_id') for s in selected_subs}
        selected_schools = [s for s in schools if s.get('id') in school_ids]
        selected_templates = []
        selected_orders = []
        label = f"selected_students_{len(selected_subs)}"
    else:
        raise ValueError("Invalid backup scope")

    payload = {
        "backup_type": "screenplant_scope_backup",
        "version": "screenplant-pro",
        "scope": scope,
        "label": label,
        "created_at": datetime.now().isoformat(),
        "data": db_payload({
            "schools": selected_schools,
            "submissions": selected_subs,
            "templates": selected_templates,
            "orders": selected_orders,
            "settings": {},
            "deleted_items": []
        })
    }
    return payload, label

def upsert_by_id(items, new_items):
    by_id = {x.get('id'): i for i, x in enumerate(items) if x.get('id')}
    for item in new_items:
        item_id = item.get('id')
        if item_id and item_id in by_id:
            items[by_id[item_id]] = item
        else:
            items.append(item)

def restore_scoped_backup(current_db, payload):
    data = ensure_db_shape(payload.get('data') or {})
    scope = payload.get('scope', '')
    if scope in {'school', 'org', 'selected_students'}:
        upsert_by_id(current_db.setdefault('schools', []), data.get('schools', []))
    if scope in {'org', 'school_students', 'selected_students'}:
        upsert_by_id(current_db.setdefault('submissions', []), data.get('submissions', []))
    if scope == 'org':
        upsert_by_id(current_db.setdefault('templates', []), data.get('templates', []))
        upsert_by_id(current_db.setdefault('orders', []), data.get('orders', []))
    return ensure_db_shape(current_db)

def gen_id(prefix=""):
    return prefix + str(uuid.uuid4())[:8]

ALLOWED_FORM_FIELD_TYPES = {
    'text', 'textarea', 'number', 'email', 'tel', 'date',
    'select', 'radio', 'checkbox', 'photo', 'signature', 'heading'
}
ALLOWED_SUBMISSION_STATUSES = {'pending', 'generated', 'printed', 'paid', 'partial', 'pending_pay'}
MAX_SCHEMA_FIELDS = 60
MAX_IMAGE_BYTES = 8 * 1024 * 1024
MAX_DOC_BYTES = 25 * 1024 * 1024
MAX_ZIP_BYTES = 30 * 1024 * 1024
MAX_ZIP_UNCOMPRESSED = 150 * 1024 * 1024
MAX_ZIP_FILES = 1000
MAX_BULK_GENERATION = int(os.environ.get('SCREENPLANT_MAX_BULK_GENERATION', '500'))
STUDENT_SUBMISSION_LIMIT = max(1, int(os.environ.get('SCREENPLANT_STUDENT_SUBMISSION_LIMIT', '3')))
IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.bmp', '.tif', '.tiff'}
IMAGE_MIMES = {'image/jpeg', 'image/png', 'image/webp'}
DOC_EXTS = IMAGE_EXTS | {'.pdf'}
DOC_MIMES = IMAGE_MIMES | {'application/pdf'}

def ensure_db_shape(db):
    db.setdefault("schools", [])
    db.setdefault("submissions", [])
    db.setdefault("templates", [])
    db.setdefault("orders", [])
    db.setdefault("settings", {})
    db["settings"].setdefault("recovery_phone", "8553031925")
    db.setdefault("deleted_items", [])
    normalize_submission_display_fields(db)
    return db

def normalize_submission_display_fields(db):
    school_map = {s.get('id'): s for s in db.get('schools', [])}
    name_hints = ('name', 'student name', 'employee name')
    for sub in db.get('submissions', []):
        extra = sub.setdefault('extra_fields', {})
        schema = school_map.get(sub.get('school_id'), {}).get('form_schema', [])
        data_fields = [f for f in schema if f.get('type') not in {'photo', 'signature', 'heading'}]
        has_name_field = any(any(h in str(f.get('label', '')).lower() for h in name_hints) and 'father' not in str(f.get('label', '')).lower() for f in data_fields)
        if sub.get('roll_no') and sub.get('id') and str(sub.get('roll_no')).upper() == str(sub.get('id'))[-6:].upper():
            sub['roll_no'] = ''
        if sub.get('name') and not has_name_field:
            if not extra and len(data_fields) == 1:
                extra[clean_text(data_fields[0].get('label') or 'Text', 120)] = clean_text(sub.get('name'), 1000)
                sub['name'] = ''
            elif extra and any(str(v).strip() == str(sub.get('name')).strip() for v in extra.values()):
                sub['name'] = ''
        if not sub.get('unique_id'):
            sub['unique_id'] = sub.get('roll_no') or sub.get('name') or next((v for v in extra.values() if v), '') or sub.get('id')

def clean_text(value, max_len=250):
    value = '' if value is None else str(value)
    value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)
    value = re.sub(r'<[^>]*?>', '', value)
    return value.strip()[:max_len]

def identity_part(value):
    return re.sub(r'[^a-z0-9]+', '', clean_text(value, 250).lower())

def pick_extra_identity(extra_fields):
    strong_hints = (
        'admission', 'admission no', 'student id', 'student code', 'roll',
        'register', 'registration', 'reg no', 'enrollment', 'enrolment',
        'sr no', 'scholar', 'employee id', 'emp id'
    )
    for key, value in (extra_fields or {}).items():
        label = str(key or '').lower()
        if any(hint in label for hint in strong_hints):
            val = identity_part(value)
            if val:
                return val
    return ''

def student_identity_basis(school_id, name='', roll_no='', class_dept='', dob='', father='', extra_fields=None, fallback=''):
    """
    Build a stable student identity for repeat-submission limits.

    Rules:
    - Strong ID wins: roll/admission/student-id fields identify the student.
    - Siblings are safe: phone/email are intentionally excluded.
    - Corrections are safe: address/blood/photo/other changing fields do not
      reset the attempt counter.
    - No-ID forms fall back to name + class + DOB/father when available.
    """
    sid = identity_part(school_id)
    roll = identity_part(roll_no)
    if roll:
        return {"school_id": sid, "mode": "strong_id", "id": roll}
    extra_id = pick_extra_identity(extra_fields or {})
    if extra_id:
        return {"school_id": sid, "mode": "strong_id", "id": extra_id}
    nm = identity_part(name)
    cls = identity_part(class_dept)
    birth = identity_part(dob)
    parent = identity_part(father)
    if nm and cls:
        return {"school_id": sid, "mode": "name_class", "name": nm, "class": cls, "dob": birth, "guardian": parent}
    if nm:
        return {"school_id": sid, "mode": "name_only", "name": nm, "dob": birth, "guardian": parent}
    fb = identity_part(fallback)
    if fb:
        return {"school_id": sid, "mode": "fallback", "value": fb, "class": cls}
    return {"school_id": sid, "mode": "anonymous", "nonce": gen_id("anon_")}

def student_identity_hash_from_basis(basis):
    return hashlib.sha256(json.dumps(basis, sort_keys=True).encode('utf-8')).hexdigest()

def submission_identity_hash(sub):
    basis = student_identity_basis(
        sub.get('school_id', ''),
        sub.get('name', ''),
        sub.get('roll_no', ''),
        sub.get('class_dept', ''),
        sub.get('dob', ''),
        sub.get('father', ''),
        sub.get('extra_fields') or {},
        sub.get('unique_id') or next((v for v in (sub.get('extra_fields') or {}).values() if v), '')
    )
    return student_identity_hash_from_basis(basis)

def parse_float(value, default, min_value=None, max_value=None):
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return default
    if min_value is not None:
        parsed = max(min_value, parsed)
    if max_value is not None:
        parsed = min(max_value, parsed)
    return parsed

def parse_int(value, default, min_value=None, max_value=None):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    if min_value is not None:
        parsed = max(min_value, parsed)
    if max_value is not None:
        parsed = min(max_value, parsed)
    return parsed

def file_size(file_storage):
    pos = file_storage.stream.tell()
    file_storage.stream.seek(0, os.SEEK_END)
    size = file_storage.stream.tell()
    file_storage.stream.seek(pos)
    return size

def validate_upload(file_storage, allowed_exts, allowed_mimes, max_bytes):
    if not file_storage or not file_storage.filename:
        return None
    size = file_size(file_storage)
    if size > max_bytes:
        raise ValueError(f"File is too large. Max allowed is {max_bytes // (1024 * 1024)} MB.")
    ext = os.path.splitext(secure_filename(file_storage.filename))[1].lower()
    if ext not in allowed_exts:
        raise ValueError("Unsupported file type")
    mime = file_storage.mimetype or ''
    if allowed_mimes and mime and mime not in allowed_mimes:
        raise ValueError("Unsupported MIME type")
    if ext in IMAGE_EXTS:
        try:
            img = Image.open(file_storage.stream)
            img.verify()
            file_storage.stream.seek(0)
        except Exception as e:
            file_storage.stream.seek(0)
            raise ValueError("Invalid or corrupted image") from e
    return ext or os.path.splitext(file_storage.filename)[1].lower()

def validate_zip_bytes(raw):
    if len(raw) > MAX_ZIP_BYTES:
        raise ValueError("ZIP file is too large")
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        infos = zf.infolist()
        if len(infos) > MAX_ZIP_FILES:
            raise ValueError("ZIP has too many files")
        total = sum(i.file_size for i in infos)
        if total > MAX_ZIP_UNCOMPRESSED:
            raise ValueError("ZIP expands to too much data")
        for info in infos:
            if os.path.isabs(info.filename) or '..' in info.filename.replace('\\', '/').split('/'):
                raise ValueError("Unsafe ZIP entry path")
    return True

def deleted_record(item_type, item):
    return {
        "id": gen_id("del_"),
        "type": item_type,
        "item_id": item.get('id'),
        "data": item,
        "deleted_at": datetime.now().isoformat()
    }

def safe_archive_name(name, fallback='Export', max_len=80):
    """Return a filesystem-safe name for ZIP folders and filenames."""
    value = clean_text(str(name or fallback), 200)
    value = re.sub(r'[^\w\s\-().]+', '', value).strip()
    value = re.sub(r'\s+', '_', value)
    return (value or fallback)[:max_len]

class QueueZipWriter:
    def __init__(self, out_queue):
        self.out_queue = out_queue
        self.pos = 0

    def write(self, data):
        if data:
            self.pos += len(data)
            self.out_queue.put(bytes(data))
        return len(data)

    def tell(self):
        return self.pos

    def flush(self):
        return None

    def seekable(self):
        return False

    def writable(self):
        return True

def stream_zip_response(entries, download_name):
    """
    Stream a ZIP response using a small queue. The producer writes one archive
    entry at a time; the response yields chunks as zipfile emits them.
    """
    out = queue.Queue(maxsize=8)
    sentinel = object()

    def producer():
        writer = QueueZipWriter(out)
        try:
            with zipfile.ZipFile(writer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for arcname, data_iter in entries:
                    with zf.open(arcname, 'w') as dest:
                        for chunk in data_iter:
                            if chunk:
                                dest.write(chunk)
        except Exception as exc:
            logger.exception("Streaming ZIP failed: %s", exc)
        finally:
            out.put(sentinel)

    def generate():
        thread = threading.Thread(target=producer, daemon=True, name='zip-stream-producer')
        thread.start()
        while True:
            chunk = out.get()
            if chunk is sentinel:
                break
            yield chunk

    headers = {'Content-Disposition': f'attachment; filename="{secure_filename(download_name) or "export.zip"}"'}
    return Response(stream_with_context(generate()), mimetype='application/zip', headers=headers)

def file_bytes_iter(path, chunk_size=1024 * 1024):
    with open(path, 'rb') as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            yield chunk

def r2_bytes_iter(r2_key):
    client = _get_r2_client()
    resp = client.get_object(Bucket=_R2_BUCKET_NAME, Key=r2_key)
    body = resp['Body']
    try:
        for chunk in body.iter_chunks(chunk_size=1024 * 1024):
            if chunk:
                yield chunk
    finally:
        try:
            body.close()
        except Exception:
            pass

def sanitize_form_schema(fields):
    if not isinstance(fields, list):
        return []
    cleaned = []
    seen_ids = set()
    for raw in fields[:MAX_SCHEMA_FIELDS]:
        if not isinstance(raw, dict):
            continue
        field_type = clean_text(raw.get('type'), 30).lower()
        if field_type not in ALLOWED_FORM_FIELD_TYPES:
            continue
        field_id = clean_text(raw.get('id'), 40) or gen_id('f_')
        if field_id in seen_ids:
            field_id = gen_id('f_')
        seen_ids.add(field_id)
        label = clean_text(raw.get('label') or raw.get('headingText') or field_type.title(), 120)
        item = {
            "id": field_id,
            "type": field_type,
            "label": label,
            "placeholder": clean_text(raw.get('placeholder'), 160),
            "required": bool(raw.get('required')),
            "width": "full" if raw.get('width') == 'full' or field_type in {'photo', 'signature', 'textarea', 'heading'} else "half"
        }
        if field_type in {'select', 'radio'}:
            opts = [clean_text(x, 80) for x in str(raw.get('options', '')).split(',')]
            item["options"] = ','.join([x for x in opts if x][:80])
        if field_type == 'heading':
            item["headingText"] = clean_text(raw.get('headingText') or label, 140)
            item["required"] = False
        cleaned.append(item)
    return cleaned


def sanitize_classes(classes):
    """Validate and clean the classes/sections list from the frontend."""
    if not isinstance(classes, list):
        return []
    result = []
    for cls in classes[:100]:
        if not isinstance(cls, dict):
            continue
        name = clean_text(cls.get('name', ''), 80).strip()
        if not name:
            continue
        has_sections = bool(cls.get('hasSections'))
        raw_sections = cls.get('sections', [])
        sections = []
        if has_sections and isinstance(raw_sections, list):
            sections = [clean_text(s, 40).strip() for s in raw_sections[:30]]
            sections = [s for s in sections if s]
        result.append({'name': name, 'hasSections': has_sections, 'sections': sections})
    return result


def unique_upload_path(stem, ext):
    path = os.path.join(UPLOADS_DIR, stem + ext)
    if not os.path.exists(path):
        return path
    for i in range(2, 1000):
        candidate = os.path.join(UPLOADS_DIR, f"{stem}_{i}{ext}")
        if not os.path.exists(candidate):
            return candidate
    return os.path.join(UPLOADS_DIR, f"{stem}_{gen_id()}{ext}")

def is_admin():
    return bool(session.get('admin'))

def csrf_token():
    token = session.get('csrf_token')
    if not token:
        token = hashlib.sha256(os.urandom(32)).hexdigest()
        session['csrf_token'] = token
    return token

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_admin():
            return jsonify({"error": "Authentication required"}), 401
        if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
            sent = request.headers.get('X-CSRF-Token', '')
            if not sent or not hmac.compare_digest(sent, csrf_token()):
                return jsonify({"error": "Invalid CSRF token"}), 403
        return fn(*args, **kwargs)
    return wrapper

def rate_limited(key, limit, window_seconds):
    now = time.time()
    bucket = rate_bucket.setdefault(key, [])
    bucket[:] = [t for t in bucket if now - t < window_seconds]
    if len(bucket) >= limit:
        return True
    bucket.append(now)
    return False

_LOGIN_MAX_FAILURES = int(os.environ.get('SCREENPLANT_LOGIN_MAX_FAILURES', '5'))
_LOGIN_LOCKOUT_SECONDS = int(os.environ.get('SCREENPLANT_LOGIN_LOCKOUT_SECONDS', '900'))  # 15 min

def is_login_locked(ip):
    entry = login_failures.get(ip)
    if not entry:
        return False
    if entry['count'] >= _LOGIN_MAX_FAILURES:
        if time.time() < entry['locked_until']:
            return True
        # Lockout expired — reset
        login_failures.pop(ip, None)
    return False

def record_login_failure(ip):
    entry = login_failures.setdefault(ip, {'count': 0, 'locked_until': 0.0})
    entry['count'] += 1
    if entry['count'] >= _LOGIN_MAX_FAILURES:
        entry['locked_until'] = time.time() + _LOGIN_LOCKOUT_SECONDS
        logger.warning("Login locked for IP %s after %d failures", ip, entry['count'])

def reset_login_failures(ip):
    login_failures.pop(ip, None)

@app.before_request
def before_request_hardening():
    g._req_start = time.monotonic()
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()
    if request.endpoint in {'submit_form'}:
        school_id = request.form.get('school_id', '') if request.method == 'POST' else ''
        rate_key = f"submit:{client_ip}:{school_id}" if school_id else f"submit_form:{client_ip}"
        if rate_limited(rate_key, 20, 60):
            return jsonify({"error": "Too many requests. Please wait and try again."}), 429

@app.after_request
def after_request_hardening(response):
    # MF1: Structured request logging
    duration_ms = round((time.monotonic() - getattr(g, '_req_start', time.monotonic())) * 1000)
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or '-').split(',')[0].strip()
    if not request.path.startswith('/uploads/'):  # skip noisy photo fetches
        logger.info('REQUEST method=%s path=%s status=%d duration_ms=%d ip=%s',
                    request.method, request.path, response.status_code, duration_ms, client_ip)

    # ── Security headers ──────────────────────────────────────────────────────
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['X-XSS-Protection'] = '1; mode=block'

    # HSTS: tell browsers to always use HTTPS (only sent over HTTPS)
    if _is_https:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'

    # Permissions-Policy: disable unused browser APIs (camera, mic, geolocation)
    response.headers['Permissions-Policy'] = (
        'camera=(), microphone=(), geolocation=(), payment=()'
    )

    # Cache-Control: API responses must not be cached by proxies or browsers
    if request.path.startswith('/api/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, private'
        response.headers['Pragma'] = 'no-cache'

    # Content-Security-Policy
    # img-src includes R2 public URL when configured, for direct photo delivery
    r2_img_src = f' {_R2_PUBLIC_URL}' if (_USE_R2 and _R2_PUBLIC_URL) else ''
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        f"img-src 'self' data: blob:{r2_img_src}; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )
    return response

@app.teardown_request
def teardown_request_hardening(exc):
    pass  # db_lock removed — PostgreSQL handles row-level locking natively

@app.errorhandler(400)
def bad_request(_err):
    return jsonify({"error": "Bad request"}), 400

@app.errorhandler(404)
def not_found(_err):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(405)
def method_not_allowed(_err):
    return jsonify({"error": "Method not allowed"}), 405

@app.errorhandler(413)
def payload_too_large(_err):
    return jsonify({"error": "Uploaded data is too large. Maximum allowed size is 25 MB."}), 413

@app.errorhandler(429)
def too_many_requests(_err):
    return jsonify({"error": "Too many requests. Please wait and try again."}), 429

@app.errorhandler(500)
def internal_error(err):
    logger.exception("Unhandled 500 error: %s", err)
    return jsonify({"error": "An internal server error occurred. Please try again."}), 500

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()

    if is_login_locked(client_ip):
        logger.warning("Login attempt from locked IP %s", client_ip)
        return jsonify({"error": "Too many failed attempts. Please try again in 15 minutes."}), 429

    data = request.json or {}

    configured_user = os.environ.get('SCREENPLANT_ADMIN_USER', 'admin')
    user_ok = hmac.compare_digest(str(data.get('username', '')), configured_user)
    pass_ok = current_admin_password_ok(data.get('password', ''))

    if user_ok and pass_ok:
        reset_login_failures(client_ip)
        session.clear()
        session['admin'] = True
        session.permanent = True
        token = csrf_token()
        logger.info("Admin logged in from %s", client_ip)
        return jsonify({"ok": True, "csrf_token": token})

    record_login_failure(client_ip)
    logger.warning("Failed login from %s", client_ip)
    return jsonify({"error": "Invalid credentials."}), 401

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    session.clear()
    return jsonify({"ok": True})

@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    return jsonify({"admin": is_admin(), "username": os.environ.get('SCREENPLANT_ADMIN_USER', 'admin'), "csrf_token": csrf_token() if is_admin() else ""})

@app.route('/api/auth/request-password-reset', methods=['POST'])
def auth_request_password_reset():
    # [NEW] Forgot-password OTP request for the configured recovery phone.
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()
    if rate_limited(f"reset:{client_ip}", 3, 15 * 60):
        return jsonify({"error": "Too many reset requests. Try again later.", "retry_after": 15 * 60}), 429
    data = request.json or {}
    phone = normalize_phone(data.get('phone', ''))
    db = load_db()
    recovery_phone = normalize_phone(db.get('settings', {}).get('recovery_phone') or os.environ.get('SCREENPLANT_RECOVERY_PHONE', '8553031925'))
    generic = {"ok": True, "message": "If this phone is registered, an OTP has been sent."}
    if not phone or not hmac.compare_digest(phone, recovery_phone):
        logger.warning("Password reset requested for non-recovery phone from %s", client_ip)
        return jsonify(generic)
    otp = f"{secrets.randbelow(1000000):06d}"
    reset_otps[phone] = {"hash": otp_hash(phone, otp), "expires": time.time() + 10 * 60, "attempts": 0}
    try:
        delivery = send_reset_otp(phone, otp)
    except Exception as exc:
        logger.exception("Could not deliver reset OTP: %s", exc)
        return jsonify({"error": "Could not send OTP right now. Try again later."}), 503
    response = dict(generic, delivery=delivery)
    if delivery == "local_log" and client_ip in {'127.0.0.1', '::1', 'local'}:
        response["local_otp"] = otp
        response["message"] = "Local development OTP generated."
    return jsonify(response)

@app.route('/api/auth/reset-password', methods=['POST'])
def auth_reset_password():
    # [NEW] Complete forgot-password flow using phone + OTP + new password.
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()
    if rate_limited(f"reset-complete:{client_ip}", 10, 15 * 60):
        return jsonify({"error": "Too many reset attempts. Try again later.", "retry_after": 15 * 60}), 429
    data = request.json or {}
    phone = normalize_phone(data.get('phone', ''))
    otp = re.sub(r'\D+', '', str(data.get('otp', '')))[:6]
    new_password = str(data.get('new_password', ''))
    if len(new_password) < 8:
        return jsonify({"error": "New password must be at least 8 characters"}), 400
    record = reset_otps.get(phone)
    if not record or record.get('expires', 0) < time.time():
        reset_otps.pop(phone, None)
        return jsonify({"error": "OTP expired or invalid. Request a new OTP."}), 400
    record['attempts'] = int(record.get('attempts', 0)) + 1
    if record['attempts'] > 5:
        reset_otps.pop(phone, None)
        return jsonify({"error": "Too many invalid OTP attempts. Request a new OTP."}), 429
    if not otp or not hmac.compare_digest(record.get('hash', ''), otp_hash(phone, otp)):
        return jsonify({"error": "Invalid OTP"}), 400
    new_hash = password_hash(new_password)
    settings = db_get_settings()
    settings['admin_password_hash'] = new_hash
    db_update_settings(settings)
    global admin_password_hash_override
    admin_password_hash_override = new_hash
    reset_otps.pop(phone, None)
    login_failures.pop(client_ip, None)
    audit('auth.password_reset_otp', 'phone=****' + phone[-4:])
    return jsonify({"ok": True})

@app.route('/api/auth/change-password', methods=['PATCH'])
@admin_required
def auth_change_password():
    # [NEW] Persistent password change stored in the app settings payload.
    global admin_password_hash_override
    data = request.json or {}
    current = data.get('current_password', '')
    new_password = str(data.get('new_password', ''))
    if not current_admin_password_ok(current):
        return jsonify({"error": "Current password is incorrect"}), 400
    if len(new_password) < 8:
        return jsonify({"error": "New password must be at least 8 characters"}), 400
    new_hash = password_hash(new_password)
    settings = db_get_settings()
    settings['admin_password_hash'] = new_hash
    db_update_settings(settings)
    admin_password_hash_override = new_hash
    login_failures.clear()
    audit('auth.password_changed')
    return jsonify({"ok": True, "message": "Password changed. Please sign in again with the new password."})

# ── Serve Frontend ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.route('/school-dashboard')
def school_dashboard_page():
    return send_from_directory(FRONTEND_DIR, 'index.html')

@app.route('/<path:path>')
def static_files(path):
    if path.startswith('api/'):
        return jsonify({"error": "API endpoint not found"}), 404
    requested = os.path.abspath(os.path.join(FRONTEND_DIR, path))
    frontend_root = os.path.abspath(FRONTEND_DIR)
    if not requested.startswith(frontend_root + os.sep) and requested != frontend_root:
        return jsonify({"error": "File not found"}), 404
    if os.path.exists(requested) and os.path.isfile(requested):
        return send_from_directory(FRONTEND_DIR, path)
    return send_from_directory(FRONTEND_DIR, 'index.html')

# ── Health check (Railway / uptime monitors ping this) ────────────────────────
@app.route('/api/health', methods=['GET'])
def health_check():
    """
    Lightweight liveness probe.
    Railway and uptime monitors call this every 30s.
    Returns 200 when the app is running, 503 when the DB is unreachable.
    """
    db_ok = False
    db_backend = 'postgres' if use_postgres() else 'sqlite'
    try:
        if use_postgres():
            with pg_connection() as conn:
                conn.execute('SELECT 1').fetchone()
        else:
            with closing(sqlite_conn()) as conn:
                conn.execute('SELECT 1').fetchone()
        db_ok = True
    except Exception as e:
        logger.warning("Health check DB ping failed: %s", e)

    r2_configured = bool(
        _R2_ACCOUNT_ID and _R2_ACCESS_KEY_ID and _R2_SECRET_ACCESS_KEY and _R2_BUCKET_NAME
    )
    r2_ok = None
    if _USE_R2 and r2_configured:
        try:
            _get_r2_client().head_bucket(Bucket=_R2_BUCKET_NAME)
            r2_ok = True
        except Exception as r2_err:
            logger.warning("Health check R2 ping failed: %s", r2_err)
            r2_ok = False
    status = 200 if db_ok else 503
    return jsonify({
        "ok": db_ok,
        "db": db_backend,
        "db_ok": db_ok,
        "storage": "r2" if _USE_R2 else "local",
        "r2_configured": r2_configured if _USE_R2 else False,
        "r2_ok": r2_ok,
        "r2_sdk_available": BOTO3_OK if _USE_R2 else None,
        "ephemeral_fs": _on_railway or _on_render,
        "version": "screenplant-pro"
    }), status

# ── Audit log ─────────────────────────────────────────────────────────────────
def audit(action, detail=''):
    """Write a one-line audit entry to the logger. Extend to DB as needed."""
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'local').split(',')[0].strip()
    logger.info("AUDIT | ip=%s action=%s detail=%s", ip, action, detail)

# [ENHANCED] Shared analytics helpers.
def parse_dt(value):
    try:
        return datetime.fromisoformat(str(value).replace('Z', '+00:00'))
    except (TypeError, ValueError):
        return None

def health_label(completion_pct):
    if completion_pct >= 90:
        return 'excellent'
    if completion_pct >= 70:
        return 'good'
    if completion_pct >= 40:
        return 'fair'
    return 'poor'

def school_health_payload(db, school):
    subs = [s for s in db.get('submissions', []) if s.get('school_id') == school.get('id')]
    total = len(subs)
    with_photo = sum(1 for s in subs if s.get('photo_path'))
    pct = round((with_photo / total) * 100, 1) if total else 0
    last = max((s.get('submitted_at', '') for s in subs), default='')
    return {
        "school_id": school.get('id'),
        "id": school.get('id'),
        "name": school.get('name', ''),
        "total": total,
        "with_photo": with_photo,
        "completion_pct": pct,
        "health": health_label(pct),
        "last_submission": last,
        "pending": sum(1 for s in subs if s.get('status') == 'pending'),
        "generated": sum(1 for s in subs if s.get('status') == 'generated'),
        "printed": sum(1 for s in subs if s.get('status') == 'printed'),
    }

# ── Serve uploaded files (photos, signatures) — admin only ────────────────────
@app.route('/uploads/<filename>')
@admin_required
def serve_upload(filename):
    """
    Photo/signature files contain student PII — require admin auth.
    The client submission form never needs to fetch these directly.
    Serves from local disk first; falls back to R2 when running in cloud mode.
    """
    filename = secure_filename(filename)
    if not filename:
        return jsonify({"error": "File not found"}), 404
    if os.path.exists(os.path.join(UPLOADS_DIR, filename)):
        return send_from_directory(UPLOADS_DIR, filename)
    if os.path.exists(os.path.join(TEMPLATES_DIR, filename)):
        return send_from_directory(TEMPLATES_DIR, filename)
    # R2 fallback — fetch from cloud and stream to client
    if _USE_R2:
        r2_key = 'photos/' + filename
        try:
            data = r2_download_bytes(r2_key)
            ext = os.path.splitext(filename)[1].lower()
            mime = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
                    'png': 'image/png', 'webp': 'image/webp'}.get(ext.lstrip('.'), 'image/jpeg')
            return send_file(io.BytesIO(data), mimetype=mime)
        except Exception as e:
            logger.warning("R2 fetch failed for %s: %s", filename, e)
    return jsonify({"error": "File not found"}), 404

# ══════════════════════════════════════════════════════════════════════════════
#  SCHOOLS API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/schools', methods=['GET'])
def get_schools():
    if not is_admin():
        return jsonify(db_list_schools(public_only=True))
    return jsonify(db_list_schools(public_only=False))

@app.route('/api/schools/<school_id>/health', methods=['GET'])
@admin_required
def get_school_health(school_id):
    # [NEW] School completion and health score endpoint.
    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "School not found"}), 404
    stats = db_school_submission_stats(school_id)
    total = stats['students']
    with_photo = stats['with_photo']
    pct = round((with_photo / total) * 100, 1) if total else 0
    payload = {
        "school_id": school_id,
        "id": school_id,
        "name": school.get('name', ''),
        "total": total,
        "with_photo": with_photo,
        "completion_pct": pct,
        "health": health_label(pct),
        "last_submission": '',
        "pending": stats['pending'],
        "generated": stats['generated'],
        "printed": stats['printed'],
    }
    return jsonify(payload)

@app.route('/api/schools', methods=['POST'])
@admin_required
def add_school():
    data = request.json or {}
    school = {
        "id": gen_id("sch_"),
        "name": clean_text(data.get("name"), 140),
        "type": clean_text(data.get("type") or "School", 40),
        "contact": clean_text(data.get("contact"), 120),
        "phone": clean_text(data.get("phone"), 60),
        "email": clean_text(data.get("email"), 120),
        "address": clean_text(data.get("address"), 300),
        "color": clean_text(data.get("color") or "#1a73e8", 20),
        "notes": clean_text(data.get("notes"), 500),
        "classes": sanitize_classes(data.get("classes", [])),
        "created_at": datetime.now().isoformat()
    }
    if not school["name"]:
        return jsonify({"error": "School name is required"}), 400
    db_insert_school(school)
    audit('school.create', school['id'])
    return jsonify(school), 201

@app.route('/api/schools/<school_id>', methods=['PUT'])
@admin_required
def update_school(school_id):
    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "School not found"}), 404
    data = request.json or {}
    for k in ['name','type','contact','phone','email','address','color','notes']:
        if k in data:
            school[k] = clean_text(data[k], 500 if k == 'notes' else 300)
    if 'classes' in data:
        school['classes'] = sanitize_classes(data['classes'])
    db_update_school(school)
    return jsonify(school)

@app.route('/api/schools/<school_id>', methods=['DELETE'])
@admin_required
def delete_school(school_id):
    school = db_get_school(school_id)
    if school:
        db_insert_deleted_item(deleted_record('school', school))
    deleted_subs = db_delete_submissions_by_school(school_id)
    for sub in deleted_subs:
        db_insert_deleted_item(deleted_record('submission', sub))
    db_delete_school(school_id)
    audit('school.delete', f"id={school_id} cascade_subs={len(deleted_subs)}")
    return jsonify({"ok": True})

@app.route('/api/schools/<school_id>/submissions', methods=['DELETE'])
@admin_required
def delete_school_submissions(school_id):
    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "School not found"}), 404
    deleted_subs = db_delete_submissions_by_school(school_id)
    for sub in deleted_subs:
        db_insert_deleted_item(deleted_record('submission', sub))
    audit('submission.bulk_delete_school', f"school_id={school_id} count={len(deleted_subs)}")
    return jsonify({"ok": True, "deleted": len(deleted_subs), "school_id": school_id})

# ══════════════════════════════════════════════════════════════════════════════
#  TEMPLATE UPLOAD API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/templates', methods=['GET'])
@admin_required
def get_templates():
    return jsonify(db_list_templates())

@app.route('/api/templates', methods=['POST'])
@admin_required
def upload_template():
    """
    Accepts:
      - file: the CorelDraw-exported PNG/JPG/PDF template
      - school_id, name, card_w_mm, card_h_mm,
        fields JSON: {name:{x,y,size,color}, rollno:{...}, ...}
    """
    school_id  = request.form.get('school_id', '')
    tpl_name   = clean_text(request.form.get('name', 'Template'), 140)
    card_w_mm  = parse_float(request.form.get('card_w_mm', 85.6), 85.6, 20, 300)
    card_h_mm  = parse_float(request.form.get('card_h_mm', 54), 54, 20, 300)
    style      = clean_text(request.form.get('style', 'classic'), 40)
    color1     = clean_text(request.form.get('color1', '#1a73e8'), 20)
    color2     = clean_text(request.form.get('color2', '#ffffff'), 20)
    fields_raw = request.form.get('fields', '{}')
    try:
        fields = json.loads(fields_raw)
    except json.JSONDecodeError:
        fields = {}

    file = request.files.get('file')
    file_path = None
    if file and file.filename:
        try:
            ext = validate_upload(file, DOC_EXTS, DOC_MIMES, MAX_DOC_BYTES)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        fname = gen_id("tpl_") + ext
        file_path = os.path.join(TEMPLATES_DIR, fname)
        file.save(file_path)

    template = {
        "id": gen_id("tpl_"),
        "name": tpl_name,
        "school_id": school_id,
        "file": file_path,
        "card_w_mm": card_w_mm,
        "card_h_mm": card_h_mm,
        "style":  style,
        "color1": color1,
        "color2": color2,
        "fields": fields,        # field layout positions
        "created_at": datetime.now().isoformat()
    }
    db_insert_template(template)
    return jsonify(template), 201

@app.route('/api/templates/<tpl_id>', methods=['DELETE'])
@admin_required
def delete_template(tpl_id):
    tpl = db_get_template(tpl_id)
    if tpl and tpl.get('file') and os.path.exists(tpl['file']):
        os.remove(tpl['file'])
    db_delete_template(tpl_id)
    return jsonify({"ok": True})

@app.route('/api/templates/<tpl_id>/duplicate', methods=['POST'])
@admin_required
def duplicate_template(tpl_id):
    # [NEW] Clone a template config while reusing the same stored asset.
    tpl = db_get_template(tpl_id)
    if not tpl:
        return jsonify({"error": "Template not found"}), 404
    clone = json.loads(json.dumps(tpl, default=str))
    clone['id'] = gen_id('tpl_')
    clone['name'] = 'Copy of ' + clean_text(tpl.get('name', 'Template'), 120)
    clone['created_at'] = datetime.now().isoformat()
    db_insert_template(clone)
    audit('template.duplicate', f"id={tpl_id} clone={clone['id']}")
    return jsonify(clone), 201

@app.route('/api/templates/<tpl_id>', methods=['PUT'])
@admin_required
def update_template(tpl_id):
    tpl = db_get_template(tpl_id)
    if not tpl:
        return jsonify({"error": "Template not found"}), 404
    # Update via multipart form (same shape as upload)
    for k in ['name', 'school_id', 'style', 'color1', 'color2']:
        v = request.form.get(k)
        if v is not None:
            tpl[k] = clean_text(v, 200)
    for k in ['card_w_mm', 'card_h_mm']:
        v = request.form.get(k)
        if v is not None:
            tpl[k] = parse_float(v, tpl.get(k, 85.6 if k == 'card_w_mm' else 54), 20, 300)
    fields_raw = request.form.get('fields')
    if fields_raw:
        try:
            tpl['fields'] = json.loads(fields_raw)
        except json.JSONDecodeError:
            pass
    file = request.files.get('file')
    if file and file.filename:
        # Remove old file
        if tpl.get('file') and os.path.exists(tpl['file']):
            os.remove(tpl['file'])
        try:
            ext = validate_upload(file, DOC_EXTS, DOC_MIMES, MAX_DOC_BYTES)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        fname = gen_id("tpl_") + ext
        file_path = os.path.join(TEMPLATES_DIR, fname)
        file.save(file_path)
        tpl['file'] = file_path
    db_update_template(tpl)
    return jsonify(tpl)

# ══════════════════════════════════════════════════════════════════════════════
#  FORM SCHEMA API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/form-schema/<school_id>', methods=['GET'])
def get_form_schema(school_id):
    school = db_get_school(school_id)       # single-row PK lookup — no full-table scan
    if not school:
        return jsonify({"error": "Not found"}), 404
    return jsonify({"fields": school.get('form_schema', [])})

@app.route('/api/form-schema/<school_id>', methods=['PUT'])
@admin_required
def save_form_schema(school_id):
    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "Not found"}), 404
    data = request.json or {}
    school['form_schema'] = sanitize_form_schema(data.get('fields', []))
    db_update_school(school)
    return jsonify({"fields": school['form_schema']})

# ══════════════════════════════════════════════════════════════════════════════
#  SUBMISSIONS API
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/submissions', methods=['GET', 'DELETE'])
@admin_required
def get_submissions():
    if request.method == 'DELETE':
        deleted_subs = db_delete_all_submissions()
        for sub in deleted_subs:
            db_insert_deleted_item(deleted_record('submission', sub))
        audit('submission.bulk_delete_all', f"count={len(deleted_subs)}")
        return jsonify({"ok": True, "deleted": len(deleted_subs)})

    school_id = request.args.get('school_id', '')
    status    = request.args.get('status', '')
    if request.args.get('paginate') == '1':
        page = parse_int(request.args.get('page'), 1, 1, 100000)
        per_page = parse_int(request.args.get('per_page'), 50, 1, 200)
        total = db_count_submissions(school_id, status)
        subs = db_list_submissions(school_id, status, page, per_page)
        pages = max(1, (total + per_page - 1) // per_page)
        return jsonify({"items": subs, "total": total, "page": page, "per_page": per_page, "pages": pages})
    return jsonify(db_list_submissions(school_id, status))

@app.route('/api/schools/<school_id>', methods=['GET'])
@app.route('/api/school-dashboard/<school_id>', methods=['GET'])
@admin_required
def get_school_dashboard(school_id):
    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "School not found"}), 404

    page = parse_int(request.args.get('page'), 1, 1, 100000)
    per_page = parse_int(request.args.get('per_page'), 50, 1, 200)
    total_count = db_count_submissions(school_id)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    submissions = db_list_submissions(school_id=school_id, page=page, per_page=per_page)

    def split_class_section(sub):
        class_dept = clean_text(sub.get('class_dept', ''), 160)
        class_name = clean_text(sub.get('class_name', ''), 80)
        section_name = clean_text(sub.get('section_name', ''), 80)
        if not class_name and class_dept:
            parts = [p.strip() for p in class_dept.split('-', 1)]
            class_name = parts[0]
            if len(parts) > 1 and not section_name:
                section_name = parts[1]
        return class_name, section_name

    safe_submissions = []
    for sub in submissions:
        class_name, section_name = split_class_section(sub)
        photo_url = ''
        if sub.get('photo_path'):
            photo_url = photo_url_for_sub(sub, request.host_url)
        safe_submissions.append({
            "id": sub.get('id', ''),
            "name": sub.get('name') or next((v for v in (sub.get('extra_fields') or {}).values() if v), '') or sub.get('unique_id', ''),
            "roll_no": sub.get('roll_no') or sub.get('unique_id', ''),
            "unique_id": sub.get('unique_id', ''),
            "class_name": class_name,
            "section_name": section_name,
            "class_dept": sub.get('class_dept', ''),
            "status": sub.get('status', 'pending'),
            "submitted_at": sub.get('submitted_at', ''),
            "photo_url": photo_url,
            "has_photo": bool(sub.get('photo_path')),
            "extra_fields": sub.get('extra_fields', {})
        })

    classes = school.get('classes', [])
    if not classes:
        discovered = {}
        for sub in safe_submissions:
            cls = sub.get('class_name') or 'Unassigned'
            sec = sub.get('section_name') or ''
            discovered.setdefault(cls, set())
            if sec:
                discovered[cls].add(sec)
        classes = [
            {"name": cls, "hasSections": bool(sections), "sections": sorted(sections)}
            for cls, sections in sorted(discovered.items())
        ]

    stats = db_school_submission_stats(school_id)

    return jsonify({
        "school": {
            "id": school.get('id'),
            "name": school.get('name', ''),
            "type": school.get('type', 'School'),
            "contact": school.get('contact', ''),
            "phone": school.get('phone', ''),
            "email": school.get('email', ''),
            "address": school.get('address', ''),
            "color": school.get('color', '#ff6b2b'),
            "notes": school.get('notes', ''),
            "classes": classes
        },
        "classes": classes,
        "stats": stats,
        "submissions": safe_submissions,
        "total_count": total_count,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "form_url": f"{request.host_url.rstrip('/')}/?school={school_id}"
    })

@app.route('/api/submissions', methods=['POST'])
def submit_form():
    """
    Multipart form:
      - Fields: name, roll_no (optional), class_dept, dob, blood, phone,
                father, address, email, validity, school_id
      - Files:  photo, signature (optional canvas PNG)

    Photo naming strategy:
      1. If roll_no / id / admission number is in the form → use it:  roll_{rollno}_{name}.jpg
      2. If no ID field but name exists                   → use name: {name}.jpg
      3. Last resort                                      → use sub_id (random)

    The unique_id field is always set to whichever identifier is used —
    this is what the photo injector uses to match photos to cards.
    """
    sub_id = gen_id("sub_")
    school_id = request.form.get('school_id', '')
    if not db_get_school(school_id):
        return jsonify({"error": "Please select a valid school"}), 400

    # Helper: find a form value by a list of possible key names (fixed + dynamic label)
    def fget(fixed_key, *label_hints):
        v = request.form.get(fixed_key, '')
        if v:
            return v
        for key in request.form:
            kl = key.lower()
            for hint in label_hints:
                if hint in kl:
                    return request.form[key]
        return ''

    # ── Extract all standard fields first ─────────────────────────────────────
    name       = fget('name', 'name').strip()
    roll_no    = fget('roll_no', 'roll', 'id number', 'admission', 'reg no', 'emp id', 'student id').strip()
    class_dept = fget('class_dept', 'class', 'dept', 'department', 'section', 'grade', 'branch').strip()
    dob        = fget('dob', 'date of birth', 'birth').strip()
    blood      = fget('blood', 'blood').strip()
    phone      = fget('phone', 'phone', 'mobile', 'contact').strip()
    father     = fget('father', 'father', 'parent', 'guardian').strip()
    address    = fget('address', 'address').strip()
    email      = fget('email', 'email').strip()
    validity   = fget('validity', 'valid', 'expiry', 'expire', 'session').strip()

    # Resolve name — dynamic submit sends label-keyed fields
    system_keys = {'school_id', 'name', 'roll_no', 'class_dept', 'dob', 'blood',
                   'phone', 'father', 'address', 'email', 'validity', 'signature_data'}
    extra_fields = {}
    for key in request.form:
        if key not in system_keys:
            safe_key = clean_text(key, 120)
            safe_val = clean_text(request.form[key], 1000)
            if safe_key and safe_val != '':
                extra_fields[safe_key] = safe_val
    first_extra_value = next((v for v in extra_fields.values() if v), '')

    # ── Determine the unique identifier for this student ──────────────────────
    # Priority: roll_no > name > sub_id fragment
    # This identifier is used to name the photo file AND for injection matching.
    def safe_filename_part(s):
        """Strip characters unsafe for filenames."""
        return secure_filename(re.sub(r'[^\w\-]', '_', s.strip()))[:40] or gen_id('file_')

    if roll_no:
        unique_id = roll_no
        photo_label = f"{safe_filename_part(roll_no)}_{safe_filename_part(name or roll_no)}"
    elif name:
        unique_id = name
        photo_label = safe_filename_part(name)
    elif first_extra_value:
        unique_id = first_extra_value
        photo_label = safe_filename_part(first_extra_value)
    else:
        unique_id = sub_id
        photo_label = sub_id

    # ── Duplicate detection: run BEFORE saving any files to avoid orphaned R2 objects ──
    first_extra_value = next((v for v in extra_fields.values() if v), '')
    identity_basis = student_identity_basis(
        school_id, name, roll_no, class_dept, dob, father, extra_fields, first_extra_value
    )
    identity_hash = student_identity_hash_from_basis(identity_basis)
    same_student_count = db_count_submissions_by_identity(school_id, identity_hash)
    if same_student_count >= STUDENT_SUBMISSION_LIMIT:
        plural = '' if STUDENT_SUBMISSION_LIMIT == 1 else 's'
        return jsonify({"error": f"This student has already submitted the form {STUDENT_SUBMISSION_LIMIT} time{plural}. Please contact the school/admin for changes."}), 409

    # ── Save photo (named after the student's identifier) ─────────────────────
    photo_path = None
    photo_file = request.files.get('photo')
    if photo_file and photo_file.filename:
        try:
            ext = validate_upload(photo_file, IMAGE_EXTS, IMAGE_MIMES, MAX_IMAGE_BYTES) or '.jpg'
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        try:
            photo_path = unique_upload_path(photo_label, ext)
            # Re-save through Pillow to strip EXIF metadata (GPS, device info) and
            # auto-correct orientation before storage. Raw save would preserve PII.
            try:
                from PIL import ImageOps
                img_raw = Image.open(photo_file.stream)
                img_raw = ImageOps.exif_transpose(img_raw)  # fix rotation from EXIF
                img_save = img_raw.convert('RGB')
                save_ext = ext.lower().lstrip('.')
                fmt = 'JPEG' if save_ext in ('jpg', 'jpeg') else save_ext.upper()
                if fmt not in ('JPEG', 'PNG', 'WEBP'):
                    fmt = 'JPEG'
                # Resize oversized photos — students often upload 3-5MB phone shots.
                # Compression strategy:
                # - Photos under 1MB: save at high quality (95), no resize — preserve original quality
                # - Photos over 1MB: resize to 1200px max and compress to quality 82
                # - Photos under 400px (genuinely low-res): always save at quality 95
                _MAX_PHOTO_PX = int(os.environ.get('SCREENPLANT_MAX_PHOTO_PX', '1200'))
                _COMPRESS_THRESHOLD_BYTES = int(os.environ.get('SCREENPLANT_COMPRESS_THRESHOLD_BYTES', str(1 * 1024 * 1024)))  # 1MB default
                # Get original file size
                photo_file.stream.seek(0, 2)  # seek to end
                original_size_bytes = photo_file.stream.tell()
                photo_file.stream.seek(0)  # reset
                is_small_res = max(img_save.width, img_save.height) < 400
                needs_compression = original_size_bytes > _COMPRESS_THRESHOLD_BYTES
                if needs_compression and not is_small_res:
                    # Large photo — resize and compress
                    if max(img_save.width, img_save.height) > _MAX_PHOTO_PX:
                        img_save.thumbnail((_MAX_PHOTO_PX, _MAX_PHOTO_PX), Image.LANCZOS)
                    img_save.save(photo_path, format=fmt, quality=82, optimize=True)
                else:
                    # Small photo or low-res — preserve quality, no resize
                    img_save.save(photo_path, format=fmt, quality=95, optimize=True)
            except Exception:
                # Pillow re-save failed — fall back to raw save (still validated above)
                photo_file.stream.seek(0)
                photo_file.save(photo_path)
            if _USE_R2:
                mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                            '.png': 'image/png', '.webp': 'image/webp'}
                ct = mime_map.get(ext.lower(), 'image/jpeg')
                # Upload to R2 in background — don't block the student's response.
                # The photo is already saved locally; R2 sync happens within seconds.
                _r2_upload_key = r2_key_from_path(photo_path)
                _r2_upload_path = photo_path
                _r2_upload_ct = ct
                def _bg_r2_upload(path=_r2_upload_path, key=_r2_upload_key, content_type=_r2_upload_ct):
                    try:
                        r2_upload(path, key, content_type=content_type)
                    except Exception as _r2e:
                        logger.error("Background R2 upload failed for %s: %s", key, _r2e)
                threading.Thread(target=_bg_r2_upload, daemon=True, name='r2-upload').start()
        except Exception as e:
            logger.exception("Could not save photo for %s: %s", photo_label, e)
            if photo_path and os.path.exists(photo_path):
                try:
                    os.remove(photo_path)
                except OSError:
                    pass
            return jsonify({"error": "Photo could not be saved. Please try a smaller file or contact support."}), 500

    # ── Save signature ─────────────────────────────────────────────────────────
    sig_path = None
    sig_data = request.form.get('signature_data', '')
    if sig_data and sig_data.startswith('data:image'):
        try:
            _header, b64 = sig_data.split(',', 1)
            sig_bytes = base64.b64decode(b64, validate=True)
        except (ValueError, binascii.Error):
            return jsonify({"error": "Invalid signature image"}), 400
        if len(sig_bytes) > MAX_IMAGE_BYTES:
            return jsonify({"error": "Signature is too large"}), 400
        sig_path = unique_upload_path(photo_label + '_sig', '.png')
        with open(sig_path, 'wb') as f:
            f.write(sig_bytes)
        if _USE_R2:
            try:
                _sig_key = r2_key_from_path(sig_path)
                _sig_path = sig_path
                def _bg_sig_upload(path=_sig_path, key=_sig_key):
                    try:
                        r2_upload(path, key, content_type='image/png')
                    except Exception as _se:
                        logger.error("Background R2 signature upload failed for %s: %s", key, _se)
                threading.Thread(target=_bg_sig_upload, daemon=True, name='r2-sig-upload').start()
            except Exception as e:
                logger.exception("Could not schedule signature upload for %s to R2: %s", photo_label, e)

    # ── Store all raw dynamic fields so nothing is lost ────────────────────────
    # Every key the frontend sent (excluding system keys and files) is stored in
    # extra_fields dict. This means custom labels like "Employee Code" are preserved.
    sub = {
        "id":           sub_id,
        "school_id":    school_id,
        "name":         name,
        "roll_no":      roll_no,          # Empty string if not in form — NO auto-generation
        "unique_id":    unique_id,        # What to use for matching and display
        "class_dept":   class_dept,
        "dob":          dob,
        "blood":        blood,
        "phone":        phone,
        "father":       father,
        "address":      address,
        "email":        email,
        "validity":     validity,
        "extra_fields": extra_fields,     # All custom dynamic labels preserved
        "photo_path":   photo_path,
        "photo_label":  photo_label,      # The actual filename stem used
        "sig_path":     sig_path,
        "status":       "pending",
        "submitted_at": datetime.now().isoformat()
    }

    # Identity already computed above (before file saves)
    sub["submission_identity"] = identity_basis
    sub["submission_identity_hash"] = identity_hash
    sub["duplicate_hash"] = identity_hash  # Backward-compatible alias for older data/tools.

    try:
        db_insert_submission(sub)
    except Exception as save_err:
        logger.exception("Failed to save submission %s: %s", sub_id, save_err)
        for fp in [photo_path, sig_path]:
            if fp and os.path.exists(fp):
                try: os.remove(fp)
                except OSError: pass
        if _USE_R2:
            r2_delete_many([r2_key_from_path(fp) for fp in [photo_path, sig_path] if fp])
        return jsonify({"error": "Could not save your submission. Please try again."}), 500
    attempt_no = same_student_count + 1
    return jsonify({
        "ok": True,
        "id": sub_id,
        "unique_id": unique_id,
        "attempt": attempt_no,
        "attempts_remaining": max(0, STUDENT_SUBMISSION_LIMIT - attempt_no)
    }), 201

@app.route('/api/submissions/<sub_id>', methods=['GET'])
@admin_required
def get_submission(sub_id):
    sub = db_get_submission(sub_id)
    if not sub:
        return jsonify({"error": "Not found"}), 404
    return jsonify(sub)

@app.route('/api/submissions/<sub_id>', methods=['PATCH'])
@admin_required
def patch_submission(sub_id):
    # [NEW] Partial submission editor for inline edits and detail modal edits.
    sub = db_get_submission(sub_id)
    if not sub:
        return jsonify({"error": "Not found"}), 404
    data = request.json or {}
    allowed = {
        'name', 'roll_no', 'class_dept', 'dob', 'blood', 'phone', 'father',
        'address', 'email', 'validity', 'unique_id', 'status', 'photo_status',
        'photo_rejection_note'
    }
    for key, value in data.items():
        if key == 'extra_fields' and isinstance(value, dict):
            cleaned = {}
            for k, v in value.items():
                safe_key = clean_text(k, 120)
                if safe_key:
                    cleaned[safe_key] = clean_text(v, 1000)
            sub['extra_fields'] = cleaned
        elif key in allowed:
            if key == 'status' and value not in ALLOWED_SUBMISSION_STATUSES:
                return jsonify({"error": "Invalid status"}), 400
            if key == 'photo_status' and value not in {'approved', 'rejected', 'pending'}:
                return jsonify({"error": "Invalid photo status"}), 400
            sub[key] = clean_text(value, 1000)
    if not sub.get('unique_id'):
        sub['unique_id'] = sub.get('roll_no') or sub.get('name') or sub.get('id')
    db_update_submission(sub)
    audit('submission.patch', f"id={sub_id}")
    return jsonify(sub)

@app.route('/api/submissions/<sub_id>', methods=['DELETE'])
@admin_required
def delete_submission(sub_id):
    sub = db_get_submission(sub_id)
    if sub:
        # Atomically move to trash in a single DB transaction
        rec = deleted_record('submission', sub)
        db_insert_deleted_item(rec)
    db_delete_submission_by_id(sub_id)
    audit('submission.delete', f"id={sub_id}")
    return jsonify({"ok": True})

@app.route('/api/submissions/<sub_id>/status', methods=['PATCH'])
@admin_required
def update_status(sub_id):
    sub = db_get_submission(sub_id)
    if not sub:
        return jsonify({"error": "Not found"}), 404
    data = request.json or {}
    status = data.get('status', sub['status'])
    if status not in ALLOWED_SUBMISSION_STATUSES:
        return jsonify({"error": "Invalid status"}), 400
    sub['status'] = status
    db_update_submission(sub)
    return jsonify(sub)

@app.route('/api/submissions/<sub_id>/photo-status', methods=['PATCH'])
@admin_required
def update_photo_status(sub_id):
    # [NEW] Photo review state stored on the submission payload.
    sub = db_get_submission(sub_id)
    if not sub:
        return jsonify({"error": "Not found"}), 404
    data = request.json or {}
    status = data.get('status', 'pending')
    if status not in {'approved', 'rejected', 'pending'}:
        return jsonify({"error": "Invalid photo status"}), 400
    sub['photo_status'] = status
    if 'note' in data:
        sub['photo_rejection_note'] = clean_text(data.get('note'), 1000)
    db_update_submission(sub)
    audit('submission.photo_status', f"id={sub_id} status={status}")
    return jsonify(sub)

@app.route('/api/deleted', methods=['GET'])
@admin_required
def get_deleted_items():
    item_type = request.args.get('type', '')
    items = db_list_deleted_items()
    if item_type:
        items = [x for x in items if x.get('type') == item_type]
    return jsonify(items)

@app.route('/api/deleted/<deleted_id>/restore', methods=['POST'])
@admin_required
def restore_deleted_item(deleted_id):
    item = db_get_deleted_item(deleted_id)
    if not item:
        return jsonify({"error": "Deleted item not found"}), 404
    item_type = item.get('type')
    data = item.get('data') or {}
    if item_type == 'submission':
        if db_get_submission(data.get('id')):
            return jsonify({"error": "Submission already exists"}), 409
        data = db_restore_submission_from_deleted(item)
    elif item_type == 'school':
        if db_get_school(data.get('id')):
            return jsonify({"error": "School already exists"}), 409
        data = db_restore_school_from_deleted(item)
    elif item_type == 'order':
        data = db_restore_order_from_deleted(item)
    else:
        return jsonify({"error": "This item type cannot be restored"}), 400
    return jsonify({"ok": True, "restored": data})

def purge_deleted_item_files(item):
    data = item.get('data') or {}
    if _USE_R2:
        r2_keys = [
            r2_key_from_path(data.get('photo_path')),
            r2_key_from_path(data.get('sig_path')),
        ]
        r2_delete_many([k for k in r2_keys if k])
    else:
        for fp in [data.get('photo_path'), data.get('sig_path')]:
            if fp and os.path.exists(fp):
                try:
                    os.remove(fp)
                except OSError:
                    pass
    now = datetime.now().isoformat()
    item['photos_purged'] = True
    item['photos_purged_at'] = now
    return now

@app.route('/api/deleted/<deleted_id>', methods=['DELETE'])
@admin_required
def purge_deleted_item(deleted_id):
    item = db_get_deleted_item(deleted_id)
    if not item:
        return jsonify({"error": "Deleted item not found"}), 404
    item['purge_requested_at'] = purge_deleted_item_files(item)
    db_update_deleted_item(item)
    return jsonify({"ok": True, "photos_purged": True, "kept_deleted_record": True})

# ══════════════════════════════════════════════════════════════════════════════
#  CARD GENERATION ENGINE
# ══════════════════════════════════════════════════════════════════════════════

# Card dimensions in pixels at 300 DPI
CARD_DPI    = 300
MM_TO_PX    = CARD_DPI / 25.4   # 1mm = 11.81px at 300dpi

def mm_to_px(mm_val):
    return int(mm_val * MM_TO_PX)

def load_font(size=12, bold=False):
    """Try to load a system font, fall back to default."""
    key = (int(size), bool(bold))
    if key in font_cache:
        return font_cache[key]
    font_names = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf' if bold else
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf' if bold else
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/System/Library/Fonts/Helvetica.ttc',
        'C:/Windows/Fonts/arial.ttf',
    ]
    for fn in font_names:
        if os.path.exists(fn):
            try:
                font_cache[key] = ImageFont.truetype(fn, size)
                return font_cache[key]
            except OSError:
                continue
    font_cache[key] = ImageFont.load_default()
    return font_cache[key]

def template_base_image(tpl_path, width, height):
    if not tpl_path or not os.path.exists(tpl_path):
        return None
    try:
        mtime = os.path.getmtime(tpl_path)
    except OSError:
        return None
    key = (tpl_path, int(width), int(height), mtime)
    cached = template_cache.get(key)
    if cached is not None:
        return cached.copy()
    try:
        img = Image.open(tpl_path).convert('RGBA').resize((width, height), Image.LANCZOS)
    except (OSError, ValueError) as e:
        logger.warning("Template render failed for %s: %s", tpl_path, e)
        return None
    if len(template_cache) > 16:
        template_cache.clear()
    template_cache[key] = img
    return img.copy()

def hex_to_rgb(hex_color):
    h = hex_color.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def draw_text_fitted(draw, text, x, y, max_w, font, color=(255,255,255)):
    """Draw text, shrink font size if too wide."""
    draw.text((x, y), str(text), font=font, fill=color)

def generate_card_image(sub, school, style_cfg):
    """
    Generate a single ID card as a PIL Image.
    style_cfg keys:
      template_path, color1, color2, text_color,
      card_w_mm, card_h_mm, style,
      logo_text, fields
    """
    W = mm_to_px(style_cfg.get('card_w_mm', 85.6))
    H = mm_to_px(style_cfg.get('card_h_mm', 54))
    color1     = hex_to_rgb(style_cfg.get('color1', '#1a73e8'))
    color2     = hex_to_rgb(style_cfg.get('color2', '#ffffff'))
    text_color = hex_to_rgb(style_cfg.get('text_color', '#ffffff'))
    style      = style_cfg.get('style', 'classic')
    school_name = school.get('name', '') if school else ''

    # ── Base: use uploaded template image if exists ───────────────────────────
    tpl_path = style_cfg.get('template_path')
    tpl_base = template_base_image(tpl_path, W, H)
    if tpl_base is not None:
        card = tpl_base
        draw = ImageDraw.Draw(card)
        # Overlay text fields from template field config
        fields_cfg = style_cfg.get('fields', {})
        field_map = {
            'name':       sub.get('name',''),
            'roll_no':    sub.get('roll_no',''),
            'class_dept': sub.get('class_dept',''),
            'dob':        sub.get('dob',''),
            'blood':      sub.get('blood',''),
            'phone':      sub.get('phone',''),
            'validity':   sub.get('validity',''),
            'father':     sub.get('father',''),
        }
        for field_key, cfg in fields_cfg.items():
            value = field_map.get(field_key, '')
            if not value:
                continue
            fx = mm_to_px(cfg.get('x', 10))
            fy = mm_to_px(cfg.get('y', 10))
            fs = int(cfg.get('size', 10) * MM_TO_PX / 4)
            fc = hex_to_rgb(cfg.get('color', '#000000'))
            bold = cfg.get('bold', False)
            font = load_font(fs, bold)
            draw.text((fx, fy), str(value), font=font, fill=fc)
        # Paste photo
        photo_cfg = fields_cfg.get('photo', {})
        if sub.get('photo_path') and os.path.exists(sub['photo_path']):
            px = mm_to_px(photo_cfg.get('x', 55))
            py = mm_to_px(photo_cfg.get('y', 8))
            pw = mm_to_px(photo_cfg.get('w', 20))
            ph = mm_to_px(photo_cfg.get('h', 25))
            photo = Image.open(sub['photo_path']).convert('RGBA').resize((pw, ph), Image.LANCZOS)
            card.paste(photo, (px, py), photo)
    else:
        # ── Built-in styles (no template uploaded) ────────────────────────────
        card = Image.new('RGBA', (W, H), (255,255,255,255))
        draw = ImageDraw.Draw(card)
        if style == 'modern':
            card = _draw_modern(card, draw, sub, school_name, W, H, color1, color2, text_color)
        elif style == 'corporate':
            card = _draw_corporate(card, draw, sub, school_name, W, H, color1, color2, text_color)
        elif style == 'minimal':
            card = _draw_minimal(card, draw, sub, school_name, W, H, color1, color2, text_color)
        else:
            card = _draw_classic(card, draw, sub, school_name, W, H, color1, color2, text_color)

    return card.convert('RGB')


def _draw_classic(card, draw, sub, school_name, W, H, color1, color2, text_color):
    # Header bar
    header_h = int(H * 0.30)
    draw.rectangle([0, 0, W, header_h], fill=color1)
    # School name
    font_school = load_font(int(H * 0.058), bold=True)
    font_small  = load_font(int(H * 0.042))
    font_name   = load_font(int(H * 0.075), bold=True)
    font_field  = load_font(int(H * 0.048))
    draw.text((int(W*0.04), int(header_h*0.18)), school_name, font=font_school, fill=text_color)
    draw.text((int(W*0.04), int(header_h*0.62)), "IDENTITY CARD", font=font_small, fill=(*text_color[:3], 180))
    # Body bg
    draw.rectangle([0, header_h, W, H], fill=color2 if color2 != (255,255,255) else (250,250,250))
    # Photo placeholder / actual
    photo_x, photo_y = int(W*0.62), header_h + int(H*0.05)
    photo_w, photo_h = int(W*0.32), int(H*0.50)
    _paste_photo(card, sub.get('photo_path'), photo_x, photo_y, photo_w, photo_h)
    draw.rectangle([photo_x, photo_y, photo_x+photo_w, photo_y+photo_h],
                   outline=color1, width=2)
    # Text fields
    tx, ty = int(W*0.04), header_h + int(H*0.05)
    body_color = (30,30,30)
    draw.text((tx, ty), sub.get('name',''), font=font_name, fill=color1)
    ty += int(H*0.11)
    for label, val in [("Roll No", sub.get('roll_no','')),
                       ("Class",   sub.get('class_dept','')),
                       ("DOB",     sub.get('dob','')),
                       ("Blood",   sub.get('blood',''))]:
        if val:
            draw.text((tx, ty), f"{label}: ", font=font_field, fill=(120,120,120))
            lw = draw.textlength(f"{label}: ", font=font_field)
            draw.text((tx + int(lw), ty), val, font=font_field, fill=body_color)
            ty += int(H * 0.095)
    # Footer bar
    footer_h = int(H * 0.12)
    draw.rectangle([0, H - footer_h, W, H], fill=color1)
    draw.text((int(W*0.04), H - footer_h + int(footer_h*0.2)),
              f"Valid: {sub.get('validity','')}",
              font=font_small, fill=(*text_color[:3], 200))
    return card


def _draw_modern(card, draw, sub, school_name, W, H, color1, color2, text_color):
    draw.rectangle([0, 0, W, H], fill=(18, 18, 20))
    # Gradient bar top
    bar_h = int(H * 0.38)
    for i in range(bar_h):
        alpha = int(255 * (1 - i / bar_h * 0.3))
        draw.line([0, i, W, i], fill=(*color1, alpha))
    font_school = load_font(int(H * 0.055), bold=True)
    font_name   = load_font(int(H * 0.08), bold=True)
    font_field  = load_font(int(H * 0.048))
    font_tiny   = load_font(int(H * 0.038))
    # Photo circle
    pw = int(H * 0.38); ph = int(H * 0.47)
    px, py = int(W * 0.62), int(H * 0.06)
    _paste_photo(card, sub.get('photo_path'), px, py, pw, ph, circle=True)
    # School
    draw.text((int(W*0.04), int(H*0.06)), school_name, font=font_school, fill=text_color)
    draw.text((int(W*0.04), int(H*0.17)), "IDENTITY CARD", font=font_tiny, fill=(*text_color[:3], 160))
    # Divider
    draw.line([0, bar_h, W, bar_h], fill=(*color1, 80), width=1)
    # Name
    draw.text((int(W*0.04), bar_h + int(H*0.04)), sub.get('name',''), font=font_name, fill=(255,255,255))
    ty = bar_h + int(H*0.16)
    for label, val in [("ID", sub.get('roll_no','')),
                       ("Dept", sub.get('class_dept','')),
                       ("Blood", sub.get('blood',''))]:
        if val:
            draw.text((int(W*0.04), ty), label, font=font_tiny, fill=(150,150,160))
            draw.text((int(W*0.18), ty), val, font=font_field, fill=(220,220,230))
            ty += int(H*0.1)
    # Footer
    draw.rectangle([0, H-int(H*0.10), W, H], fill=(*color1, 60))
    draw.text((int(W*0.04), H-int(H*0.085)),
              f"Valid: {sub.get('validity','')}  |  {school_name}",
              font=font_tiny, fill=(*text_color[:3],150))
    return card


def _draw_corporate(card, draw, sub, school_name, W, H, color1, color2, text_color):
    # Full gradient bg
    for i in range(H):
        r = int(color1[0] + (color2[0]-color1[0]) * i / H)
        g = int(color1[1] + (color2[1]-color1[1]) * i / H)
        b = int(color1[2] + (color2[2]-color1[2]) * i / H)
        draw.line([0, i, W, i], fill=(r,g,b))
    font_large  = load_font(int(H * 0.10), bold=True)
    font_school = load_font(int(H * 0.055), bold=True)
    font_field  = load_font(int(H * 0.048))
    font_tiny   = load_font(int(H * 0.038))
    # Left stripe
    draw.rectangle([0, 0, int(W*0.06), H], fill=(*text_color[:3], 30))
    # Photo
    pw = int(H*0.45); ph = int(H*0.56)
    px, py = int(W*0.63), int(H*0.08)
    _paste_photo(card, sub.get('photo_path'), px, py, pw, ph)
    draw.rectangle([px-2, py-2, px+pw+2, py+ph+2], outline=(255,255,255,120), width=2)
    # Text
    draw.text((int(W*0.10), int(H*0.08)), school_name, font=font_school, fill=text_color)
    draw.text((int(W*0.10), int(H*0.20)), "OFFICIAL ID", font=font_tiny, fill=(*text_color[:3],180))
    draw.line([int(W*0.10), int(H*0.33), int(W*0.55), int(H*0.33)], fill=(*text_color[:3],80), width=1)
    draw.text((int(W*0.10), int(H*0.37)), sub.get('name',''), font=font_large, fill=text_color)
    ty = int(H * 0.54)
    for val in [sub.get('class_dept',''), sub.get('roll_no','')]:
        if val:
            draw.text((int(W*0.10), ty), val, font=font_field, fill=(*text_color[:3],200))
            ty += int(H * 0.11)
    # Bottom strip
    draw.rectangle([0, H-int(H*0.12), W, H], fill=(0,0,0,50))
    if sub.get('blood'):
        draw.text((int(W*0.10), H-int(H*0.10)),
                  f"Blood: {sub['blood']}   Valid: {sub.get('validity','')}",
                  font=font_tiny, fill=(*text_color[:3],200))
    return card


def _draw_minimal(card, draw, sub, school_name, W, H, color1, color2, text_color):
    draw.rectangle([0, 0, W, H], fill=(255,255,255))
    # Top accent line
    draw.rectangle([0, 0, W, int(H*0.06)], fill=color1)
    font_school = load_font(int(H * 0.050), bold=True)
    font_name   = load_font(int(H * 0.080), bold=True)
    font_field  = load_font(int(H * 0.046))
    font_tiny   = load_font(int(H * 0.036))
    body_text   = (30, 30, 30)
    muted       = (120, 120, 130)
    # Photo
    pw = int(H*0.44); ph = int(H*0.56)
    px, py = int(W*0.64), int(H*0.10)
    _paste_photo(card, sub.get('photo_path'), px, py, pw, ph)
    draw.rectangle([px, py, px+pw, py+ph], outline=(*color1, 80), width=2)
    # School name
    draw.text((int(W*0.04), int(H*0.10)), school_name, font=font_school, fill=color1)
    draw.text((int(W*0.04), int(H*0.22)), sub.get('name',''), font=font_name, fill=body_text)
    ty = int(H*0.40)
    draw.line([int(W*0.04), ty, int(W*0.56), ty], fill=(220,220,230), width=1)
    ty += int(H*0.05)
    for label, val in [("Roll No.", sub.get('roll_no','')),
                       ("Class", sub.get('class_dept','')),
                       ("Blood", sub.get('blood',''))]:
        if val:
            draw.text((int(W*0.04), ty), label, font=font_tiny, fill=muted)
            draw.text((int(W*0.24), ty), val, font=font_field, fill=body_text)
            ty += int(H*0.10)
    draw.rectangle([0, H-int(H*0.09), W, H], fill=(*color1, 20))
    draw.text((int(W*0.04), H-int(H*0.075)),
              f"Valid: {sub.get('validity','')}",
              font=font_tiny, fill=muted)
    return card


def _paste_photo(card, photo_path, x, y, w, h, circle=False):
    """Paste a photo (or placeholder) onto the card. Falls back to R2 in cloud mode."""
    img_data = None
    if photo_path:
        if os.path.exists(photo_path):
            img_data = photo_path  # local path — PIL can open directly
        elif _USE_R2:
            try:
                img_data = io.BytesIO(r2_download_bytes(r2_key_from_path(photo_path)))
            except Exception as e:
                logger.warning("R2 fetch for card render failed %s: %s", photo_path, e)
    if img_data is not None:
        try:
            photo = Image.open(img_data).convert('RGBA').resize((w, h), Image.LANCZOS)
            if circle:
                mask = Image.new('L', (w, h), 0)
                md   = ImageDraw.Draw(mask)
                md.ellipse([0, 0, w, h], fill=255)
                photo.putalpha(mask)
            card.paste(photo, (x, y), photo)
            return
        except (OSError, ValueError) as e:
            logger.warning("Photo render failed for %s: %s", photo_path, e)
    # Placeholder
    draw_c = ImageDraw.Draw(card)
    draw_c.rectangle([x, y, x+w, y+h], fill=(60, 60, 70), outline=(100,100,110), width=1)
    font = load_font(max(8, int(h*0.25)))
    draw_c.text((x + w//2 - int(w*0.25), y + h//2 - int(h*0.15)), "👤", font=font, fill=(150,150,160))


def card_to_bytes(img, fmt='JPEG', dpi=300):
    buf = io.BytesIO()
    if fmt == 'JPEG':
        img.convert('RGB').save(buf, format='JPEG', quality=95, dpi=(dpi, dpi))
    else:
        img.save(buf, format='PNG', dpi=(dpi, dpi))
    buf.seek(0)
    return buf


# ── Single card preview ────────────────────────────────────────────────────────
@app.route('/api/generate/preview/<sub_id>', methods=['GET'])
@admin_required
def preview_card(sub_id):
    sub   = db_get_submission(sub_id)
    if not sub:
        return jsonify({"error": "Submission not found"}), 404
    school = db_get_school(sub.get('school_id')) or {}
    tpl    = db_get_template_for_school(sub.get('school_id'))
    style_cfg = {
        'template_path': tpl.get('file'),
        'color1':   tpl.get('color1') or school.get('color', '#1a73e8'),
        'color2':   tpl.get('color2', '#ffffff'),
        'text_color': '#ffffff',
        'card_w_mm': tpl.get('card_w_mm', 85.6),
        'card_h_mm': tpl.get('card_h_mm', 54),
        'style':    tpl.get('style') or request.args.get('style', 'classic'),
        'fields':   tpl.get('fields', {})
    }
    img = generate_card_image(sub, school, style_cfg)
    buf = card_to_bytes(img)
    return send_file(buf, mimetype='image/jpeg')


# ── Bulk generation: returns ZIP of JPEGs ─────────────────────────────────────
def generation_style_config(school, tpl, style):
    return {
        'template_path': tpl.get('file'),
        'color1': tpl.get('color1') or school.get('color', '#1a73e8'),
        'color2': tpl.get('color2', '#ffffff'),
        'text_color': '#ffffff',
        'card_w_mm': tpl.get('card_w_mm', 85.6),
        'card_h_mm': tpl.get('card_h_mm', 54),
        'style': tpl.get('style') or style,
        'fields': tpl.get('fields', {})
    }

def prepare_generation_request(data, force_fmt=None):
    school_id = data.get('school_id', '')
    style = data.get('style', 'classic')
    fmt = (force_fmt or data.get('format', 'jpeg')).lower()
    sub_ids = data.get('sub_ids') or data.get('ids') or None

    if sub_ids:
        subs = db_list_submissions_for_generation(sub_ids=sub_ids)
        school_ids = {s.get('school_id') for s in subs}
        if len(school_ids) > 1:
            raise ValueError("Selected cards must belong to one school per PDF batch")
        school_id = next(iter(school_ids), school_id) if school_ids else school_id
    else:
        subs = db_list_submissions_for_generation(school_id=school_id)
    if not subs:
        raise LookupError("No submissions found")
    if len(subs) > MAX_BULK_GENERATION:
        raise OverflowError(f"Too many cards at once ({len(subs)}). Export in batches of {MAX_BULK_GENERATION} using class/section filters.")

    school = db_get_school(school_id) or {}
    tpl = db_get_template_for_school(school_id)
    return fmt, subs, school, generation_style_config(school, tpl, style)

def render_generation_artifact(subs, school, style_cfg, fmt):
    safe_school = safe_archive_name(school.get('name', 'cards'), 'cards')
    if fmt == 'pdf':
        buf = _generate_pdf(subs, school, style_cfg)
        return buf, 'application/pdf', f"{safe_school}_{datetime.now().strftime('%Y%m%d')}.pdf"

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for sub in subs:
            img = generate_card_image(sub, school, style_cfg)
            ibuf = card_to_bytes(img)
            stem = safe_archive_name(f"{sub.get('roll_no') or sub.get('unique_id') or 'card'}_{sub.get('name','')}", 'card')
            zf.writestr(f"{stem}.jpg", ibuf.read())
    zip_buf.seek(0)
    return zip_buf, 'application/zip', f"{safe_school}_{datetime.now().strftime('%Y%m%d')}.zip"

def generation_job_snapshot(job):
    return {
        "id": job.get("id"),
        "status": job.get("status"),
        "progress": job.get("progress", 0),
        "message": job.get("message", ""),
        "download_ready": bool(job.get("path") and job.get("status") == "done"),
        "filename": job.get("filename", ""),
        "created_at": job.get("created_at", ""),
        "finished_at": job.get("finished_at", ""),
    }

def generation_job_state_path(job_id):
    return os.path.join(GENERATED_DIR, f"{secure_filename(job_id)}.json")

def write_generation_job(job):
    job_id = job.get('id')
    if not job_id:
        return
    generation_jobs[job_id] = dict(job)
    path = generation_job_state_path(job_id)
    tmp_path = path + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(job, f, default=str)
    os.replace(tmp_path, path)

def read_generation_job(job_id):
    with generation_jobs_lock:
        if job_id in generation_jobs:
            return dict(generation_jobs[job_id])
    path = generation_job_state_path(job_id)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as f:
            job = json.load(f)
    except (OSError, json.JSONDecodeError):
        return None
    with generation_jobs_lock:
        generation_jobs[job_id] = dict(job)
    return job

def update_generation_job(job_id, **changes):
    with generation_jobs_lock:
        job = generation_jobs.get(job_id) or read_generation_job(job_id) or {"id": job_id}
        job.update(changes)
        write_generation_job(job)
        return dict(job)

def cleanup_generation_jobs():
    cutoff = time.time() - int(os.environ.get('SCREENPLANT_GENERATION_JOB_TTL_SECONDS', '3600'))
    with generation_jobs_lock:
        stale_ids = [
            job_id for job_id, job in generation_jobs.items()
            if job.get('finished_epoch', job.get('created_epoch', 0)) < cutoff
        ]
        for job_id in stale_ids:
            path = generation_jobs[job_id].get('path')
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except OSError:
                    pass
            state_path = generation_job_state_path(job_id)
            if os.path.exists(state_path):
                try:
                    os.remove(state_path)
                except OSError:
                    pass
            generation_jobs.pop(job_id, None)

def run_generation_job(job_id, data):
    try:
        update_generation_job(job_id, status='running', progress=5, message='Preparing cards')
        fmt, subs, school, style_cfg = prepare_generation_request(data)
        update_generation_job(job_id, progress=15, message=f"Rendering {len(subs)} cards")
        buf, mimetype, filename = render_generation_artifact(subs, school, style_cfg, fmt)
        out_path = os.path.join(GENERATED_DIR, f"{job_id}_{secure_filename(filename)}")
        with open(out_path, 'wb') as f:
            f.write(buf.getvalue())
        db_mark_submissions_generated(subs)
        update_generation_job(
            job_id,
            status='done',
            progress=100,
            message='Ready to download',
            path=out_path,
            mimetype=mimetype,
            filename=filename,
            finished_at=datetime.now().isoformat(),
            finished_epoch=time.time(),
        )
    except Exception as e:
        logger.exception("Generation job %s failed: %s", job_id, e)
        update_generation_job(
            job_id,
            status='error',
            progress=100,
            message=str(e),
            finished_at=datetime.now().isoformat(),
            finished_epoch=time.time(),
        )

@app.route('/api/generate/jobs', methods=['POST'])
@admin_required
def create_generation_job():
    cleanup_generation_jobs()
    data = request.json or {}
    job_id = gen_id('gen_')
    with generation_jobs_lock:
        job = {
            "id": job_id,
            "status": "queued",
            "progress": 0,
            "message": "Queued",
            "created_at": datetime.now().isoformat(),
            "created_epoch": time.time(),
        }
        write_generation_job(job)
    generation_executor.submit(run_generation_job, job_id, data)
    return jsonify(generation_job_snapshot(job)), 202

@app.route('/api/generate/jobs/<job_id>', methods=['GET'])
@admin_required
def get_generation_job(job_id):
    job = read_generation_job(job_id)
    if not job:
        return jsonify({"error": "Generation job not found"}), 404
    return jsonify(generation_job_snapshot(job))

@app.route('/api/generate/jobs/<job_id>/download', methods=['GET'])
@admin_required
def download_generation_job(job_id):
    job = read_generation_job(job_id)
    if not job:
        return jsonify({"error": "Generation job not found"}), 404
    if job.get('status') != 'done' or not job.get('path'):
        return jsonify({"error": "Generation is not ready yet"}), 409
    path = job.get('path')
    mimetype = job.get('mimetype', 'application/octet-stream')
    filename = job.get('filename', os.path.basename(path))
    return send_file(path, mimetype=mimetype, download_name=filename, as_attachment=True)

@app.route('/api/generate/bulk', methods=['POST'])
@admin_required
def bulk_generate():
    data = request.json or {}
    sub_ids = data.get('sub_ids') or data.get('ids') or []
    # SC2: Force large batches through the async job API to avoid RAM spikes
    if not sub_ids or len(sub_ids) > 50:
        return jsonify({
            "error": "Use POST /api/generate/jobs for batches over 50 cards or school-wide exports.",
            "use_jobs_api": True
        }), 400
    return bulk_generate_response()

def bulk_generate_response(force_fmt=None):
    """
    Body: { school_id, style, format: "jpeg"|"pdf", sub_ids: [] (optional) }
    Returns: ZIP file with one image per card (or one PDF)
    """
    try:
        fmt, subs, school, style_cfg = prepare_generation_request(request.json or {}, force_fmt)
    except LookupError:
        return jsonify({"error": "No submissions found"}), 404
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except OverflowError as e:
        return jsonify({"error": str(e)}), 400

    buf, mimetype, filename = render_generation_artifact(subs, school, style_cfg, fmt)
    db_mark_submissions_generated(subs)
    return send_file(buf, mimetype=mimetype, download_name=filename)


def _generate_pdf(subs, school, style_cfg):
    """
    Generate a print-ready PDF.
    Layout: 2 cards per row, with crop marks.
    """
    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=A4)
    page_w, page_h = A4
    card_w = style_cfg.get('card_w_mm', 85.6) * mm
    card_h = style_cfg.get('card_h_mm', 54)   * mm
    margin = 10 * mm
    gap    = 5  * mm
    cols   = int((page_w - 2*margin + gap) / (card_w + gap))
    rows_per_page = int((page_h - 2*margin + gap) / (card_h + gap))
    cards_per_page = cols * rows_per_page

    for idx, sub in enumerate(subs):
        pos_in_page = idx % cards_per_page
        col = pos_in_page % cols
        row = pos_in_page // cols

        if pos_in_page == 0 and idx > 0:
            c.showPage()

        x = margin + col * (card_w + gap)
        y = page_h - margin - (row + 1) * card_h - row * gap

        img = generate_card_image(sub, school, style_cfg)
        ibuf = card_to_bytes(img, fmt='JPEG', dpi=CARD_DPI)
        c.drawImage(ImageReader(ibuf), x, y, width=card_w, height=card_h)

        # Crop marks (helpful for cutting)
        c.setStrokeColorRGB(0.7, 0.7, 0.7)
        c.setLineWidth(0.3)
        mark = 3 * mm
        c.line(x - mark, y + card_h, x, y + card_h)
        c.line(x, y + card_h + mark, x, y + card_h)
        c.line(x + card_w, y + card_h, x + card_w + mark, y + card_h)
        c.line(x + card_w, y + card_h + mark, x + card_w, y + card_h)
        c.line(x - mark, y, x, y)
        c.line(x, y - mark, x, y)
        c.line(x + card_w, y, x + card_w + mark, y)
        c.line(x + card_w, y - mark, x + card_w, y)

    c.save()
    buf.seek(0)
    return buf


# ── PDF with printed sheet (multiple cards per A4) ────────────────────────────
@app.route('/api/generate/sheet', methods=['POST'])
@admin_required
def generate_sheet():
    """Same as bulk PDF but returns single A4 sheet layout."""
    return bulk_generate()

@app.route('/api/generate/batch-pdf', methods=['POST'])
@admin_required
def generate_batch_pdf():
    # [NEW] Backward-compatible alias used by the enhanced Generate tab.
    return bulk_generate_response(force_fmt='pdf')


# ══════════════════════════════════════════════════════════════════════════════
#  EXPORT / ORDERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/export/csv', methods=['GET'])
@admin_required
def export_csv():
    """
    Dynamic CSV — only includes columns that actually have data.
    Standard fields come first, then any custom extra_fields from the form builder.
    """
    school_id  = request.args.get('school_id','')
    class_name = request.args.get('class_name','').strip()
    section    = request.args.get('section','').strip()
    subs = db_list_submissions(school_id=school_id)  # SQL-filtered, no full-table RAM load
    if class_name:
        subs = [s for s in subs if (s.get('class_dept') or '').split('-')[0].strip().lower() == class_name.lower()]
    if section:
        subs = [s for s in subs if len((s.get('class_dept') or '').split('-',1)) > 1 and (s.get('class_dept') or '').split('-',1)[1].strip().lower() == section.lower()]
    if not subs:
        return send_file(io.BytesIO(b'No data'), mimetype='text/csv', download_name='submissions.csv')

    # Standard fields — only include if ANY submission has a non-empty value
    standard_fields = [
        ('Name',        'name'),
        ('Roll / ID No','roll_no'),
        ('Class/Dept',  'class_dept'),
        ('Date of Birth','dob'),
        ('Blood Group', 'blood'),
        ('Phone',       'phone'),
        ('Father/Guardian','father'),
        ('Address',     'address'),
        ('Email',       'email'),
        ('Valid Until', 'validity'),
    ]
    # Filter out roll_no if it's always the same as name (no real ID field)
    active_standard = []
    for label, key in standard_fields:
        vals = [s.get(key,'') for s in subs]
        if key == 'roll_no':
            # Only include if there's a roll_no that is genuinely different from the name
            has_real_id = any(v and v != s.get('name','') for v, s in zip(vals, subs))
            if not has_real_id:
                continue
        if any(v for v in vals):
            active_standard.append((label, key))

    # Extra dynamic fields — collect all unique keys across all submissions
    extra_keys = []
    seen = set()
    for s in subs:
        for k in (s.get('extra_fields') or {}).keys():
            if k not in seen:
                seen.add(k)
                extra_keys.append(k)

    # Build headers
    headers = [label for label, _ in active_standard] + extra_keys + ['Status', 'Submitted']

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for s in subs:
        row = [s.get(key, '') for _, key in active_standard]
        row += [(s.get('extra_fields') or {}).get(k, '') for k in extra_keys]
        row += [s.get('status',''), s.get('submitted_at','')]
        w.writerow(row)

    buf.seek(0)
    sc = db_get_school(school_id) or {}
    fname = f"{sc.get('name','submissions').replace(' ','_')}_data.csv"
    return send_file(
        io.BytesIO(buf.getvalue().encode()),
        mimetype='text/csv',
        download_name=fname
    )

@app.route('/api/orders', methods=['GET'])
@admin_required
def get_orders():
    return jsonify(db_list_orders())

@app.route('/api/orders', methods=['POST'])
@admin_required
def add_order():
    data = request.json or {}
    school = db_get_school(clean_text(data.get('school_id', ''), 60)) or {}
    qty = parse_int(data.get('qty'), 0, 0, 1000000)
    price_per = parse_float(data.get('price_per'), 25, 0, 100000)
    order = {
        "id": gen_id("ord_"),
        "school_id":    clean_text(data.get('school_id',''), 60),
        "client_name":  clean_text(school.get('name', data.get('client_name','')), 160),
        "qty":          qty,
        "price_per":    price_per,
        "total":        qty * price_per,
        "status":       clean_text(data.get('status','pending'), 40),
        "notes":        clean_text(data.get('notes',''), 1000),
        "date":         datetime.now().isoformat()
    }
    db_insert_order(order)
    return jsonify(order), 201

@app.route('/api/orders/<order_id>', methods=['PATCH'])
@admin_required
def update_order(order_id):
    order = db_get_order(order_id)
    if not order:
        return jsonify({"error": "Not found"}), 404
    data = request.json or {}
    for k in ['status','notes']:
        if k in data:
            order[k] = clean_text(data[k], 1000 if k == 'notes' else 40)
    db_update_order(order)
    return jsonify(order)

@app.route('/api/orders/<order_id>', methods=['DELETE'])
@admin_required
def delete_order(order_id):
    order = db_get_order(order_id)
    if order:
        db_insert_deleted_item(deleted_record('order', order))
    db_delete_order(order_id)
    return jsonify({"ok": True})

@app.route('/api/stats', methods=['GET'])
@admin_required
def get_stats():
    now = time.time()
    with _stats_cache_lock:
        if _stats_cache['data'] is not None and now < _stats_cache['expires']:
            return jsonify(_stats_cache['data'])
    payload = db_stats_payload()
    payload["revenue_this_month"] = payload.get("revenue", 0)
    with _stats_cache_lock:
        _stats_cache['data'] = payload
        _stats_cache['expires'] = time.time() + _STATS_CACHE_TTL
    return jsonify(payload)

@app.route('/api/analytics/daily', methods=['GET'])
@admin_required
def analytics_daily():
    # Submission trend for the last N days — cached to avoid repeated DB hits.
    days = parse_int(request.args.get('days'), 30, 1, 365)
    now = time.time()
    with _analytics_daily_cache_lock:
        cached = _analytics_daily_cache.get(days)
        if cached and now < cached['expires']:
            return jsonify(cached['data'])
    data = db_analytics_daily_counts(days)
    with _analytics_daily_cache_lock:
        _analytics_daily_cache[days] = {'data': data, 'expires': time.time() + _ANALYTICS_CACHE_TTL}
    return jsonify(data)

@app.route('/api/analytics/by-school', methods=['GET'])
@admin_required
def analytics_by_school():
    # [NEW] Per-school analytics rollup.
    return jsonify(db_analytics_by_school_rows())

@app.route('/api/export/analytics-csv', methods=['GET'])
@admin_required
def export_analytics_csv():
    # [NEW] CSV report for analytics tab.
    days = parse_int(request.args.get('days'), 30, 1, 365)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['Date', 'School', 'Submissions', 'Photos', 'Completion%'])
    for day, school_name, submissions, photos in db_analytics_csv_rows(days):
        submissions = int(submissions or 0)
        photos = int(photos or 0)
        pct = round((photos / submissions) * 100, 1) if submissions else 0
        writer.writerow([day, school_name or '', submissions, photos, pct])
    buf.seek(0)
    fname = f"analytics_{date.today().isoformat()}.csv"
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8-sig')), mimetype='text/csv', download_name=fname, as_attachment=True)

@app.route('/api/share-links', methods=['GET'])
@admin_required
def get_share_links():
    base_url = request.args.get('base_url', request.host_url.rstrip('/'))
    schools = db_list_schools(public_only=False)  # already includes submission_count via GROUP BY
    links = []
    for school in schools:
        sid = school.get('id')
        links.append({
            "school_id": sid,
            "name": school.get('name', ''),
            "type": school.get('type', 'School'),
            "color": school.get('color', '#1a73e8'),
            "url": f"{base_url}/?school={sid}",
            "submission_count": school.get('submission_count', 0),  # already in payload
            "schema_fields": len(school.get('form_schema', []))
        })
    return jsonify(links)

@app.route('/api/settings', methods=['GET'])
@admin_required
def get_settings():
    return jsonify(db_get_settings())

@app.route('/api/settings', methods=['PUT'])
@admin_required
def save_settings():
    settings = db_get_settings()
    data = request.json or {}
    for key in ['biz_name', 'owner', 'phone', 'email']:
        if key in data:
            settings[key] = clean_text(data[key], 200)
    if 'price_per_card' in data:
        settings['price_per_card'] = parse_float(data.get('price_per_card'), settings.get('price_per_card', 25), 0, 100000)
    db_update_settings(settings)
    return jsonify(settings)

@app.route('/api/backups', methods=['GET'])
@admin_required
def list_backups():
    files = []
    for name in sorted(os.listdir(BACKUPS_DIR), reverse=True):
        if not name.endswith('.json'):
            continue
        path = os.path.join(BACKUPS_DIR, name)
        meta = {}
        try:
            with open(path, 'r', encoding='utf-8') as f:
                meta = backup_metadata(json.load(f))
        except (OSError, json.JSONDecodeError):
            meta = {"backup_type": "unknown", "scope": ""}
        files.append({
            "name": name,
            "size": os.path.getsize(path),
            "created_at": datetime.fromtimestamp(os.path.getmtime(path)).isoformat(),
            **meta
        })
    return jsonify(files)

@app.route('/api/backups', methods=['POST'])
@admin_required
def make_backup():
    data = request.json or {}
    scope = clean_text(data.get('scope') or 'all', 40)
    try:
        if scope == 'all':
            name, _path = create_backup('manual')
        else:
            payload, label = scoped_backup_payload(
                load_db(),
                scope,
                clean_text(data.get('school_id'), 80),
                data.get('submission_ids') if isinstance(data.get('submission_ids'), list) else []
            )
            name, _path = write_backup_payload(label, payload)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    audit('backup.create', name)
    return jsonify({"ok": True, "name": name}), 201

@app.route('/api/backups/<name>', methods=['GET'])
@admin_required
def download_backup(name):
    safe = secure_filename(name)
    path = os.path.join(BACKUPS_DIR, safe)
    if not safe.endswith('.json') or not os.path.exists(path):
        return jsonify({"error": "Backup not found"}), 404
    return send_from_directory(BACKUPS_DIR, safe, as_attachment=True)

@app.route('/api/backups/restore', methods=['POST'])
@admin_required
def restore_backup():
    create_backup('before_restore')
    backup_name = clean_text((request.json or {}).get('name') if request.is_json else '', 120)
    uploaded = request.files.get('file')
    try:
        if uploaded and uploaded.filename:
            validate_upload(uploaded, {'.json'}, set(), MAX_DOC_BYTES)
            payload = json.load(uploaded.stream)
        else:
            safe = secure_filename(backup_name)
            path = os.path.join(BACKUPS_DIR, safe)
            if not safe.endswith('.json') or not os.path.exists(path):
                return jsonify({"error": "Backup not found"}), 404
            with open(path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
    except (ValueError, json.JSONDecodeError, OSError) as e:
        return jsonify({"error": f"Could not restore backup: {e}"}), 400
    if isinstance(payload, dict) and payload.get('backup_type') == 'screenplant_scope_backup':
        try:
            db = restore_scoped_backup(load_db(), payload)
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
    else:
        db = ensure_db_shape(payload)
    save_db(db)
    audit('backup.restore', backup_name or 'uploaded_file')
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
#  PHOTO INJECTOR ENGINE
#  Tariq's core workflow: CorelDraw PDF (text filled, empty photo boxes)
#  + student photos from DB → final print-ready PDF with photos injected
# ══════════════════════════════════════════════════════════════════════════════

def detect_cards_on_sheet(sheet_img):
    """
    Auto-detect individual ID cards on a printed sheet.
    Finds the repeating card grid by looking for crop marks or
    regular rectangular boundaries.
    Returns: list of (x, y, w, h) tuples for each card in pixels.
    """
    W, H = sheet_img.size
    if not NUMPY_OK:
        return _fallback_grid(W, H)
    gray = sheet_img.convert('L')

    arr = np.array(gray)

    # Strategy 1: Look for crop marks — thin lines at regular intervals
    # near the edges of the image (crop marks are usually within 8px of card corners)
    # Strategy 2: Find large white/light rectangular regions separated by dark borders

    # We use edge detection: find rows/cols that are mostly dark (card borders)
    row_mean = arr.mean(axis=1)   # mean brightness per row
    col_mean = arr.mean(axis=0)   # mean brightness per col

    # Find "dark" dividing lines between cards (brightness < 180)
    dark_rows = np.where(row_mean < 160)[0]
    dark_cols = np.where(col_mean < 160)[0]

    # Cluster consecutive dark rows/cols into single dividers
    def cluster(indices, gap=10):
        if len(indices) == 0:
            return []
        clusters = []
        start = indices[0]
        prev  = indices[0]
        for i in indices[1:]:
            if i - prev > gap:
                clusters.append((start + prev) // 2)
                start = i
            prev = i
        clusters.append((start + prev) // 2)
        return clusters

    h_dividers = [0] + cluster(dark_rows) + [H]
    v_dividers = [0] + cluster(dark_cols) + [W]

    # Build card bounding boxes from the grid formed by dividers
    cards = []
    min_card_w = W * 0.10  # cards must be at least 10% of sheet width
    min_card_h = H * 0.08

    for i in range(len(h_dividers) - 1):
        for j in range(len(v_dividers) - 1):
            y1, y2 = h_dividers[i], h_dividers[i+1]
            x1, x2 = v_dividers[j], v_dividers[j+1]
            cw, ch = x2 - x1, y2 - y1
            if cw >= min_card_w and ch >= min_card_h:
                # Add small inset to skip border pixels
                pad = 4
                cards.append((int(x1+pad), int(y1+pad), int(cw-pad*2), int(ch-pad*2)))

    # If detection failed, fall back to uniform grid estimation
    if len(cards) < 1:
        cards = _fallback_grid(W, H)

    return cards


def _fallback_grid(W, H, card_w_mm=85.6, card_h_mm=54, sheet_w_mm=210, sheet_h_mm=297):
    """Fallback: assume standard A4 with standard ID cards, 10mm margin."""
    scale = W / (sheet_w_mm * 3.7795)  # px per mm at screen res
    cw = int(card_w_mm * scale * 3.7795)
    ch = int(card_h_mm * scale * 3.7795)
    margin = int(10 * scale * 3.7795)
    gap    = int(3  * scale * 3.7795)
    cols = max(1, (W - margin*2 + gap) // (cw + gap))
    rows = max(1, (H - margin*2 + gap) // (ch + gap))
    cards = []
    for r in range(rows):
        for c in range(cols):
            x = margin + c * (cw + gap)
            y = margin + r * (ch + gap)
            if x + cw <= W and y + ch <= H:
                cards.append((x, y, cw, ch))
    return cards


def detect_photo_box(card_img):
    """
    Find the empty photo placeholder inside a single card image.
    Looks for: a light-colored (white/near-white) rectangle
    surrounded by darker design elements — that's the photo box.
    Returns: (x, y, w, h) relative to card, or None.
    """
    W, H = card_img.size
    if not NUMPY_OK:
        box_w, box_h = int(W * 0.30), int(H * 0.55)
        return (W - box_w - int(W*0.04), int(H*0.08), box_w, box_h)
    gray = card_img.convert('L')

    arr = np.array(gray)

    # Find the brightest rectangular region (empty photo box is white/light)
    best = None
    best_score = -1

    # Scan candidate boxes using a sliding window approach
    # Check boxes that are roughly photo-sized: 25-50% of card width, 40-70% of card height
    min_w = int(W * 0.18)
    max_w = int(W * 0.55)
    min_h = int(H * 0.35)
    max_h = int(H * 0.85)

    step = max(4, W // 30)

    for bx in range(2, W - min_w, step):
        for by in range(2, H - min_h, step):
            for bw in range(min_w, min(max_w, W - bx), step*2):
                for bh in range(min_h, min(max_h, H - by), step*2):
                    region = arr[by:by+bh, bx:bx+bw]
                    brightness = region.mean()
                    # Also check that the area AROUND this box is darker
                    border_brightness = _border_brightness(arr, bx, by, bw, bh, W, H)
                    contrast = brightness - border_brightness
                    score = brightness * 0.4 + contrast * 0.6
                    if score > best_score:
                        best_score = score
                        best = (bx, by, bw, bh)

    # Sanity check: minimum brightness threshold
    if best and best_score > 50:
        return best

    # Last resort: assume photo box is top-right or top-left (common layouts)
    box_w = int(W * 0.30)
    box_h = int(H * 0.55)
    # Try top-right
    return (W - box_w - int(W*0.04), int(H*0.08), box_w, box_h)


def _border_brightness(arr, bx, by, bw, bh, W, H, pad=8):
    """Average brightness of the border area around a rectangle."""
    regions = []
    if by - pad >= 0:
        regions.append(arr[max(0,by-pad):by, bx:bx+bw])
    if by + bh + pad <= H:
        regions.append(arr[by+bh:min(H,by+bh+pad), bx:bx+bw])
    if bx - pad >= 0:
        regions.append(arr[by:by+bh, max(0,bx-pad):bx])
    if bx + bw + pad <= W:
        regions.append(arr[by:by+bh, bx+bw:min(W,bx+bw+pad)])
    if not regions:
        return 128
    return float(np.concatenate([r.flatten() for r in regions]).mean())



def inject_photos_into_sheet(sheet_img, card_boxes, photo_box_rel, photo_map, card_order):
    """
    Core injection function.
    sheet_img:     PIL Image of the full CorelDraw sheet
    card_boxes:    list of (x,y,w,h) for each card on the sheet
    photo_box_rel: (rx,ry,rw,rh) — photo box position RELATIVE to card (0..1 floats)
    photo_map:     dict of index→photo_path (or roll_no→photo_path)
    card_order:    list of roll_nos or indices matching card_boxes order
    Returns: PIL Image with photos injected
    """
    result = sheet_img.copy().convert('RGBA')

    for idx, (cx, cy, cw, ch) in enumerate(card_boxes):
        # Find photo for this card
        photo_path = None
        if idx < len(card_order):
            key = card_order[idx]
            photo_path = photo_map.get(str(key)) or photo_map.get(key)

        if not photo_path or not os.path.exists(photo_path):
            continue

        # Calculate absolute photo box position on sheet
        rx, ry, rw, rh = photo_box_rel
        px = cx + int(rx * cw)
        py = cy + int(ry * ch)
        pw = int(rw * cw)
        ph = int(rh * ch)

        # Load and resize photo
        try:
            photo = Image.open(photo_path).convert('RGBA')
            # Smart crop: crop to same aspect ratio first, then resize
            photo = _smart_crop(photo, pw, ph)
            photo = photo.resize((pw, ph), Image.LANCZOS)

            # Paste photo onto sheet
            result.paste(photo, (px, py), photo)
        except Exception as e:
            logger.warning("Photo inject error for card %d: %s", idx, e)
            continue

    return result.convert('RGB')


def _smart_crop(img, target_w, target_h):
    """Crop image to target aspect ratio, centered."""
    iw, ih = img.size
    target_ratio = target_w / target_h
    current_ratio = iw / ih

    if current_ratio > target_ratio:
        # Too wide — crop sides
        new_w = int(ih * target_ratio)
        left = (iw - new_w) // 2
        img = img.crop((left, 0, left + new_w, ih))
    else:
        # Too tall — crop top/bottom (keep top — passport photos are top-heavy)
        new_h = int(iw / target_ratio)
        img = img.crop((0, 0, iw, new_h))
    return img


ID_FIELD_HINTS = (
    'unique', 'uid', 'id no', 'id_no', 'roll', 'roll no', 'roll_no',
    'admission', 'admission no', 'reg', 'register', 'registration',
    'student id', 'employee id', 'emp id', 'person id', 'code', 'sr no'
)
NAME_FIELD_HINTS = (
    'name', 'student name', 'full name', 'person name', 'employee name',
    'first name', 'last name', 'surname'
)


def match_key(value):
    """Normalize names/IDs so Excel rows and photo filenames match reliably."""
    value = '' if value is None else str(value)
    value = os.path.splitext(value.strip().lower())[0]
    value = re.sub(r'[_\-]+', ' ', value)
    value = re.sub(r'[^a-z0-9]+', '', value)
    return value


def split_filename_tokens(filename):
    base = os.path.splitext(os.path.basename(filename))[0].lower()
    parts = re.split(r'[\s_\-().\[\]]+', base)
    return [match_key(p) for p in parts if match_key(p)]


def label_has_hint(label, hints):
    label = f" {str(label).strip().lower().replace('_', ' ')} "
    for hint in hints:
        h = f" {hint.strip().lower().replace('_', ' ')} "
        if h in label:
            return True
    return False


def best_row_name(row):
    name_values = []
    for key, value in row.items():
        if value in (None, ''):
            continue
        if label_has_hint(key, NAME_FIELD_HINTS):
            name_values.append(str(value).strip())
    if name_values:
        return ' '.join(name_values)
    return next((str(v).strip() for k, v in row.items() if v and label_has_hint(k, ('name',))), '')


def row_aliases(row):
    aliases = set()
    name_parts = []
    for key, value in row.items():
        if value in (None, ''):
            continue
        if label_has_hint(key, ID_FIELD_HINTS) or str(key).strip().lower() in {'id', 'uid'}:
            mk = match_key(value)
            if mk:
                aliases.add(mk)
        if label_has_hint(key, NAME_FIELD_HINTS):
            name_parts.append(str(value).strip())
            mk = match_key(value)
            if mk:
                aliases.add(mk)
    if name_parts:
        full_name = match_key(' '.join(name_parts))
        if full_name:
            aliases.add(full_name)
    if not aliases:
        for value in row.values():
            mk = match_key(value)
            if mk:
                aliases.add(mk)
    return aliases


def choose_best_photo(aliases, photo_tokens):
    """
    Pick the safest matching photo for a person row.
    Exact normalized ID/name matches win. Short loose matches are rejected to
    avoid accidentally matching "1" to many filenames.
    """
    best = None
    best_score = 0.0
    best_reason = ''
    aliases = {a for a in aliases if a}
    for alias in aliases:
        for base_key, tokens, path, original in photo_tokens:
            if not base_key:
                continue
            score = 0.0
            reason = ''
            if alias == base_key:
                score, reason = 1.0, 'exact filename'
            elif len(alias) >= 4 and alias in tokens:
                score, reason = 0.96, 'filename token'
            elif len(alias) >= 5 and (alias in base_key or base_key in alias):
                overlap = min(len(alias), len(base_key)) / max(len(alias), len(base_key))
                score, reason = 0.88 + (0.08 * overlap), 'contains'
            elif len(alias) >= 5:
                ratio = difflib.SequenceMatcher(None, alias, base_key).ratio()
                if ratio >= 0.86:
                    score, reason = ratio, 'fuzzy'
            if score > best_score:
                best = (path, original)
                best_score = score
                best_reason = reason
    if not best:
        return None, 0.0, ''
    return best, best_score, best_reason


def submission_aliases(sub):
    aliases = set()
    for key in ['unique_id', 'roll_no', 'name', 'phone', 'email']:
        mk = match_key(sub.get(key, ''))
        if mk:
            aliases.add(mk)
    for value in (sub.get('extra_fields') or {}).values():
        mk = match_key(value)
        if mk:
            aliases.add(mk)
    return aliases


def extract_card_texts_from_pdf(pdf_bytes, page_num, dpi, card_boxes):
    """Return normalized text found inside each detected card box."""
    if not PYMUPDF_OK:
        return ['' for _ in card_boxes]
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        page = doc[page_num]
        scale = dpi / 72
        words = page.get_text("words") or []
        card_texts = []
        for cx, cy, cw, ch in card_boxes:
            parts = []
            for x0, y0, x1, y1, word, *_rest in words:
                wx = ((x0 + x1) / 2) * scale
                wy = ((y0 + y1) / 2) * scale
                if cx <= wx <= cx + cw and cy <= wy <= cy + ch:
                    parts.append(str(word))
            card_texts.append(' '.join(parts))
        return card_texts
    finally:
        doc.close()


def auto_order_from_card_text(card_texts, aliases_by_key, fallback_order):
    """
    Match each card to the person whose alias appears in that card's printed text.
    Falls back to sheet order for low-confidence or image-only sheets.
    """
    resolved = []
    used = set()
    hits = 0
    details = []

    for idx, text in enumerate(card_texts):
        normalized_text = match_key(text)
        best_key = None
        best_score = 0
        best_alias = ''

        if normalized_text:
            for key, aliases in aliases_by_key.items():
                if key in used:
                    continue
                for alias in aliases:
                    if len(alias) < 3:
                        continue
                    score = 0
                    if alias in normalized_text:
                        score = 100 + len(alias)
                    elif len(alias) >= 6:
                        ratio = difflib.SequenceMatcher(None, alias, normalized_text).ratio()
                        if ratio >= 0.72:
                            score = int(ratio * 80)
                    if score > best_score:
                        best_key, best_score, best_alias = key, score, alias

        if best_key and best_score >= 80:
            resolved.append(best_key)
            used.add(best_key)
            hits += 1
            details.append({"card": idx + 1, "matched": True, "key": best_key, "alias": best_alias})
        else:
            fallback = fallback_order[idx] if idx < len(fallback_order) else ''
            resolved.append(fallback)
            if fallback:
                used.add(fallback)
            details.append({"card": idx + 1, "matched": False, "key": fallback, "alias": ""})

    return resolved, {"cards": len(card_texts), "text_matched": hits, "details": details[:80]}


def parse_external_rows(file_storage):
    if not file_storage:
        return []
    if file_size(file_storage) > MAX_DOC_BYTES:
        raise RuntimeError("Data file is too large")
    raw = file_storage.read()
    name = file_storage.filename or ''
    ext = os.path.splitext(name)[1].lower()
    rows = []

    if ext in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        if not OPENPYXL_OK:
            raise RuntimeError("Excel import requires openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        headers = [clean_text(h or f"Column_{i+1}", 120) for i, h in enumerate(values[0])]
        for vals in values[1:]:
            if not vals or not any(v not in (None, '') for v in vals):
                continue
            rows.append({headers[i]: vals[i] if i < len(vals) else '' for i in range(len(headers))})
        return rows

    if ext == '.csv':
        text = raw.decode('utf-8-sig', errors='replace')
        return [dict(r) for r in csv.DictReader(io.StringIO(text)) if any((v or '').strip() for v in r.values())]

    raise RuntimeError("Upload .xlsx or .csv data")


def collect_external_photos(request_files, temp_dir):
    photos = []

    zip_file = request_files.get('photos_zip')
    if zip_file and zip_file.filename:
        raw = zip_file.read()
        validate_zip_bytes(raw)
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                ext = os.path.splitext(info.filename)[1].lower()
                if ext not in IMAGE_EXTS:
                    continue
                if info.file_size > MAX_IMAGE_BYTES:
                    continue
                safe_name = secure_filename(os.path.basename(info.filename))
                if not safe_name:
                    continue
                out_path = unique_temp_path(temp_dir, safe_name)
                with zf.open(info) as src, open(out_path, 'wb') as dst:
                    dst.write(src.read())
                photos.append((info.filename, out_path))

    for fs in request_files.getlist('photos'):
        if not fs or not fs.filename:
            continue
        ext = os.path.splitext(fs.filename)[1].lower()
        if ext not in IMAGE_EXTS:
            continue
        try:
            validate_upload(fs, IMAGE_EXTS, IMAGE_MIMES, MAX_IMAGE_BYTES)
        except ValueError:
            continue
        safe_name = secure_filename(os.path.basename(fs.filename))
        if not safe_name:
            continue
        out_path = unique_temp_path(temp_dir, safe_name)
        fs.save(out_path)
        photos.append((fs.filename, out_path))

    return photos


def unique_temp_path(folder, filename):
    stem, ext = os.path.splitext(filename)
    path = os.path.join(folder, filename)
    if not os.path.exists(path):
        return path
    for i in range(2, 1000):
        candidate = os.path.join(folder, f"{stem}_{i}{ext}")
        if not os.path.exists(candidate):
            return candidate
    return os.path.join(folder, f"{stem}_{gen_id()}{ext}")


def build_external_injection_source(data_file, request_files):
    rows = parse_external_rows(data_file)
    if not rows:
        raise RuntimeError("No people found in the uploaded data sheet")

    temp_dir = tempfile.mkdtemp(prefix='screenplant_inject_')
    photos = collect_external_photos(request_files, temp_dir)
    if not photos:
        raise RuntimeError("Upload a photos ZIP or select a photo folder")

    photo_exact = {}
    photo_tokens = []
    for original, path in photos:
        base_key = match_key(os.path.basename(original))
        if base_key:
            photo_exact[base_key] = path
        tokens = split_filename_tokens(original)
        for token in tokens:
            photo_exact.setdefault(token, path)
        photo_tokens.append((base_key, tokens, path, original))

    photo_map = {}
    card_order = []
    aliases_by_order = {}
    matched = 0
    fuzzy = 0
    missing = []
    match_details = []

    all_photo_keys = [k for k in photo_exact.keys() if k]
    for idx, row in enumerate(rows):
        aliases = row_aliases(row)
        order_key = f"row_{idx}"
        card_order.append(order_key)
        aliases_by_order[order_key] = aliases
        photo_path = None

        for alias in aliases:
            if alias in photo_exact:
                photo_path = photo_exact[alias]
                break

        if not photo_path:
            for alias in aliases:
                for base_key, tokens, path, _original in photo_tokens:
                    if alias and (alias in base_key or base_key in alias or alias in tokens):
                        photo_path = path
                        break
                if photo_path:
                    break

        if not photo_path and all_photo_keys:
            best = None
            best_score = 0.0
            for alias in aliases:
                close = difflib.get_close_matches(alias, all_photo_keys, n=1, cutoff=0.86)
                if close:
                    score = difflib.SequenceMatcher(None, alias, close[0]).ratio()
                    if score > best_score:
                        best = close[0]
                        best_score = score
            if best:
                photo_path = photo_exact[best]
                fuzzy += 1

        if photo_path:
            photo_map[order_key] = photo_path
            matched += 1
            match_details.append({
                "row": idx + 2,
                "name": best_row_name(row) or order_key,
                "matched": True,
                "photo": os.path.basename(photo_path),
                "aliases": sorted(list(aliases))[:8]
            })
        else:
            name = best_row_name(row)
            missing.append(name or f"Row {idx + 2}")
            match_details.append({
                "row": idx + 2,
                "name": name or f"Row {idx + 2}",
                "matched": False,
                "photo": "",
                "aliases": sorted(list(aliases))[:8]
            })

    return {
        "rows": rows,
        "photo_map": photo_map,
        "card_order": card_order,
        "aliases_by_order": aliases_by_order,
        "temp_dir": temp_dir,
        "report": {
            "rows": len(rows),
            "photos": len(photos),
            "matched": matched,
            "fuzzy": fuzzy,
            "missing": missing[:30],
            "details": match_details[:80]
        }
    }


@app.route('/api/injector/match-preview', methods=['POST'])
@admin_required
def injector_match_preview():
    try:
        source = build_external_injection_source(request.files.get('data_file'), request.files)
        report = source["report"]
        temp_dir = source.get("temp_dir")
        if temp_dir and os.path.isdir(temp_dir):
            shutil.rmtree(temp_dir, ignore_errors=True)
        return jsonify({"ok": True, "report": report})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


def pdf_page_to_image(pdf_bytes, page_num=0, dpi=200):
    """Convert a PDF page to PIL Image using PyMuPDF."""
    if not PYMUPDF_OK:
        raise RuntimeError("PyMuPDF (fitz) not installed. Run: pip install pymupdf")
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    page = doc[page_num]
    mat = fitz.Matrix(dpi/72, dpi/72)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    return img, len(doc)


def images_to_pdf(images, output_path=None):
    """Convert list of PIL Images to a single PDF."""
    buf = io.BytesIO()
    if len(images) == 1:
        images[0].convert('RGB').save(buf, format='PDF', resolution=200)
    else:
        images[0].convert('RGB').save(
            buf, format='PDF', save_all=True,
            append_images=[i.convert('RGB') for i in images[1:]],
            resolution=200
        )
    buf.seek(0)
    return buf


# ── API: Analyse uploaded sheet (detect cards + photo box) ────────────────────
@app.route('/api/injector/analyse', methods=['POST'])
@admin_required
def injector_analyse():
    """
    Upload one sheet image (PNG/JPG) or PDF page.
    Returns detected card grid and photo box position.
    """
    file = request.files.get('file')
    if not file:
        return jsonify({"error": "No file uploaded"}), 400

    try:
        validate_upload(file, DOC_EXTS, DOC_MIMES, MAX_DOC_BYTES)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    raw = file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    try:
        if ext == '.pdf':
            if not PYMUPDF_OK:
                return jsonify({"error": "PDF support requires: pip install pymupdf"}), 400
            sheet_img, total_pages = pdf_page_to_image(raw, page_num=0)
        else:
            sheet_img = Image.open(io.BytesIO(raw)).convert('RGB')
            total_pages = 1

        W, H = sheet_img.size

        card_boxes = detect_cards_on_sheet(sheet_img)
        num_cards = len(card_boxes)

        # Detect photo box from first card
        photo_box_abs = None
        photo_box_rel = None
        if card_boxes:
            cx, cy, cw, ch = card_boxes[0]
            card_crop = sheet_img.crop((cx, cy, cx+cw, cy+ch))
            pb = detect_photo_box(card_crop)
            if pb:
                bx, by, bw, bh = pb
                photo_box_abs = {"x": bx, "y": by, "w": bw, "h": bh}
                photo_box_rel = {
                    "x": round(bx/cw, 4), "y": round(by/ch, 4),
                    "w": round(bw/cw, 4), "h": round(bh/ch, 4)
                }

        # Save sheet temporarily for preview
        preview_id = gen_id("prev_")
        prev_path = os.path.join(UPLOADS_DIR, preview_id + '.jpg')

        # Draw detected boxes on a preview copy
        preview = sheet_img.copy()
        draw = ImageDraw.Draw(preview)
        for i, (cx, cy, cw, ch) in enumerate(card_boxes):
            draw.rectangle([cx, cy, cx+cw, cy+ch], outline=(0,200,100), width=3)
            if photo_box_rel:
                px = cx + int(photo_box_rel['x'] * cw)
                py = cy + int(photo_box_rel['y'] * ch)
                pw = int(photo_box_rel['w'] * cw)
                ph = int(photo_box_rel['h'] * ch)
                draw.rectangle([px, py, px+pw, py+ph], outline=(255,80,80), width=2)
            draw.text((cx+6, cy+6), str(i+1), fill=(0,200,100))
        preview.save(prev_path, quality=85)

        return jsonify({
            "ok": True,
            "sheet_size": {"w": W, "h": H},
            "total_pages": total_pages,
            "num_cards": num_cards,
            "card_boxes": [{"x":x,"y":y,"w":w,"h":h} for x,y,w,h in card_boxes],
            "photo_box_rel": photo_box_rel,
            "photo_box_abs": photo_box_abs,
            "preview_url": f"/uploads/{preview_id}.jpg"
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── API: Inject photos into CorelDraw PDF ─────────────────────────────────────
@app.route('/api/injector/inject', methods=['POST'])
@admin_required
def injector_inject():
    """
    Multipart form:
      - file:          The CorelDraw PDF (text-filled, empty photo boxes)
      - school_id:     Which school's submissions to use
      - photo_box:     JSON {x,y,w,h} relative (0..1) — from analyse step
      - card_order:    JSON array of roll_nos in sheet order (optional)
      - match_mode:    "position" (default) or "id" (future OCR)
      - dpi:           Render DPI (default 200)
    Returns: print-ready PDF with photos injected
    """
    file        = request.files.get('file')
    source_mode = request.form.get('source_mode', 'saved')
    school_id   = request.form.get('school_id', '')
    pb_json     = request.form.get('photo_box', '{}')
    order_json  = request.form.get('card_order', '[]')
    match_mode  = request.form.get('match_mode', 'auto')
    dpi         = parse_int(request.form.get('dpi', 200), 200, 72, 300)

    if not file:
        return jsonify({"error": "No PDF/image file uploaded"}), 400
    try:
        validate_upload(file, DOC_EXTS, DOC_MIMES, MAX_DOC_BYTES)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    try:
        photo_box_rel = json.loads(pb_json)
        card_order_input = json.loads(order_json)
    except json.JSONDecodeError:
        photo_box_rel = {}
        card_order_input = []

    db = load_db()
    external_report = None
    external_temp_dir = None
    aliases_by_key = {}
    try:
        if source_mode == 'external':
            source = build_external_injection_source(request.files.get('data_file'), request.files)
            photo_map = source["photo_map"]
            card_order_input = card_order_input or source["card_order"]
            aliases_by_key = source.get("aliases_by_order", {})
            external_temp_dir = source.get("temp_dir")
            subs = [
                {
                    "unique_id": source["card_order"][i],
                    "roll_no": source["card_order"][i],
                    "name": best_row_name(row) or source["card_order"][i]
                }
                for i, row in enumerate(source["rows"])
            ]
            external_report = source["report"]
        else:
            subs = [s for s in db['submissions'] if s.get('school_id') == school_id]
            if not subs and not school_id:
                subs = db['submissions']  # fallback: all submissions
            if not subs:
                return jsonify({"error": "No submissions found for this school"}), 404

            # Build photo map: exact and normalized unique_id/name/roll_no fallbacks.
            photo_map = {}
            for sub in subs:
                p = sub.get('photo_path')
                if not p or not os.path.exists(p):
                    continue
                for key in [sub.get('unique_id',''), sub.get('roll_no',''), sub.get('name','')]:
                    k = str(key).strip()
                    nk = match_key(k)
                    if k:
                        photo_map[k] = p
                    if nk:
                        photo_map[nk] = p
            for sub in subs:
                key = match_key(sub.get('unique_id') or sub.get('roll_no') or sub.get('name', ''))
                if key:
                    aliases_by_key[key] = submission_aliases(sub)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    raw = file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    output_images = []

    try:
        if ext == '.pdf':
            if not PYMUPDF_OK:
                return jsonify({"error": "PDF support requires: pip install pymupdf"}), 400
            doc = fitz.open(stream=raw, filetype="pdf")
            total_pages = len(doc)
            doc.close()
        else:
            total_pages = 1

        # Track which submission index we're on across all pages
        sub_offset = 0

        for page_num in range(total_pages):
            if ext == '.pdf':
                sheet_img, _ = pdf_page_to_image(raw, page_num=page_num, dpi=dpi)
            else:
                sheet_img = Image.open(io.BytesIO(raw)).convert('RGB')

            W, H = sheet_img.size
            card_boxes = detect_cards_on_sheet(sheet_img)

            # Use provided photo_box or detect it
            if photo_box_rel and all(k in photo_box_rel for k in ['x','y','w','h']):
                pb = (
                    float(photo_box_rel['x']),
                    float(photo_box_rel['y']),
                    float(photo_box_rel['w']),
                    float(photo_box_rel['h'])
                )
            else:
                # Auto-detect from first card of this page
                if card_boxes:
                    cx, cy, cw, ch = card_boxes[0]
                    card_crop = sheet_img.crop((cx, cy, cx+cw, cy+ch))
                    detected = detect_photo_box(card_crop)
                    if detected:
                        bx, by, bw, bh = detected
                        pb = (bx/cw, by/ch, bw/cw, bh/ch)
                    else:
                        pb = (0.62, 0.08, 0.30, 0.55)  # safe default
                else:
                    pb = (0.62, 0.08, 0.30, 0.55)

            # Build card order for this page: auto text-match first, then positional fallback.
            if card_order_input:
                page_order = card_order_input[sub_offset:sub_offset+len(card_boxes)]
                if source_mode != 'external':
                    page_order = [match_key(x) for x in page_order]
            else:
                # Positional: match subs in order they appear in DB
                page_order = [
                    match_key(subs[sub_offset + i].get('unique_id') or subs[sub_offset + i].get('roll_no') or subs[sub_offset + i].get('name', ''))
                    for i in range(min(len(card_boxes), len(subs) - sub_offset))
                ]

            if match_mode == 'auto' and ext == '.pdf' and aliases_by_key:
                card_texts = extract_card_texts_from_pdf(raw, page_num, dpi, card_boxes)
                page_order, text_report = auto_order_from_card_text(card_texts, aliases_by_key, page_order)
                if external_report is not None:
                    external_report.setdefault("text_match", {"cards": 0, "text_matched": 0})
                    external_report["text_match"]["cards"] += text_report["cards"]
                    external_report["text_match"]["text_matched"] += text_report["text_matched"]

            injected = inject_photos_into_sheet(sheet_img, card_boxes, pb, photo_map, page_order)
            output_images.append(injected)
            sub_offset += len(card_boxes)

        # Output as PDF
        out_buf = images_to_pdf(output_images)
        if external_temp_dir and os.path.isdir(external_temp_dir):
            shutil.rmtree(external_temp_dir, ignore_errors=True)
        school = next((s for s in db['schools'] if s['id'] == school_id), {})
        label = school.get('name') if source_mode != 'external' else 'external_data'
        fname = f"{label or 'cards'}_with_photos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        response = send_file(out_buf, mimetype='application/pdf', download_name=fname)
        if external_report:
            response.headers['X-Injection-Report'] = json.dumps(external_report)
            response.headers['Access-Control-Expose-Headers'] = 'X-Injection-Report'
        return response

    except Exception as e:
        if external_temp_dir and os.path.isdir(external_temp_dir):
            shutil.rmtree(external_temp_dir, ignore_errors=True)
        logger.exception("injector_inject failed: %s", e)
        return jsonify({"error": str(e)}), 500


# ── API: Export Excel for CorelDraw mail merge ────────────────────────────────

def unique_zip_name(used, name):
    base, ext = os.path.splitext(name)
    candidate = name
    i = 2
    while candidate.lower() in used:
        candidate = f"{base}_{i}{ext}"
        i += 1
    used.add(candidate.lower())
    return candidate


@app.route('/api/export/excel', methods=['GET'])
@admin_required
def export_excel():
    """
    Dynamic Excel export — only includes columns that actually have data.
    Standard fields first, then custom extra_fields from the form builder.
    CorelDraw mail merge reads column headers as field names.
    """
    if not OPENPYXL_OK:
        return jsonify({"error": "Excel export requires openpyxl. Install it with: pip install openpyxl"}), 503
    school_id  = request.args.get('school_id', '')
    class_name = request.args.get('class_name', '').strip()
    section    = request.args.get('section', '').strip()
    subs = db_list_submissions(school_id=school_id)  # SQL-filtered
    if class_name:
        subs = [s for s in subs if (s.get('class_dept') or '').split('-')[0].strip().lower() == class_name.lower()]
    if section:
        subs = [s for s in subs if len((s.get('class_dept') or '').split('-',1)) > 1 and (s.get('class_dept') or '').split('-',1)[1].strip().lower() == section.lower()]
    if not subs:
        return jsonify({"error": "No submissions found"}), 404

    # Sort by submission time so Excel row order matches photo sequence numbers
    subs = sorted(subs, key=lambda x: x.get('submitted_at', ''))

    school = db_get_school(school_id) or {}

    # Standard fields — only include if any submission has a real value
    standard_fields = [
        ('Name',            'name'),
        ('Roll_No',         'roll_no'),
        ('Class_Dept',      'class_dept'),
        ('DOB',             'dob'),
        ('Blood_Group',     'blood'),
        ('Phone',           'phone'),
        ('Father_Name',     'father'),
        ('Address',         'address'),
        ('Email',           'email'),
        ('Valid_Until',     'validity'),
    ]
    active_standard = []
    for label, key in standard_fields:
        vals = [s.get(key, '') for s in subs]
        if key == 'roll_no':
            has_real_id = any(v and v != s.get('name', '') for v, s in zip(vals, subs))
            if not has_real_id:
                continue
        if any(v for v in vals):
            active_standard.append((label, key))

    # Extra dynamic fields from form builder
    extra_keys = []
    seen = set()
    for s in subs:
        for k in (s.get('extra_fields') or {}).keys():
            if k not in seen:
                seen.add(k)
                extra_keys.append(k)

    # Always add School_Name and Photo_File (sequential number matching photos ZIP)
    headers = [label for label, _ in active_standard] + extra_keys + ['School_Name', 'Photo_File']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ID Card Data"

    header_fill  = PatternFill("solid", fgColor="1A1A2E")
    header_font  = Font(color="F5A623", bold=True, name="Calibri", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[cell.column_letter].width = max(14, len(h) + 4)
    ws.row_dimensions[1].height = 22

    row_fill_alt = PatternFill("solid", fgColor="F8F8F8")
    data_font    = Font(name="Calibri", size=10)
    data_align   = Alignment(horizontal="left", vertical="center")

    for row_idx, sub in enumerate(subs, 2):
        seq = row_idx - 1  # 1-based sequence number matching the photos ZIP
        photo_p = sub.get('photo_path')
        photo_ext = os.path.splitext(photo_p)[1] if photo_p else '.jpg'
        photo_filename = f"{seq}{photo_ext}" if photo_p and os.path.exists(photo_p) else ''

        row_data = [sub.get(key, '') for _, key in active_standard]
        row_data += [(sub.get('extra_fields') or {}).get(k, '') for k in extra_keys]
        row_data += [school.get('name', ''), photo_filename]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=str(val) if val else '')
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border
            if row_idx % 2 == 0:
                cell.fill = row_fill_alt
        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A2"

    # Instructions sheet
    ws2 = wb.create_sheet("CorelDraw Instructions")
    ws2['A1'].value = "HOW TO USE THIS FILE WITH CORELDRAW"
    ws2['A1'].font = Font(bold=True, size=13, color="1A73E8")
    ws2.merge_cells('A1:B1')
    steps = [
        ("1", "Open CorelDraw and your ID card template"),
        ("2", "Go to: File → Print Merge → Create/Load Merge Fields"),
        ("3", "Select 'From a file' and browse to this Excel file"),
        ("4", "Map each field: <<Name>> → Name column, etc."),
        ("5", "Click 'Perform Merge' — all cards fill automatically"),
        ("6", "Export/Print to PDF (leave photo boxes empty)"),
        ("7", "Upload that PDF to ScreenPlant → Photo Injector tab"),
        ("8", "ScreenPlant adds all photos automatically → download final PDF"),
    ]
    for r, (step, action) in enumerate(steps, 3):
        ws2.cell(row=r, column=1, value=step).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=action)
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filter_tag = (f"_{class_name.replace(' ','_')}" if class_name else '') + (f"_sec{section.replace(' ','_')}" if section else '')
    fname = f"{school.get('name', 'screenplant').replace(' ','_')}{filter_tag}_merge_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=fname)


@app.route('/api/export/photos-zip', methods=['GET'])
@admin_required
def export_photos_zip():
    """
    Download all student photos for a school, renamed as sequential files.
    Streams the ZIP response so large R2 exports do not accumulate in RAM.
    """
    school_id  = request.args.get('school_id', '')
    class_name = request.args.get('class_name', '').strip()
    section    = request.args.get('section', '').strip()
    approved_only = request.args.get('approved_only') == '1'
    if not school_id and not approved_only:
        return jsonify({"error": "school_id is required"}), 400
    subs = db_list_submissions(school_id=school_id)  # SQL-filtered
    if approved_only:
        subs = [s for s in subs if s.get('photo_status') == 'approved']
    school = db_get_school(school_id) or {} if school_id else {}
    if class_name:
        subs = [s for s in subs if (s.get('class_dept') or '').split('-')[0].strip().lower() == class_name.lower()]
    if section:
        subs = [s for s in subs if len((s.get('class_dept') or '').split('-',1)) > 1 and (s.get('class_dept') or '').split('-',1)[1].strip().lower() == section.lower()]

    subs = sorted(subs, key=lambda x: x.get('submitted_at', ''))
    photo_entries = []
    seq = 1
    for sub in subs:
        p = sub.get('photo_path')
        if not p or (not os.path.exists(p) and not _USE_R2):
            continue
        ext = os.path.splitext(p)[1] or '.jpg'
        photo_entries.append((f"{seq}{ext}", p))
        seq += 1

    if not photo_entries:
        return jsonify({"error": "No photos found"}), 404

    filter_tag = (f"_{class_name.replace(' ','_')}" if class_name else '') + (f"_sec{section.replace(' ','_')}" if section else '')
    fname = f"{school.get('name','approved' if approved_only else 'photos')}{filter_tag}_{datetime.now().strftime('%Y%m%d')}_photos.zip"

    def entries():
        for arcname, path in photo_entries:
            if os.path.exists(path):
                yield arcname, file_bytes_iter(path)
            elif _USE_R2:
                yield arcname, r2_bytes_iter(r2_key_from_path(path))

    return stream_zip_response(entries(), fname)

@app.route('/api/export/school-bundle', methods=['GET'])
@admin_required
def export_school_bundle():
    """
    Download one selected school as a ZIP containing an Excel sheet and photos.
    Photos are streamed into the ZIP one at a time to avoid large RAM spikes.
    """
    if not OPENPYXL_OK:
        return jsonify({"error": "School bundle requires openpyxl. Install it with: pip install openpyxl"}), 503

    school_id  = request.args.get('school_id', '')
    class_name = request.args.get('class_name', '').strip()
    section    = request.args.get('section', '').strip()
    if not school_id:
        return jsonify({"error": "Select a school first"}), 400

    school = db_get_school(school_id)
    if not school:
        return jsonify({"error": "School not found"}), 404

    subs = db_list_submissions(school_id=school_id)  # SQL-filtered
    if class_name:
        subs = [s for s in subs if (s.get('class_dept') or '').split('-')[0].strip().lower() == class_name.lower()]
    if section:
        subs = [s for s in subs if len((s.get('class_dept') or '').split('-',1)) > 1 and (s.get('class_dept') or '').split('-',1)[1].strip().lower() == section.lower()]
    if not subs:
        return jsonify({"error": "No submissions found for this school"}), 404

    subs = sorted(subs, key=lambda x: x.get('submitted_at', ''))
    filter_tag = (f"_{class_name.replace(' ','_')}" if class_name else '') + (f"_sec{section.replace(' ','_')}" if section else '')
    folder = safe_archive_name(school.get('name'), 'School') + filter_tag

    standard_fields = [
        ('Name', 'name'), ('Roll_No', 'roll_no'), ('Class_Dept', 'class_dept'),
        ('DOB', 'dob'), ('Blood_Group', 'blood'), ('Phone', 'phone'),
        ('Father_Name', 'father'), ('Address', 'address'), ('Email', 'email'),
        ('Valid_Until', 'validity'), ('Status', 'status'), ('Submitted', 'submitted_at'),
    ]
    active_standard = []
    for label, key in standard_fields:
        vals = [s.get(key, '') for s in subs]
        if key == 'roll_no':
            has_real_id = any(v and v != s.get('name', '') for v, s in zip(vals, subs))
            if not has_real_id:
                continue
        if any(vals):
            active_standard.append((label, key))

    extra_keys = []
    seen = set()
    for sub in subs:
        for key in (sub.get('extra_fields') or {}).keys():
            if key not in seen:
                seen.add(key)
                extra_keys.append(key)

    photo_names_by_id = {}
    seq = 1
    for sub in subs:
        p = sub.get('photo_path')
        has_photo = (p and os.path.exists(p)) or (_USE_R2 and p)
        if has_photo:
            ext = os.path.splitext(p)[1] or '.jpg'
            photo_names_by_id[sub.get('id')] = f"{seq}{ext}"
            seq += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "School Data"
    headers = [label for label, _ in active_standard] + extra_keys + ['School_Name', 'Photo_File']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = max(14, len(header) + 3)

    for row_idx, sub in enumerate(subs, 2):
        row = [sub.get(key, '') for _, key in active_standard]
        row += [(sub.get('extra_fields') or {}).get(key, '') for key in extra_keys]
        row += [school.get('name', ''), photo_names_by_id.get(sub.get('id'), '')]
        for col, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=str(value) if value else '')
    ws.freeze_panes = "A2"

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    photo_entries = []
    for sub in subs:
        p = sub.get('photo_path')
        photo_name = photo_names_by_id.get(sub.get('id'))
        if p and photo_name and (os.path.exists(p) or _USE_R2):
            photo_entries.append((f"{folder}/photos/{photo_name}", p))

    summary = [
        f"School: {school.get('name', '')}",
        f"Submissions: {len(subs)}",
        f"Photos included: {len(photo_entries)}",
        f"Exported: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "Open data.xlsx for the complete school data sheet.",
        "Photos are inside the photos folder.",
    ]

    fname = f"{folder}_complete_bundle_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"

    def entries():
        yield f"{folder}/data.xlsx", (excel_bytes,)
        for arcname, path in photo_entries:
            if os.path.exists(path):
                yield arcname, file_bytes_iter(path)
            elif _USE_R2:
                yield arcname, r2_bytes_iter(r2_key_from_path(path))
        yield f"{folder}/README.txt", ("\n".join(summary).encode('utf-8'),)

    return stream_zip_response(entries(), fname)


@app.route('/api/export/photos-zip-all', methods=['GET'])
@admin_required
def export_photos_zip_all():
    """
    One ZIP with all schools, each in its own folder.
    Structure:
      ScreenPlant_All_Photos_YYYYMMDD_HHMM.zip
        LTM_School/
          A001_Rayan.jpg
        Rayan_College/
          B001_Student.jpg
        No_School/
          Name_Only.jpg

    For large datasets (>500 photos) prefer per-school export with class filter
    to avoid gunicorn timeout. This endpoint is best for full backups.
    """
    schools = db_list_schools(public_only=False)
    school_map = {s['id']: s for s in schools}
    all_subs = db_list_submissions()  # all submissions via SQL, not full JSON blob load

    # Warn early if the export is very large — client sees this before waiting
    photo_count = sum(1 for s in all_subs if s.get('photo_path') and (os.path.exists(s['photo_path']) or _USE_R2))
    if photo_count == 0:
        return jsonify({"error": "No photos found"}), 404
    if photo_count > int(os.environ.get('SCREENPLANT_MAX_ZIP_PHOTOS', '2000')):
        return jsonify({
            "error": f"Too many photos ({photo_count}) for a single ZIP. "
                     "Use per-school export with class/section filters to download in batches.",
            "photo_count": photo_count
        }), 413

    photo_entries = []
    used_names = set()
    for sub in all_subs:
        p = sub.get('photo_path')
        if not p or (not os.path.exists(p) and not _USE_R2):
            continue
        sc = school_map.get(sub.get('school_id', ''), {})
        folder = re.sub(r'[^\w\s\-]', '', sc.get('name', 'No_School')).strip().replace(' ', '_') or 'No_School'
        ext = os.path.splitext(p)[1] or '.jpg'
        uid = sub.get('unique_id') or sub.get('roll_no') or ''
        nm  = secure_filename(sub.get('name', 'unknown').replace(' ', '_')) or 'unknown'
        safe_uid = re.sub(r'[^\w\-]', '_', str(uid))[:30]
        if safe_uid.lower() == nm.lower().replace(' ', '_'):
            file_name = f"{nm}{ext}"
        else:
            file_name = f"{safe_uid}_{nm}{ext}" if safe_uid else f"{nm}{ext}"
        arcname = unique_zip_name(used_names, f"{folder}/{file_name}")
        photo_entries.append((arcname, p))

    fname = f"ScreenPlant_All_Photos_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"

    def entries():
        for arcname, path in photo_entries:
            if os.path.exists(path):
                yield arcname, file_bytes_iter(path)
            elif _USE_R2:
                yield arcname, r2_bytes_iter(r2_key_from_path(path))

    return stream_zip_response(entries(), fname)

# ── Graceful SIGTERM shutdown (MF3) ────────────────────────────────────────────
import signal as _signal

def _handle_shutdown(signum, frame):
    logger.info("Received SIGTERM — shutting down gracefully")
    generation_executor.shutdown(wait=True, cancel_futures=False)
    global _pg_pool
    if _pg_pool:
        try:
            _pg_pool.close()
        except Exception:
            pass
    raise SystemExit(0)

try:
    _signal.signal(_signal.SIGTERM, _handle_shutdown)
except (OSError, ValueError):
    pass  # Not applicable in some contexts (e.g. subprocess workers)


# ── Automated daily backup (MF5) ──────────────────────────────────────────────
_BACKUP_INTERVAL_SECONDS = int(os.environ.get('SCREENPLANT_BACKUP_INTERVAL_HOURS', '24')) * 3600
_BACKUP_KEEP_COUNT = int(os.environ.get('SCREENPLANT_BACKUP_KEEP', '7'))

def _prune_old_auto_backups():
    try:
        auto_files = sorted(
            [f for f in os.listdir(BACKUPS_DIR) if f.startswith('screenplant_auto_') and f.endswith('.json')],
            key=lambda f: os.path.getmtime(os.path.join(BACKUPS_DIR, f)),
            reverse=True
        )
        for old in auto_files[_BACKUP_KEEP_COUNT:]:
            try:
                os.remove(os.path.join(BACKUPS_DIR, old))
                logger.info("Auto-backup: pruned old backup %s", old)
            except OSError:
                pass
    except Exception as e:
        logger.warning("Auto-backup: prune failed: %s", e)

def _auto_backup():
    # Stagger startup by 60s so this doesn't race with trash-cleanup at boot
    time.sleep(60)
    while True:
        # Retry up to 3 times with 15s gaps on pool timeout — handles cold starts
        for attempt in range(3):
            try:
                name, _path = create_backup('auto')
                logger.info("Auto-backup created: %s", name)
                _prune_old_auto_backups()
                break
            except Exception as e:
                if attempt < 2 and 'PoolTimeout' in type(e).__name__:
                    logger.warning("Auto-backup pool timeout (attempt %d/3), retrying in 15s", attempt + 1)
                    time.sleep(15)
                else:
                    logger.exception("Auto-backup failed: %s", e)
                    break
        time.sleep(_BACKUP_INTERVAL_SECONDS)

def _start_auto_backup_thread():
    t = threading.Thread(target=_auto_backup, daemon=True, name='auto-backup')
    t.start()
    logger.info("Auto-backup thread started — interval %dh, keep last %d",
                _BACKUP_INTERVAL_SECONDS // 3600, _BACKUP_KEEP_COUNT)

_start_auto_backup_thread()


# ── 7-day auto-trash cleanup ───────────────────────────────────────────────────
TRASH_TTL_DAYS = int(os.environ.get('SCREENPLANT_TRASH_TTL_DAYS', '7'))
_CLEANUP_INTERVAL_SECONDS = 24 * 60 * 60  # run once per day

def _auto_purge_old_trash():
    """
    Background thread: runs daily.
    - Deletes photos from R2 (or local disk) for trash items older than TRASH_TTL_DAYS.
    - DB trash records are kept forever (free storage, useful for audit history).
    - Marks items with 'photos_purged': True ONLY after confirmed success.
    """
    # Stagger startup by 120s so backup thread runs first and pool is warm
    time.sleep(120)
    while True:
        try:
            cutoff = (datetime.now() - timedelta(days=TRASH_TTL_DAYS)).isoformat()
            to_clean = [
                x for x in db_list_deleted_items()
                if x.get('deleted_at', '') < cutoff
                and not x.get('photos_purged')
            ]
            if to_clean:
                logger.info("Auto-trash: cleaning photos for %d item(s) older than %d days",
                            len(to_clean), TRASH_TTL_DAYS)
                for item in to_clean:
                    purge_deleted_item_files(item)
                    if item.get('photos_purged'):   # only update if purge actually set the flag
                        db_update_deleted_item(item)
                    else:
                        logger.warning("Auto-trash: could not purge photos for %s, will retry next cycle", item.get('id'))
                logger.info("Auto-trash: cycle complete")
        except Exception as e:
            logger.exception("Auto-trash cleanup failed: %s", e)
        time.sleep(_CLEANUP_INTERVAL_SECONDS)

def _start_auto_trash_thread():
    t = threading.Thread(target=_auto_purge_old_trash, daemon=True, name='auto-trash-cleanup')
    t.start()
    logger.info("Auto-trash cleanup started — items purged after %d days", TRASH_TTL_DAYS)

def _start_rate_bucket_pruner():
    """
    Background thread: prune stale entries from rate_bucket and login_failures
    every hour to prevent unbounded memory growth on long-running instances.
    rate_bucket entries are already self-expiring per-request, but old IP keys
    accumulate indefinitely if never touched again.
    """
    def _prune():
        while True:
            try:
                time.sleep(3600)  # run every hour
                cutoff = time.time() - 3600  # remove buckets idle for >1 hour
                keys_to_del = [k for k, v in list(rate_bucket.items()) if all(t < cutoff for t in v)]
                for k in keys_to_del:
                    rate_bucket.pop(k, None)
                # Prune expired login lockouts
                expired_ips = [
                    ip for ip, entry in list(login_failures.items())
                    if entry.get('locked_until', 0) < time.time() and entry.get('count', 0) < _LOGIN_MAX_FAILURES
                ]
                for ip in expired_ips:
                    login_failures.pop(ip, None)
                if keys_to_del or expired_ips:
                    logger.debug("Rate-bucket pruner: removed %d stale keys, %d expired lockouts",
                                 len(keys_to_del), len(expired_ips))
            except Exception:
                pass
    t = threading.Thread(target=_prune, daemon=True, name='rate-bucket-pruner')
    t.start()

# Start the cleanup thread when the app loads (works on Railway, Render, gunicorn, etc.)
_start_auto_trash_thread()
_start_rate_bucket_pruner()


if __name__ == '__main__':
    print("\n" + "="*50)
    print("  ScreenPlant Pro — Backend Running")
    print("  URL: http://localhost:5000")
    print("  Admin: http://localhost:5000")
    print("="*50 + "\n")
    app.run(debug=os.environ.get('FLASK_DEBUG') == '1', port=parse_int(os.environ.get('PORT'), 5000, 1, 65535), use_reloader=False)
